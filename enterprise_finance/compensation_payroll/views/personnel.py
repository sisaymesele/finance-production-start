from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator
from django.urls import reverse
from django.contrib import messages
from django.db.models import Sum, Count, F, Case, When, Value, IntegerField
from django.db.models import Q
from datetime import datetime
from django.contrib.auth.decorators import login_required
from compensation_payroll.models import PersonnelList
from compensation_payroll.forms import PersonnelListForm
from django.http import HttpResponse
import plotly.graph_objs as go
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment


@login_required
def personnel_list(request):
    # Filter personnels by logged-in user first
    personnels_queryset = PersonnelList.objects.filter(organization_name=request.user.organization_name)

    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        personnels_queryset = personnels_queryset.filter(
            Q(personnel_id__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(father_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )

    # Employment order logic
    employment_order = Case(
        When(employment_type="full_time", then=Value(1)),
        When(employment_type="part_time", then=Value(2)),
        When(employment_type="contract", then=Value(3)),
        When(employment_type="freelancers", then=Value(4)),
        When(employment_type="remote", then=Value(5)),
        When(employment_type="daily_worker", then=Value(6)),
        When(employment_type="interns", then=Value(7)),
        When(employment_type="trainee", then=Value(8)),
        When(employment_type="other", then=Value(9)),
        output_field=IntegerField(),
    )

    # Apply ordering after filtering and searching
    personnels_queryset = personnels_queryset.order_by(employment_order, "registration_date")

    # Pagination
    paginator = Paginator(personnels_queryset, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Group by 'employment_type' and count occurrences
    count_by_employment = personnels_queryset.values('employment_type') \
        .annotate(employment_count=Count('employment_type'))
    count_total_personnels = personnels_queryset.count()

    # Pass data to the template
    context = {
        'page_obj': page_obj,
        'count_by_employment': count_by_employment,
        'count_total_personnels': count_total_personnels,
    }

    return render(request, 'personnel_list/list.html', context)


@login_required
def create_personnel(request):
    next_url = request.GET.get('next')

    if request.method == 'POST':
        form = PersonnelListForm(request.POST)
        if form.is_valid():
            personnel = form.save(commit=False)
            personnel.organization_name = request.user.organization_name
            personnel.save()
            messages.success(request, "Personnel created successfully!")

            if next_url:
                return redirect(next_url)
            return redirect('personnel_list')
    else:
        form = PersonnelListForm()

    context = {
        'form': form,
        'form_title': 'Create Personnel',
        'form_description': 'Each Information Has Financial Impact',
        'submit_button_text': 'Create Personnel',
        'back_url': request.GET.get('next', reverse('personnel_list')),
    }
    return render(request, 'personnel_list/form.html', context)


@login_required
def update_personnel(request, pk):
    personnel = get_object_or_404(PersonnelList, pk=pk, organization_name=request.user.organization_name)

    if request.method == 'POST':
        form = PersonnelListForm(request.POST, instance=personnel)
        if form.is_valid():
            form.save()
            messages.success(request, "Personnel updated successfully!")
            return redirect('personnel_list')
    else:
        form = PersonnelListForm(instance=personnel)

    context = {
        'form': form,
        'form_title': 'Update Personnel',
        'form_description': 'Each Information Has Financial Impact',
        'submit_button_text': 'Update Personnel',
        'back_url': request.GET.get('next', reverse('personnel_list')),

        # ✅ Add these two lines:
        'edit_mode': True,
        'editing_personnel': personnel,
    }
    return render(request, 'personnel_list/form.html', context)


@login_required
def delete_personnel(request, pk):
    # Fetch the personnel to delete, ensuring it belongs to the logged-in user
    personnel = get_object_or_404(PersonnelList, pk=pk, organization_name=request.user.organization_name)

    if request.method == 'POST':
        # If POST request, delete the personnel
        personnel.delete()
        messages.success(request, "Personnel deleted successfully!")
        return redirect('personnel_list')  # Redirect to personnel list after deletion

    # If GET request, render a confirmation page
    return render(request, 'personnel_list/delete_confirm.html', {'personnel': personnel})


# export personnel


@login_required
def export_personnels_to_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Personnels"

    excluded_fields = {'slug'}
    field_names = [
        field.name for field in PersonnelList._meta.get_fields()
        if not field.is_relation and field.name not in excluded_fields
    ]

    # Two-line headers: Replace underscores and split long names into 2 lines
    headers = [field.replace('_', ' ').title().replace(' ', '\n') for field in field_names]
    ws.append(headers)

    # Header styles
    header_fill = PatternFill(start_color="FF0070C0", end_color="FF0070C0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Fetch data from DB
    personnels = PersonnelList.objects.filter(
        organization_name=request.user.organization_name
    ).values_list(*field_names)

    for personnel in personnels:
        row = []
        for value in personnel:
            if isinstance(value, datetime):
                row.append(value.replace(tzinfo=None))
            elif isinstance(value, str) and '_' in value:
                row.append(value.replace('_', ' ').title())  # Example: 'government_official' → 'Government Official'
            else:
                row.append(value)
        ws.append(row)

    # Auto-adjust column widths
    for i, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = max(max_length + 2, 10)

    # Prepare response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename=personnels.xlsx'
    wb.save(response)
    return response





@login_required
def personnel_graph_view(request):
    # Count personnels by employment_type, gender, and section for the logged-in user
    employment_counts = PersonnelList.objects.filter(organization_name=request.user.organization_name).values(
        'employment_type') \
        .annotate(employment_count=Count('employment_type'))

    gender_counts = PersonnelList.objects.filter(organization_name=request.user.organization_name).values('gender') \
        .annotate(gender_count=Count('gender'))

    section_counts = PersonnelList.objects.filter(organization_name=request.user.organization_name).values('section') \
        .annotate(section_count=Count('section'))

    # If no data is available, hide the graphs
    if not employment_counts and not gender_counts and not section_counts:
        return render(request, 'personnel_list/graph.html', {
            'plot_html_bar': None,
            'plot_html_pie': None,
            'plot_html_gender': None,
            'plot_html_section': None,
        })

    # Get employment type and gender labels and values
    EMPLOYMENT_TYPE_CHOICES = dict(PersonnelList.EMPLOYMENT_TYPE_CHOICES)
    GENDER_CHOICES = dict(PersonnelList.GENDER_CHOICES)

    employment_labels = [EMPLOYMENT_TYPE_CHOICES.get(entry['employment_type'], entry['employment_type']) for entry in
                         employment_counts]
    employment_values = [entry['employment_count'] for entry in employment_counts]

    gender_labels = [GENDER_CHOICES.get(entry['gender'], entry['gender']) for entry in gender_counts]
    gender_values = [entry['gender_count'] for entry in gender_counts]

    section_labels = [entry['section'] if entry['section'] else 'Unknown' for entry in section_counts]
    section_values = [entry['section_count'] for entry in section_counts]

    # Predefined color palette for all charts
    color_palette = ["ForestGreen", "RoyalBlue", "Magenta", "Indigo", "Cyan", "Orange", "Gold", "mediumVioletRed",
                     "Crimson", ]

    # Create Bar Chart for Employment Type
    fig_bar = go.Figure(data=[
        go.Bar(x=employment_labels, y=employment_values, marker=dict(color=color_palette[:len(employment_labels)]))])
    fig_bar.update_layout(
        title='Personnels by Employment Type',
        xaxis_title='Employment Type',
        yaxis_title='Personnel Count',
        template='plotly_white',
        barmode='group',
        xaxis=dict(tickangle=45),
        yaxis=dict(showgrid=True, zeroline=True, showline=True)
    )

    # Create Pie Chart for Employment Type
    fig_pie = go.Figure(data=[go.Pie(labels=employment_labels, values=employment_values,
                                     marker=dict(colors=color_palette[:len(employment_labels)]), hole=0.3)])
    fig_pie.update_layout(title='Personnel Distribution by Employment Type')

    # Create Bar Chart for Personnels by Gender
    fig_gender = go.Figure(
        data=[go.Bar(x=gender_labels, y=gender_values, marker=dict(color=color_palette[:len(gender_labels)]))])
    fig_gender.update_layout(
        title='Personnels by Gender',
        xaxis_title='Gender',
        yaxis_title='Personnel Count',
        template='plotly_white',
        barmode='group',
        xaxis=dict(tickangle=45),
        yaxis=dict(showgrid=True, zeroline=True, showline=True)
    )

    # Create Bar Chart for Personnels by Section
    fig_section = go.Figure(
        data=[go.Bar(x=section_labels, y=section_values, marker=dict(color=color_palette[:len(section_labels)]))])
    fig_section.update_layout(
        title='Personnels by Section',
        xaxis_title='Section',
        yaxis_title='Personnel Count',
        template='plotly_white',
        barmode='group',
        xaxis=dict(tickangle=45),
        yaxis=dict(showgrid=True, zeroline=True, showline=True)
    )

    # Convert figures to HTML
    plot_html_bar = fig_bar.to_html(full_html=False)
    plot_html_pie = fig_pie.to_html(full_html=False)
    plot_html_gender = fig_gender.to_html(full_html=False)
    plot_html_section = fig_section.to_html(full_html=False)

    return render(request, 'personnel_list/graph.html', {
        'plot_html_bar': plot_html_bar,
        'plot_html_pie': plot_html_pie,
        'plot_html_gender': plot_html_gender,
        'plot_html_section': plot_html_section,  # ← Add this
    })
