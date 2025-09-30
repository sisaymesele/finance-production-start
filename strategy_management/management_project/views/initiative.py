from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from management_project.models import Initiative
from management_project.forms import InitiativeForm
from django.db.models import Q, Count
from django.core.paginator import Paginator
from django.http import HttpResponse

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

import plotly.graph_objects as go
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Count

# -------------------- INITIATIVE LIST --------------------
@login_required
def initiative_list(request):
    query = request.GET.get('search', '').strip()
    selected_focus = request.GET.get('initiative_focus_area', '').strip()
    page_number = request.GET.get('page', 1)

    initiatives = Initiative.objects.filter(
        organization_name=request.user.organization_name
    )

    # Filter by focus area
    if selected_focus:
        initiatives = initiatives.filter(initiative_focus_area=selected_focus)

    # Search filter across multiple fields
    if query:
        initiatives = initiatives.filter(
            Q(initiative_focus_area__icontains=query) |
            Q(initiative_dimension__icontains=query) |
            Q(initiative_name__icontains=query) |
            Q(description__icontains=query)
        )

    initiatives = initiatives.order_by('priority', 'risk_level', '-id')

    # Get distinct focus areas for dropdown
    focus_areas = Initiative.objects.filter(
        organization_name=request.user.organization_name
    ).values_list('initiative_focus_area', flat=True).distinct()

    paginator = Paginator(initiatives, 10)
    page_obj = paginator.get_page(page_number)

    return render(request, 'initiative/list.html', {
        'initiatives': page_obj,
        'page_obj': page_obj,
        'search_query': query,
        'focus_areas': focus_areas,
        'selected_focus': selected_focus,
    })


# -------------------- CREATE INITIATIVE --------------------
@login_required
def create_initiative(request):
    if request.method == 'POST':
        form = InitiativeForm(request.POST)
        if 'save' in request.POST and form.is_valid():
            initiative = form.save(commit=False)
            initiative.organization_name = request.user.organization_name
            initiative.save()
            messages.success(request, "Initiative created successfully!")
            return redirect('initiative_list')
    else:
        form = InitiativeForm()
    return render(request, 'initiative/form.html', {'form': form})


# -------------------- UPDATE INITIATIVE --------------------
@login_required
def update_initiative(request, pk):
    initiative = get_object_or_404(
        Initiative, pk=pk, organization_name=request.user.organization_name
    )

    if request.method == 'POST':
        form = InitiativeForm(request.POST, instance=initiative)
        if 'save' in request.POST and form.is_valid():
            form.save()
            messages.success(request, "Initiative updated successfully!")
            return redirect('initiative_list')
        else:
            messages.error(request, "Error updating initiative. Check the form.")
    else:
        form = InitiativeForm(instance=initiative)

    return render(request, 'initiative/form.html', {'form': form})


# -------------------- DELETE INITIATIVE --------------------
@login_required
def delete_initiative(request, pk):
    initiative = get_object_or_404(
        Initiative, pk=pk, organization_name=request.user.organization_name
    )

    if request.method == 'POST':
        initiative.delete()
        messages.success(request, "Initiative deleted successfully!")
        return redirect('initiative_list')

    return render(request, 'initiative/delete_confirm.html', {'entry': initiative})


# -------------------- EXPORT INITIATIVE TO EXCEL --------------------
@login_required
def export_initiative_to_excel(request):
    query = request.GET.get('search', '').strip()
    selected_focus = request.GET.get('initiative_focus_area', '').strip()

    initiatives = Initiative.objects.filter(
        organization_name=request.user.organization_name
    )

    if selected_focus:
        initiatives = initiatives.filter(initiative_focus_area=selected_focus)
    if query:
        initiatives = initiatives.filter(
            Q(initiative_focus_area__icontains=query) |
            Q(initiative_dimension__icontains=query) |
            Q(initiative_name__icontains=query) |
            Q(description__icontains=query)
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Initiatives"

    # === Styles ===
    title_font = Font(size=14, bold=True, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_fill = PatternFill(start_color="4BACC6", end_color="4BACC6", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    # === Title Row ===
    total_columns = 12
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = ws.cell(row=1, column=1, value="Initiative Report")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = title_fill

    # === Headers ===
    headers = [
        "Focus Area", "Dimension", "Name", "Description",
        "Priority", "Impact", "Likelihood", "Risk Level", "Status",
        "Organization", "Created At", "Updated At"
    ]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # === Data Rows ===
    for row_num, ini in enumerate(initiatives, start=3):
        row_data = [
            ini.initiative_focus_area,
            ini.initiative_dimension,
            ini.initiative_name,
            ini.description or "",
            ini.get_priority_display(),
            ini.get_impact_display(),
            ini.get_likelihood_display(),
            ini.get_risk_level_display(),
            ini.get_status_display(),
            str(ini.organization_name),  # readable org name
            ini.created_at.strftime("%Y-%m-%d %H:%M"),
            ini.updated_at.strftime("%Y-%m-%d %H:%M"),
        ]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = border

    # === Auto column width ===
    for i, col in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(i)].width = max_length + 5

    # === Response ===
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Initiative_Report.xlsx'
    wb.save(response)
    return response



# -------------------- INITIATIVE CHART --------------------


@login_required
def initiative_chart(request):
    qs = Initiative.objects.filter(
        organization_name=request.user.organization_name
    )

    # ------------------ PRIORITY DISTRIBUTION ------------------
    priority_counts = qs.values('priority').annotate(count=Count('id'))
    priority_labels = [p['priority'] for p in priority_counts]
    priority_values = [p['count'] for p in priority_counts]

    fig_priority = go.Figure(data=[go.Bar(
        x=priority_labels,
        y=priority_values,
        text=priority_values,
        textposition='auto',
        marker_color=['#d62728', '#ff7f0e', '#ffbb78', '#98df8a', '#2ca02c']
    )])
    fig_priority.update_layout(
        title='Initiatives Distribution by Priority',
        xaxis_title='Priority',
        yaxis_title='Count',
        template='plotly_white',
        height=400
    )

    # ------------------ IMPACT DISTRIBUTION ------------------
    impact_counts = qs.values('impact').annotate(count=Count('id'))
    impact_labels = [i['impact'] for i in impact_counts]
    impact_values = [i['count'] for i in impact_counts]

    fig_impact = go.Figure(data=[go.Bar(
        x=impact_labels,
        y=impact_values,
        text=impact_values,
        textposition='auto',
        marker_color=['#98df8a', '#2ca02c', '#ffbb78', '#ff7f0e', '#d62728']
    )])
    fig_impact.update_layout(
        title='Initiatives Distribution by Impact',
        xaxis_title='Impact',
        yaxis_title='Count',
        template='plotly_white',
        height=400
    )

    # ------------------ LIKELIHOOD DISTRIBUTION ------------------
    likelihood_counts = qs.values('likelihood').annotate(count=Count('id'))
    likelihood_labels = [l['likelihood'] if l['likelihood'] else 'Unknown' for l in likelihood_counts]
    likelihood_values = [l['count'] for l in likelihood_counts]

    fig_likelihood = go.Figure(data=[go.Bar(
        x=likelihood_labels,
        y=likelihood_values,
        text=likelihood_values,
        textposition='auto',
        marker_color=['#17becf', '#1f77b4', '#ff7f0e', '#2ca02c', '#d62728', '#7f7f7f']
    )])
    fig_likelihood.update_layout(
        title='Initiatives Distribution by Likelihood',
        xaxis_title='Likelihood',
        yaxis_title='Count',
        template='plotly_white',
        height=400
    )

    # ------------------ RISK LEVEL DISTRIBUTION ------------------
    risk_counts = qs.values('risk_level').annotate(count=Count('id'))
    risk_labels = [r['risk_level'] for r in risk_counts]
    risk_values = [r['count'] for r in risk_counts]

    fig_risk = go.Figure(data=[go.Bar(
        x=risk_labels,
        y=risk_values,
        text=risk_values,
        textposition='auto',
        marker_color=['#d62728', '#ff7f0e', '#ffbb78', '#98df8a', '#2ca02c']
    )])
    fig_risk.update_layout(
        title='Initiatives Distribution by Risk Level',
        xaxis_title='Risk Level',
        yaxis_title='Count',
        template='plotly_white',
        height=400
    )

    # ------------------ STATUS DISTRIBUTION (PIE CHART) ------------------
    status_counts = qs.values('status').annotate(count=Count('id'))
    status_labels = [s['status'] for s in status_counts]
    status_values = [s['count'] for s in status_counts]

    fig_status = go.Figure(data=[go.Pie(
        labels=status_labels,
        values=status_values,
        textinfo='label+percent',
        hole=0.3
    )])
    fig_status.update_layout(
        title='Initiatives Distribution by Status',
        template='plotly_white',
        height=400
    )

    # ------------------ DIMENSION ANALYSIS ------------------
    dimension_counts = qs.values('initiative_dimension').annotate(count=Count('id')).order_by('-count')
    dimension_labels = [d['initiative_dimension'] for d in dimension_counts]
    dimension_values = [d['count'] for d in dimension_counts]

    fig_dimension = go.Figure(data=[go.Bar(
        x=dimension_labels,
        y=dimension_values,
        text=dimension_values,
        textposition='auto',
        marker_color=['#636efa', '#ef553b', '#00cc96', '#ab63fa', '#ffa15a']
    )])
    fig_dimension.update_layout(
        title='Initiatives Count per Dimension',
        xaxis_title='Dimension',
        yaxis_title='Count',
        template='plotly_white',
        height=400
    )

    # ------------------ SUMMARY DATA ------------------
    summary_data = {
        'total_initiatives': qs.count(),
        'high_priority_count': qs.filter(priority__in=['High', 'Very High']).count(),
        'impact_high_count': qs.filter(impact__in=['High', 'Very High']).count(),
        'likelihood_high_count': qs.filter(likelihood__in=['High', 'Very High']).count(),
        'risk_critical_count': qs.filter(risk_level='Critical').count(),
        'status_inprogress_count': qs.filter(status='In Progress').count(),
    }

    return render(request, 'initiative/chart.html', {
        'plot_html_priority': fig_priority.to_html(full_html=False),
        'plot_html_impact': fig_impact.to_html(full_html=False),
        'plot_html_likelihood': fig_likelihood.to_html(full_html=False),
        'plot_html_risk': fig_risk.to_html(full_html=False),
        'plot_html_status': fig_status.to_html(full_html=False),
        'plot_html_dimension': fig_dimension.to_html(full_html=False),
        'summary_data': summary_data,
    })

