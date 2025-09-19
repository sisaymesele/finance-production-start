from django.shortcuts import render
from django.http import HttpResponse
from decimal import Decimal
from django.contrib.auth.decorators import login_required
from management_project.services.combined.yearly_context import get_combined_yearly_detail
from django.contrib.auth.decorators import login_required
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from datetime import datetime


@login_required
def yearly_combined_detail_view(request):
    context = get_combined_yearly_detail(request)
    return render(request, 'combined_payroll/yearly_detail.html', context)

#yearly summary
@login_required
def yearly_combined_summary_view(request):

    context = get_combined_yearly_detail(request)

    return render(request, 'combined_payroll/yearly_summary.html', context)
#



@login_required
def export_combined_yearly_detail_to_excel(request):
    context = get_combined_yearly_detail(request)
    yearly_list = context['page_obj'].object_list

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Combined Yearly Payroll"

    # ======================
    # STYLE DEFINITIONS
    # ======================
    # Fonts
    title_font = Font(bold=True, size=14, color="FFFFFF")
    section_font = Font(bold=True, size=12, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF")
    normal_font = Font(size=11)
    bold_font = Font(bold=True, size=11)
    total_font = Font(bold=True, size=11, color="FFFFFF")

    # Colors (matching HTML template)
    color_palette = {
        'primary': '4472C4',  # Blue (card header)
        'success': '70AD47',  # Green (strategic action plan)
        'warning': 'FFC000',  # Yellow (strategic reports)
        'danger': 'ED7D31',  # Orange (deduction adjustments)
        'info': '5B9BD5',  # Light blue (adjustment summary)
        'purple': '7030A0',  # Purple (severance)
        'totals': '4472C4',  # Blue (totals)
    }

    # Fills
    title_fill = PatternFill(start_color=color_palette['primary'], end_color=color_palette['primary'],
                             fill_type="solid")
    regular_fill = PatternFill(start_color=color_palette['success'], end_color=color_palette['success'],
                               fill_type="solid")
    earning_fill = PatternFill(start_color=color_palette['warning'], end_color=color_palette['warning'],
                               fill_type="solid")
    deduction_fill = PatternFill(start_color=color_palette['danger'], end_color=color_palette['danger'],
                                 fill_type="solid")
    summary_fill = PatternFill(start_color=color_palette['info'], end_color=color_palette['info'], fill_type="solid")
    severance_fill = PatternFill(start_color=color_palette['purple'], end_color=color_palette['purple'],
                                 fill_type="solid")
    totals_fill = PatternFill(start_color=color_palette['totals'], end_color=color_palette['totals'], fill_type="solid")

    # Borders
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    # Number formats
    money_format = '#,##0.00'

    # Alignment
    center_align = Alignment(horizontal="center")
    left_align = Alignment(horizontal="left")
    right_align = Alignment(horizontal="right")

    row_num = 1

    for item in yearly_list:
        # ======================
        # MONTH HEADER
        # ======================
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
        header_cell = ws.cell(row=row_num, column=1, value=f"Combined Yearly Payroll Detail for {item['year']}")
        header_cell.font = title_font
        header_cell.fill = title_fill
        header_cell.alignment = center_align
        row_num += 2

        # ======================
        # REGULAR PAYROLL SECTION
        # ======================
        if 'regular' in item:
            # Section Header
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
            header_cell = ws.cell(row=row_num, column=1, value="REGULAR PAYROLL")
            header_cell.font = Font(bold=True, size=12, color=color_palette['success'])
            row_num += 1

            # Regular Components
            if 'regular_item_by_component' in item:
                headers = ["Component", "Amount"]
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=header)
                    cell.font = header_font
                    cell.fill = regular_fill
                    cell.alignment = center_align
                    cell.border = border
                row_num += 1

                for comp, amount in item['regular_item_by_component'].items():
                    if amount:
                        ws.cell(row=row_num, column=1, value=comp).font = normal_font
                        ws.cell(row=row_num, column=1).alignment = left_align
                        ws.cell(row=row_num, column=1).border = border

                        amt_cell = ws.cell(row=row_num, column=2, value=float(amount))
                        amt_cell.number_format = money_format
                        amt_cell.alignment = right_align
                        amt_cell.border = border
                        row_num += 1
                row_num += 1

            # Regular Summary
            headers = ["Strategic Action Plan Summary Item", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = regular_fill
                cell.alignment = center_align
                cell.border = border
            row_num += 1

            regular_summary = [
                ("Taxable Gross", item['regular'].get('taxable_gross')),
                ("Non-Taxable Gross", item['regular'].get('non_taxable_gross')),
                ("Gross Pay", item['regular'].get('gross'), 'primary'),
                ("Pensionable Amount", item['regular'].get('pensionable')),
                ("Employee Pension", item['regular'].get('employee_pension')),
                ("Employer Pension", item['regular'].get('employer_pension')),
                ("Total Pension", item['regular'].get('total_pension')),
                ("Employment Income Tax", item['regular'].get('employment_income_tax')),
                ("Total Regular Deduction", item['regular'].get('total_regular_deduction')),
                ("Net Pay", item['regular'].get('net_pay'), 'success'),
                ("Expense", item['regular'].get('expense')),
            ]

            for label, value, *style in regular_summary:
                if value:
                    cell = ws.cell(row=row_num, column=1, value=label)
                    cell.font = bold_font
                    cell.alignment = left_align
                    cell.border = border

                    amt_cell = ws.cell(row=row_num, column=2, value=float(value))
                    amt_cell.number_format = money_format
                    amt_cell.alignment = right_align
                    amt_cell.border = border

                    if style:
                        amt_cell.font = total_font
                        amt_cell.fill = PatternFill(start_color=color_palette[style[0]],
                                                    end_color=color_palette[style[0]], fill_type="solid")

                    row_num += 1
            row_num += 2

        # ======================
        # EARNING ADJUSTMENTS SECTION
        # ======================
        if item.get('show_earning', False):
            # Section Header
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
            header_cell = ws.cell(row=row_num, column=1, value="EARNING ADJUSTMENTS")
            header_cell.font = Font(bold=True, size=12, color=color_palette['warning'])
            row_num += 1

            # Earning Components
            if 'earning_adj_by_component' in item:
                headers = ["Component", "Earning Amount", "Taxable", "Non-Taxable",
                           "Employee Pension", "Employer Pension", "Total Pension"]
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=header)
                    cell.font = header_font
                    cell.fill = earning_fill
                    cell.alignment = center_align
                    cell.border = border
                row_num += 1

                for comp, vals in item['earning_adj_by_component'].items():
                    if vals.get('earning_amount'):
                        ws.cell(row=row_num, column=1, value=comp).font = normal_font
                        ws.cell(row=row_num, column=1).alignment = left_align
                        ws.cell(row=row_num, column=1).border = border

                        columns = [
                            ('earning_amount', 2),
                            ('taxable', 3),
                            ('non_taxable', 4),
                            ('employee_pension_contribution', 5),
                            ('employer_pension_contribution', 6),
                            ('total_pension', 7)
                        ]

                        for key, col in columns:
                            val = float(vals.get(key, Decimal('0.00')))
                            ws.cell(row=row_num, column=col, value=val).number_format = money_format
                            ws.cell(row=row_num, column=col).alignment = right_align
                            ws.cell(row=row_num, column=col).border = border

                        row_num += 1
                row_num += 1

            # Earning Summary
            headers = ["Strategic Report Summary Item", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = earning_fill
                cell.alignment = center_align
                cell.border = border
            row_num += 1

            earning_summary = [
                ("Taxable Gross", item['adjustment'].get('taxable_gross')),
                ("Non-Taxable Gross", item['adjustment'].get('non_taxable_gross')),
                ("Gross Adjustment", item['adjustment'].get('gross'), 'primary'),
                ("Adjusted Pensionable", item['adjustment'].get('adjusted_pensionable')),
                ("Employee Pension", item['adjustment'].get('employee_pension')),
                ("Employer Pension", item['adjustment'].get('employer_pension')),
                ("Total Pension", item['adjustment'].get('total_pension')),
                ("Employment Income Tax", item['adjustment'].get('employment_income_tax')),
                ("Strategic Report Deduction", item['adjustment'].get('strategic_report_deduction')),
                ("Expense", item['adjustment'].get('expense')),
            ]

            for label, value, *style in earning_summary:
                if value:
                    cell = ws.cell(row=row_num, column=1, value=label)
                    cell.font = bold_font
                    cell.alignment = left_align
                    cell.border = border

                    amt_cell = ws.cell(row=row_num, column=2, value=float(value))
                    amt_cell.number_format = money_format
                    amt_cell.alignment = right_align
                    amt_cell.border = border

                    if style:
                        amt_cell.font = total_font
                        amt_cell.fill = PatternFill(start_color=color_palette[style[0]],
                                                    end_color=color_palette[style[0]], fill_type="solid")

                    row_num += 1
            row_num += 2

        # ======================
        # DEDUCTION ADJUSTMENTS SECTION
        # ======================
        if item.get('show_deduction', False):
            # Section Header
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
            header_cell = ws.cell(row=row_num, column=1, value="DEDUCTION ADJUSTMENTS")
            header_cell.font = Font(bold=True, size=12, color=color_palette['danger'])
            row_num += 1

            # Deduction Components
            if 'deduction_adj_by_component' in item:
                headers = ["Component", "Amount"]
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=header)
                    cell.font = header_font
                    cell.fill = deduction_fill
                    cell.alignment = center_align
                    cell.border = border
                row_num += 1

                for comp, amount in item['deduction_adj_by_component'].items():
                    if amount:
                        ws.cell(row=row_num, column=1, value=comp).font = normal_font
                        ws.cell(row=row_num, column=1).alignment = left_align
                        ws.cell(row=row_num, column=1).border = border

                        amt_cell = ws.cell(row=row_num, column=2, value=float(amount))
                        amt_cell.number_format = money_format
                        amt_cell.alignment = right_align
                        amt_cell.border = border
                        row_num += 1
                row_num += 1

            # Deduction Summary
            headers = ["Deduction Adjustment Summary Item", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = deduction_fill
                cell.alignment = center_align
                cell.border = border
            row_num += 1

            deduction_summary = [
                ("Individual Adjustment Deduction", item['adjustment'].get('individual_adjustment_deduction')),
            ]

            for label, value in deduction_summary:
                if value:
                    cell = ws.cell(row=row_num, column=1, value=label)
                    cell.font = bold_font
                    cell.alignment = left_align
                    cell.border = border

                    amt_cell = ws.cell(row=row_num, column=2, value=float(value))
                    amt_cell.number_format = money_format
                    amt_cell.alignment = right_align
                    amt_cell.border = border
                    row_num += 1
            row_num += 1

        # ======================
        # ======================
        # ADJUSTMENT SUMMARY SECTION
        # ======================
        if (
            item.get('show_earning', False) or
            item.get('show_deduction', False)
        ) and 'adjustment' in item:
            # Section Header
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
            header_cell = ws.cell(row=row_num, column=1, value="ADJUSTMENT SUMMARY")
            header_cell.font = Font(bold=True, size=12, color=color_palette['info'])
            row_num += 1

            # Adjustment Summary
            headers = ["Total Adjustment Summary Item", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = summary_fill
                cell.alignment = center_align
                cell.border = border
            row_num += 1

            adjustment_summary = [
                ("Taxable Gross", item['adjustment'].get('taxable_gross')),
                ("Non-Taxable Gross", item['adjustment'].get('non_taxable_gross')),
                ("Gross Adjustment", item['adjustment'].get('gross'), 'primary'),
                ("Adjusted Pensionable", item['adjustment'].get('adjusted_pensionable')),
                ("Employee Pension", item['adjustment'].get('employee_pension')),
                ("Employer Pension", item['adjustment'].get('employer_pension')),
                ("Total Pension", item['adjustment'].get('total_pension')),
                ("Employment Income Tax", item['adjustment'].get('employment_income_tax')),
                ("Strategic Report Deduction", item['adjustment'].get('strategic_report_deduction')),
                ("Individual Adjustment Deduction", item['adjustment'].get('individual_adjustment_deduction')),
                ("Total Yearly Adjustment Deduction", item['adjustment'].get('total_yearly_adjustment_deduction'),
                 'primary'),
                ("Net Yearly Adjustment", item['adjustment'].get('net_yearly_adjustment'), 'success'),
                ("Expense", item['adjustment'].get('expense')),
            ]

            for label, value, *style in adjustment_summary:
                if value:
                    cell = ws.cell(row=row_num, column=1, value=label)
                    cell.font = bold_font
                    cell.alignment = left_align
                    cell.border = border

                    amt_cell = ws.cell(row=row_num, column=2, value=float(value))
                    amt_cell.number_format = money_format
                    amt_cell.alignment = right_align
                    amt_cell.border = border

                    if style:
                        amt_cell.font = total_font
                        amt_cell.fill = PatternFill(
                            start_color=color_palette[style[0]],
                            end_color=color_palette[style[0]],
                            fill_type="solid"
                        )
                    row_num += 1
            row_num += 2


        # ======================
        # SEVERANCE PAY SECTION
        # ======================
        if 'severance' in item and any(item['severance'].values()):
            # Section Header
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
            header_cell = ws.cell(row=row_num, column=1, value="SEVERANCE PAY")
            header_cell.font = Font(bold=True, size=12, color=color_palette['purple'])
            row_num += 1

            # Severance Details
            headers = ["Component", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = severance_fill
                cell.alignment = center_align
                cell.border = border
            row_num += 1

            severance_items = [
                ("Taxable Gross", item['severance'].get('taxable_gross')),
                ("Gross Pay", item['severance'].get('gross'), 'primary'),
                ("Employment Income Tax", item['severance'].get('employment_income_tax')),
                ("Total Severance Deduction", item['severance'].get('total_severance_deduction')),
                ("Net Pay", item['severance'].get('net'), 'success'),
                ("Expense", item['severance'].get('expense')),
            ]

            for label, value, *style in severance_items:
                if value:
                    cell = ws.cell(row=row_num, column=1, value=label)
                    cell.font = bold_font
                    cell.alignment = left_align
                    cell.border = border

                    amt_cell = ws.cell(row=row_num, column=2, value=float(value))
                    amt_cell.number_format = money_format
                    amt_cell.alignment = right_align
                    amt_cell.border = border

                    if style:
                        amt_cell.font = total_font
                        amt_cell.fill = PatternFill(start_color=color_palette[style[0]],
                                                    end_color=color_palette[style[0]], fill_type="solid")

                    row_num += 1
            row_num += 2

        # ======================
        # TOTAL SUMMARY SECTION
        # ======================
        if 'totals' in item:
            # Section Header
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
            header_cell = ws.cell(row=row_num, column=1, value="TOTAL SUMMARY")
            header_cell.font = Font(bold=True, size=12, color=color_palette['totals'])
            row_num += 1

            # Total Details
            headers = ["Total Summary Items", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = totals_fill
                cell.alignment = center_align
                cell.border = border
            row_num += 1

            summary_items = [
                ("Taxable Gross", item['totals'].get('taxable_gross'), 'info'),
                ("Non-Taxable Gross", item['totals'].get('non_taxable_gross'), 'info'),
                ("Total Gross", item['totals'].get('gross'), 'info'),
                ("Pensionable", item['totals'].get('pensionable'), 'warning'),
                ("Employee Pension", item['totals'].get('employee_pension'), 'warning'),
                ("Employer Pension", item['totals'].get('employer_pension'), 'warning'),
                ("Total Pension", item['totals'].get('total_pension'), 'warning'),
                ("Employment Income Tax", item['totals'].get('employment_income_tax')),
                ("Total Deduction", item['totals'].get('total_deduction')),
                ("Final Net Pay", item['totals'].get('final_net_pay'), 'success'),
                ("Expense", item['totals'].get('expense')),
            ]

            for label, value, *style in summary_items:
                if value:
                    cell = ws.cell(row=row_num, column=1, value=label)
                    cell.font = bold_font
                    cell.alignment = left_align
                    cell.border = border

                    amt_cell = ws.cell(row=row_num, column=2, value=float(value))
                    amt_cell.number_format = money_format
                    amt_cell.alignment = right_align
                    amt_cell.border = border

                    if style:
                        amt_cell.font = total_font
                        amt_cell.fill = PatternFill(start_color=color_palette[style[0]],
                                                    end_color=color_palette[style[0]], fill_type="solid")

                    row_num += 1
            row_num += 3  # Extra space between years

    # ======================
    # FINAL FORMATTING
    # ======================
    # Adjust column widths - handle merged cells properly
    column_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue  # Skip merged cells

            if cell.value:
                col_letter = openpyxl.utils.get_column_letter(cell.column)
                content_length = len(str(cell.value))
                if col_letter in column_widths:
                    if content_length > column_widths[col_letter]:
                        column_widths[col_letter] = content_length
                else:
                    column_widths[col_letter] = content_length

    for col_letter, max_len in column_widths.items():
        adjusted_width = (max_len + 2) * 1.2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"combined_yearly_payroll_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


#


#
@login_required
def export_combined_yearly_summary_to_excel(request):
    context = get_combined_yearly_detail(request)
    yearly_list = context['page_obj'].object_list

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Yearly Payroll Summary"

    # ======================
    # STYLE DEFINITIONS MATCHING HTML TEMPLATE
    # ======================
    # Fonts
    title_font = Font(bold=True, size=14, color="FFFFFF")
    section_font = Font(bold=True, size=12)
    header_font = Font(bold=True, color="FFFFFF")
    normal_font = Font(size=11)
    bold_font = Font(bold=True, size=11)
    positive_font = Font(size=11, color="000000")  # Black for amounts
    negative_font = Font(size=11, color="000000")  # Black for amounts
    neutral_font = Font(size=11, color="000000")  # Black for neutral amounts

    # Colors matching HTML template classes
    color_palette = {
        'title': '3498DB',  # Blue (card header)
        'regular': '2E86C1',  # Primary blue (section header)
        'adjustment': 'F39C12',  # Orange (section header)
        'severance': 'E74C3C',  # Red (section header)
        'totals': '2ECC71',  # Green (section header)
        'primary': 'D6EAF8',  # Light blue (table-primary)
        'success': 'D5F5E3',  # Light green (table-success)
        'info': 'D6EAF8',  # Light blue (table-info)
        'warning': 'FDEBD0',  # Light orange (table-warning)
        'danger': 'FADBD8',  # Light red (table-danger)
        'neutral': 'FFFFFF',  # White
    }

    # Fills
    title_fill = PatternFill(start_color=color_palette['title'], end_color=color_palette['title'], fill_type="solid")
    regular_fill = PatternFill(start_color=color_palette['regular'], end_color=color_palette['regular'],
                               fill_type="solid")
    adjustment_fill = PatternFill(start_color=color_palette['adjustment'], end_color=color_palette['adjustment'],
                                  fill_type="solid")
    severance_fill = PatternFill(start_color=color_palette['severance'], end_color=color_palette['severance'],
                                 fill_type="solid")
    totals_fill = PatternFill(start_color=color_palette['totals'], end_color=color_palette['totals'], fill_type="solid")
    primary_fill = PatternFill(start_color=color_palette['primary'], end_color=color_palette['primary'],
                               fill_type="solid")
    success_fill = PatternFill(start_color=color_palette['success'], end_color=color_palette['success'],
                               fill_type="solid")
    info_fill = PatternFill(start_color=color_palette['info'], end_color=color_palette['info'], fill_type="solid")
    warning_fill = PatternFill(start_color=color_palette['warning'], end_color=color_palette['warning'],
                               fill_type="solid")
    danger_fill = PatternFill(start_color=color_palette['danger'], end_color=color_palette['danger'], fill_type="solid")
    neutral_fill = PatternFill(start_color=color_palette['neutral'], end_color=color_palette['neutral'],
                               fill_type="solid")

    # Borders
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    # Number formats
    money_format = '#,##0.00'

    # Alignment
    center_align = Alignment(horizontal="center")
    left_align = Alignment(horizontal="left")
    right_align = Alignment(horizontal="right")

    row_num = 1

    for item in yearly_list:
        # ======================
        # MONTH HEADER - matches card header
        # ======================
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
        header_cell = ws.cell(row=row_num, column=1, value=f"Yearly Payroll Summary For {item['year']}")
        header_cell.font = title_font
        header_cell.fill = title_fill
        header_cell.alignment = center_align
        row_num += 2

        # ======================
        # REGULAR PAYROLL SUMMARY - matches table-primary/success
        # ======================
        if 'regular' in item and any(val != Decimal('0.00') for val in item['regular'].values()):
            # Section Header
            ws.cell(row=row_num, column=1, value="REGULAR PAYROLL SUMMARY").font = Font(bold=True, size=12,
                                                                                        color=color_palette['regular'])
            row_num += 1

            headers = ["Item", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = regular_fill
                cell.alignment = center_align
                cell.border = border
            row_num += 1

            regular_summary = [
                ("taxable_gross", item['regular'].get('taxable_gross'), 'neutral'),
                ("non_taxable_gross", item['regular'].get('non_taxable_gross'), 'neutral'),
                ("gross", item['regular'].get('gross'), 'primary'),
                ("pensionable", item['regular'].get('pensionable'), 'neutral'),
                ("employee_pension", item['regular'].get('employee_pension'), 'neutral'),
                ("employer_pension", item['regular'].get('employer_pension'), 'neutral'),
                ("total_pension", item['regular'].get('total_pension'), 'neutral'),
                ("employment_income_tax", item['regular'].get('employment_income_tax'), 'neutral'),
                ("total_regular_deduction", item['regular'].get('total_regular_deduction'), 'neutral'),
                ("net_pay", item['regular'].get('net_pay'), 'success'),
                ("expense", item['regular'].get('expense'), 'neutral'),
            ]

            for field_name, value, color_type in regular_summary:
                if value != Decimal('0.00'):
                    label = field_name.replace('_', ' ').title()
                    ws.cell(row=row_num, column=1, value=label).font = bold_font
                    ws.cell(row=row_num, column=1).alignment = left_align
                    ws.cell(row=row_num, column=1).border = border

                    amt_cell = ws.cell(row=row_num, column=2, value=float(value))
                    amt_cell.number_format = money_format
                    amt_cell.alignment = right_align
                    amt_cell.border = border

                    # Apply color coding
                    if color_type == 'primary':
                        amt_cell.fill = primary_fill
                        amt_cell.font = bold_font
                    elif color_type == 'success':
                        amt_cell.fill = success_fill
                        amt_cell.font = bold_font
                    else:
                        amt_cell.fill = neutral_fill
                        amt_cell.font = normal_font

                    row_num += 1
            row_num += 2

        # ======================
        # ADJUSTMENT SUMMARY - matches table-primary/success
        # ======================
        if 'adjustment' in item:
            adj_summary = [
                ("taxable_gross", item['adjustment'].get('taxable_gross'), 'neutral'),
                ("non_taxable_gross", item['adjustment'].get('non_taxable_gross'), 'neutral'),
                ("gross", item['adjustment'].get('gross'), 'primary'),
                ("adjusted_pensionable", item['adjustment'].get('adjusted_pensionable'), 'neutral'),
                ("employee_pension", item['adjustment'].get('employee_pension'), 'neutral'),
                ("employer_pension", item['adjustment'].get('employer_pension'), 'neutral'),
                ("total_pension", item['adjustment'].get('total_pension'), 'neutral'),
                ("employment_income_tax", item['adjustment'].get('employment_income_tax'), 'neutral'),
                ("strategic_report_deduction", item['adjustment'].get('strategic_report_deduction'), 'neutral'),
                ("individual_adjustment_deduction", item['adjustment'].get('individual_adjustment_deduction'),
                 'neutral'),
                ("total_yearly_adjustment_deduction", item['adjustment'].get('total_yearly_adjustment_deduction'),
                 'primary'),
                ("net_yearly_adjustment", item['adjustment'].get('net_yearly_adjustment'), 'success'),
                ("expense", item['adjustment'].get('expense'), 'neutral'),
            ]

            # Check if there are any non-zero values in the adjustment items
            has_non_zero_adjustments = any(value != Decimal('0.00') for _, value, _ in adj_summary)

            if has_non_zero_adjustments:
                # Section Header
                ws.cell(row=row_num, column=1, value="ADJUSTMENT SUMMARY").font = Font(bold=True, size=12,
                                                                                       color=color_palette[
                                                                                           'adjustment'])
                row_num += 1

                headers = ["Item", "Amount"]
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=header)
                    cell.font = header_font
                    cell.fill = adjustment_fill
                    cell.alignment = center_align
                    cell.border = border
                row_num += 1

                for field_name, value, color_type in adj_summary:
                    if value != Decimal('0.00'):
                        label = field_name.replace('_', ' ').title()
                        ws.cell(row=row_num, column=1, value=label).font = bold_font
                        ws.cell(row=row_num, column=1).alignment = left_align
                        ws.cell(row=row_num, column=1).border = border

                        amt_cell = ws.cell(row=row_num, column=2, value=float(value))
                        amt_cell.number_format = money_format
                        amt_cell.alignment = right_align
                        amt_cell.border = border

                        # Apply color coding
                        if color_type == 'primary':
                            amt_cell.fill = primary_fill
                            amt_cell.font = bold_font
                        elif color_type == 'success':
                            amt_cell.fill = success_fill
                            amt_cell.font = bold_font
                        else:
                            amt_cell.fill = neutral_fill
                            amt_cell.font = normal_font

                        row_num += 1
                row_num += 2

        # ======================
        # SEVERANCE SUMMARY - matches table-primary/success/danger
        # ======================
        if 'severance' in item and any(val != Decimal('0.00') for val in item['severance'].values()):
            # Section Header
            ws.cell(row=row_num, column=1, value="SEVERANCE SUMMARY").font = Font(bold=True, size=12,
                                                                                  color=color_palette['severance'])
            row_num += 1

            headers = ["Item", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = severance_fill
                cell.alignment = center_align
                cell.border = border
            row_num += 1

            severance_items = [
                ("taxable_gross", item['severance'].get('taxable_gross'), 'neutral'),
                ("gross", item['severance'].get('gross'), 'primary'),
                ("employment_income_tax", item['severance'].get('employment_income_tax'), 'neutral'),
                ("total_severance_deduction", item['severance'].get('total_severance_deduction'), 'neutral'),
                ("net", item['severance'].get('net'), 'success'),
                ("expense", item['severance'].get('expense'), 'neutral'),
            ]

            for field_name, value, color_type in severance_items:
                if value != Decimal('0.00'):
                    label = field_name.replace('_', ' ').title()
                    ws.cell(row=row_num, column=1, value=label).font = bold_font
                    ws.cell(row=row_num, column=1).alignment = left_align
                    ws.cell(row=row_num, column=1).border = border

                    amt_cell = ws.cell(row=row_num, column=2, value=float(value))
                    amt_cell.number_format = money_format
                    amt_cell.alignment = right_align
                    amt_cell.border = border

                    # Apply color coding
                    if color_type == 'primary':
                        amt_cell.fill = primary_fill
                        amt_cell.font = bold_font
                    elif color_type == 'success':
                        amt_cell.fill = success_fill
                        amt_cell.font = bold_font
                    else:
                        amt_cell.fill = neutral_fill
                        amt_cell.font = normal_font

                    row_num += 1
            row_num += 2

        # ======================
        # TOTAL SUMMARY - matches table-info/warning/success
        # ======================
        if 'totals' in item and any(val != Decimal('0.00') for val in item['totals'].values()):
            # Section Header
            ws.cell(row=row_num, column=1, value="TOTAL SUMMARY").font = Font(bold=True, size=12,
                                                                              color=color_palette['totals'])
            row_num += 1

            headers = ["Item", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = totals_fill
                cell.alignment = center_align
                cell.border = border
            row_num += 1

            summary_items = [
                ("taxable_gross", item['totals'].get('taxable_gross'), 'info'),
                ("non_taxable_gross", item['totals'].get('non_taxable_gross'), 'info'),
                ("gross", item['totals'].get('gross'), 'info'),
                ("pensionable", item['totals'].get('pensionable'), 'warning'),
                ("employee_pension", item['totals'].get('employee_pension'), 'warning'),
                ("employer_pension", item['totals'].get('employer_pension'), 'warning'),
                ("total_pension", item['totals'].get('total_pension'), 'warning'),
                ("employment_income_tax", item['totals'].get('employment_income_tax'), 'neutral'),
                ("total_deduction", item['totals'].get('total_deduction'), 'neutral'),
                ("final_net_pay", item['totals'].get('final_net_pay'), 'success'),
                ("expense", item['totals'].get('expense'), 'neutral'),
            ]

            for field_name, value, color_type in summary_items:
                if value != Decimal('0.00'):
                    label = field_name.replace('_', ' ').title()
                    ws.cell(row=row_num, column=1, value=label).font = bold_font
                    ws.cell(row=row_num, column=1).alignment = left_align
                    ws.cell(row=row_num, column=1).border = border

                    amt_cell = ws.cell(row=row_num, column=2, value=float(value))
                    amt_cell.number_format = money_format
                    amt_cell.alignment = right_align
                    amt_cell.border = border

                    # Apply color coding
                    if color_type == 'info':
                        amt_cell.fill = info_fill
                        amt_cell.font = bold_font
                    elif color_type == 'warning':
                        amt_cell.fill = warning_fill
                        amt_cell.font = bold_font
                    elif color_type == 'success':
                        amt_cell.fill = success_fill
                        amt_cell.font = bold_font
                    else:
                        amt_cell.fill = neutral_fill
                        amt_cell.font = normal_font

                    row_num += 1
            row_num += 3  # Extra space between years

    # ======================
    # FINAL FORMATTING
    # ======================
    # Adjust column widths
    column_max_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                col_index = cell.column
                content_length = len(str(cell.value))
                column_max_widths[col_index] = max(column_max_widths.get(col_index, 0), content_length)

    for col_index, max_len in column_max_widths.items():
        col_letter = get_column_letter(col_index)
        if max_len <= 10:
            adjusted_width = max_len + 8
        elif max_len <= 20:
            adjusted_width = int(max_len * 1.2)
        else:
            adjusted_width = int(max_len * 1.4)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"yearly_payroll_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response



#
#
# #yearly detail excel
# @login_required
# def export_combined_yearly_summary_to_excel(request):
#     context = get_combined_yearly_detail(request)
#     yearly_summary = context.get('yearly_summary', {})
#
#     section_fields = {
#         'regular': {
#             'taxable_gross': 'Taxable Gross',
#             'non_taxable_gross': 'Non-Taxable Gross',
#             'gross': 'Gross Pay',
#             'pensionable': 'Pensionable',
#             'employee_pension': 'Employee Pension',
#             'employer_pension': 'Employer Pension',
#             'total_pension': 'Total Pension Contribution',
#             'employment_income_tax': 'Income Tax',
#             'total_regular_deduction': 'Total Deductions',
#             'net_pay': 'Net Pay',
#             'expense': 'Expense',
#         },
#         'adjustment': {
#             'taxable_gross': 'Adjusted Taxable Gross',
#             'non_taxable_gross': 'Adjusted Non-Taxable Gross',
#             'gross': 'Adjusted Gross Pay',
#             'adjusted_pensionable': 'Adjusted Pensionable',
#             'employee_pension': 'Adjusted Employee Pension',
#             'employer_pension': 'Adjusted Employer Pension',
#             'total_pension': 'Adjusted Total Pension Contribution',
#             'employment_income_tax': 'Income Tax on Adjustment',
#             'total_adjustment_deduction': 'Adjustment Deductions',
#             'expense': 'Adjusted Expense',
#         },
#         'severance': {
#             'taxable_gross': 'Severance Gross (Taxable)',
#             'gross': 'Severance Gross',
#             'employment_income_tax': 'Severance Income Tax',
#             'total_severance_deduction': 'Total Severance Deductions',
#             'net': 'Severance Net Pay',
#             'expense': 'Severance Expense',
#         },
#         'totals': {
#             'taxable_gross': 'Total Taxable Gross',
#             'non_taxable_gross': 'Total Non-Taxable Gross',
#             'gross': 'Total Gross Pay',
#             'pensionable': 'Total Pensionable',
#             'employee_pension': 'Total Employee Pension',
#             'employer_pension': 'Total Employer Pension',
#             'total_pension': 'Total Pension Contribution',
#             'employment_income_tax': 'Total Income Tax',
#             'total_deduction': 'Total Deductions',
#             'expense': 'Total Expense',
#             'final_net_pay': 'Final Net Pay',
#         }
#     }
#
#     section_colors = {
#         'regular': 'BDD7EE',
#         'adjustment': 'FDE9D9',
#         'severance': 'F8CBAD',
#         'totals': 'D9EAD3',
#     }
#
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Yearly Payroll Summary"
#     row = 1
#
#     for year, data in yearly_summary.items():
#         year_cell = ws.cell(row=row, column=1, value=f"{year} Payroll Summary")
#         year_cell.font = Font(bold=True, size=14)
#         row += 2
#
#         for section_title, key in [('Regular', 'regular'), ('Adjustment', 'adjustment'),
#                                    ('Severance', 'severance'), ('Totals', 'totals')]:
#             section_data = data.get(key, {})
#             labels = section_fields.get(key, {})
#
#             # Skip section if all values are None or 0
#             non_zero = any(
#                 v is not None and float(v) != 0.0
#                 for k, v in section_data.items()
#                 if k in labels
#             )
#             if not non_zero:
#                 continue
#
#             # Section Header
#             sec_cell = ws.cell(row=row, column=1, value=section_title)
#             sec_cell.font = Font(bold=True, size=12)
#             fill_color = section_colors.get(key)
#             if fill_color:
#                 sec_cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
#             row += 1
#
#             # Table Header
#             comp_cell = ws.cell(row=row, column=1, value="Component")
#             amt_cell = ws.cell(row=row, column=2, value="Amount")
#             comp_cell.font = amt_cell.font = Font(bold=True)
#             comp_cell.fill = amt_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
#             comp_cell.alignment = amt_cell.alignment = Alignment(horizontal="center")
#             row += 1
#
#             # Data rows with conditional check
#             for field_key, label in labels.items():
#                 value = section_data.get(field_key)
#                 if value is None or float(value) == 0.0:
#                     continue
#
#                 amount = float(value)
#                 ws.cell(row=row, column=1, value=label)
#                 amt_cell = ws.cell(row=row, column=2, value=amount)
#                 amt_cell.number_format = '#,##0.00'
#                 row += 1
#
#             row += 3  # Extra space between sections
#
#         row += 2  # Space between years
#
#     # Adjust column widths
#     for col in range(1, 3):
#         max_length = 0
#         col_letter = get_column_letter(col)
#         for cell in ws[col_letter]:
#             if cell.value:
#                 max_length = max(max_length, len(str(cell.value)))
#         ws.column_dimensions[col_letter].width = max_length + 5
#
#     response = HttpResponse(
#         content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#     )
#     response['Content-Disposition'] = 'attachment; filename="Yearly_Payroll_Summary.xlsx"'
#     wb.save(response)
#     return response

