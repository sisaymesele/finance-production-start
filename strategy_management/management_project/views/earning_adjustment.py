from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import Sum
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from openpyxl import Workbook
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from management_project.models import EarningAdjustment
from management_project.forms import EarningAdjustmentForm
from management_project.services.earning_adjustment.business import EarningAdjustmentBusinessService
from management_project.services.earning_adjustment.context import get_earning_adjustment_context
from management_project.services.excel_export import ExportUtilityService


#
@login_required
def earning_object_list(request):
    context = get_earning_adjustment_context(request)
    return render(request, 'earning_adjustment/list.html', context)


@login_required
def earnings_object_detail(request):
    context = get_earning_adjustment_context(request)
    return render(request, 'earning_adjustment/detail.html', context)


@login_required
def earning_per_adjusted_month(request):
    context = get_earning_adjustment_context(request)
    return render(request, 'earning_adjustment/per_adjusted_month.html', context)


@login_required
def monthly_earning_adjustment(request):
    context = get_earning_adjustment_context(request)
    return render(request, 'earning_adjustment/monthly_earning.html', context)
#
@login_required
def monthly_earning_adjustment_total(request):
    context = get_earning_adjustment_context(request)
    return render(request, 'earning_adjustment/monthly_earning_total.html', context)

#
@login_required
def create_earning_adjustment(request):
    if request.method == 'POST':
        form = EarningAdjustmentForm(request.POST, request=request)
        if form.is_valid():
            earning_adjustment = form.save(commit=False)
            earning_adjustment.organization_name = request.user.organization_name
            earning_adjustment.save()
            messages.success(request, "Earning Adjustment created successfully!")
            return redirect('earning_adjustment_list')
        else:
            messages.error(request, "Error creating the earning adjustment. Check the form.")
    else:
        form = EarningAdjustmentForm(request=request)

    context = {
        'form': form,
        'form_title': 'Create Earning Adjustment',
        'submit_button_text': 'Create Earning Adjustment',
    }
    return render(request, 'earning_adjustment/form.html', context)


@login_required
def update_earning_adjustment(request, pk):
    earning_adjustment = get_object_or_404(EarningAdjustment, pk=pk, organization_name=request.user.organization_name)

    if request.method == 'POST':
        form = EarningAdjustmentForm(request.POST, instance=earning_adjustment, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, "Earning Adjustment updated successfully!")
            return redirect('earning_adjustment_list')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = EarningAdjustmentForm(instance=earning_adjustment, request=request)

    context = {
        'form': form,
        'form_title': 'Update Earning Adjustment',
        'submit_button_text': 'Update Earning Adjustment',

    }
    return render(request, 'earning_adjustment/form.html', context)


@login_required
def delete_earning_adjustment(request, pk):
    # Fetch the payroll entry to delete
    earning_adjustment = get_object_or_404(EarningAdjustment, pk=pk, organization_name=request.user.organization_name)

    if request.method == "POST":
        earning_adjustment.delete()
        messages.success(request, "Earning Adjustment deleted successfully!")
        return redirect('earning_adjustment_list')

    context = {'earning_adjustment': earning_adjustment}

    return render(request, 'earning_adjustment/delete_confirm.html', context)

#
# #
def export_earning_adjustment_list_to_excel(request):
    context = get_earning_adjustment_context(request)
    adjustments = context.get('earning_adjustments', [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Earning Adjustment Individual"

    # Title row (1st)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=18)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Individual Earning Adjustment"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Subtitle row (2nd)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=18)
    subtitle_cell = ws.cell(row=2, column=1)
    subtitle_cell.value = "Details of earning adjustments"
    subtitle_cell.font = Font(size=10, italic=True)
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Headers (3rd row)
    headers = [
        "Record Month", "Adjusted Month",
        "First Name", "Father Name", "Last Name",
        "Case", "Component", "Earning Amount",
        "Taxable", "Non-Taxable", "Employee Pension",
        "Employer Pension", "Total Pension Contribution",
        "Period Start", "Period End", "Months Covered",
        "Created At", "Updated At"
    ]

    export_util = ExportUtilityService()
    ws.append([export_util.split_header_to_lines(h) for h in headers])

    # Header style
    header_fill = PatternFill(start_color="FF0070C0", end_color="FF0070C0", fill_type="solid")  # Blue
    header_font = Font(bold=True, color="FFFFFFFF")  # White
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(right=Side(style='thin', color='FF000000'))

    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Simple safe_getattr helper for one-level safety
    def safe_getattr(obj, attr, default=""):
        return getattr(obj, attr, default) if obj else default

    # Data rows (start at row 4)
    for ea in adjustments:
        payroll_to_record = getattr(ea, 'payroll_to_record', None)
        payroll_needing_adjustment = getattr(ea, 'payroll_needing_adjustment', None)

        # Manually doing one more safe_getattr nesting for payroll_month (3 levels):
        record_month_obj = safe_getattr(payroll_to_record, 'payroll_month', None)
        record_month = safe_getattr(record_month_obj, 'payroll_month', "")
        record_month_final = safe_getattr(record_month, 'payroll_month', "")  # extra level

        adjusted_month_obj = safe_getattr(payroll_needing_adjustment, 'payroll_month', None)
        adjusted_month = safe_getattr(adjusted_month_obj, 'payroll_month', "")
        adjusted_month_final = safe_getattr(adjusted_month, 'payroll_month', "")  # extra level

        ws.append([
            record_month_final,
            adjusted_month_final,
            safe_getattr(safe_getattr(payroll_to_record, 'personnel_full_name', None), 'first_name', ""),
            safe_getattr(safe_getattr(payroll_to_record, 'personnel_full_name', None), 'father_name', ""),
            safe_getattr(safe_getattr(payroll_to_record, 'personnel_full_name', None), 'last_name', ""),
            ea.get_case_display() if ea else "",
            ea.get_component_display() if ea else "",
            float(ea.earning_amount or 0),
            float(ea.taxable or 0),
            float(ea.non_taxable or 0),
            float(ea.employee_pension_contribution or 0),
            float(ea.employer_pension_contribution or 0),
            float(ea.total_pension or 0),
            ea.period_start.strftime("%Y-%m-%d") if ea.period_start else "",
            ea.period_end.strftime("%Y-%m-%d") if ea.period_end else "",
            ea.months_covered or "",
            ea.created_at.strftime("%Y-%m-%d %H:%M") if ea.created_at else "",
            ea.updated_at.strftime("%Y-%m-%d %H:%M") if ea.updated_at else "",
        ])

    # Adjust column widths
    MIN_WIDTH = 10
    MAX_WIDTH = 25
    for idx, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col_cells:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_length + 2, MIN_WIDTH), MAX_WIDTH)

    # Prepare response
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=earning_adjustments_individual.xlsx'
    return response



#per adjusted month export
@login_required
def export_earning_per_adjusted_month_to_excel(request):
    context = get_earning_adjustment_context(request)
    data = context.get('earning_per_adjusted_month', [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Adjusted Month Earnings"

    # Title row (1st)
    total_columns = 20  # Number of columns in headers
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Adjusted Month Earnings Report"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Headers row (split lines using ExportUtilityService)
    headers = [
        "#", "Record Month", "Adjusted Month", "Personnel Id",
        "First Name", "Father Name", "Last Name",
        "Gross Taxable (Before Adjustment)", "Income Tax (Before Adjustment)",
        "Gross Taxable Pay", "Gross Non-Taxable Pay", "Gross Pay",
        "Total Taxable Pay", "Employment Income Tax Total", "Employment Income Tax",
        "Employee Pension Contribution", "Employer Pension Contribution",
        "Total Pension Contribution", "Total Deduction", "Total Expense"
    ]
    export_util = ExportUtilityService()
    ws.append([export_util.split_header_to_lines(h) for h in headers])

    # Header styles: amber fill, bold, center align, wrap text
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(right=Side(style='thin', color='FF000000'))

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Append data rows
    for idx, adj in enumerate(data, start=1):
        ws.append([
            idx,
            adj.get("payroll_to_record__payroll_month__payroll_month__payroll_month", ""),
            adj.get("payroll_needing_adjustment__payroll_month__payroll_month__payroll_month", ""),
            adj.get("payroll_to_record__personnel_full_name__personnel_id", ""),
            adj.get("payroll_to_record__personnel_full_name__first_name", ""),
            adj.get("payroll_to_record__personnel_full_name__father_name", ""),
            adj.get("payroll_to_record__personnel_full_name__last_name", ""),
            adj.get("payroll_needing_adjustment__gross_taxable_pay", 0),
            adj.get("payroll_needing_adjustment__employment_income_tax", 0),
            adj.get("adjusted_month_gross_taxable_pay", 0),
            adj.get("adjusted_month_gross_non_taxable_pay", 0),
            adj.get("adjusted_month_gross_pay", 0),
            adj.get("adjusted_month_total_taxable_pay", 0),
            adj.get("adjusted_month_employment_income_tax_total", 0),
            adj.get("adjusted_month_employment_income_tax", 0),
            adj.get("adjusted_month_employee_pension_contribution", 0),
            adj.get("adjusted_month_employer_pension_contribution", 0),
            adj.get("adjusted_month_total_pension", 0),
            adj.get("adjusted_month_total_earning_deduction", 0),
            adj.get("adjusted_month_expense", 0),
        ])

    # Adjust column widths with limits
    MIN_WIDTH = 10
    MAX_WIDTH = 15
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except Exception:
                pass
        adjusted_width = max(MIN_WIDTH, min(MAX_WIDTH, max_length + 2))
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Output to HTTP response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=adjusted_month_earning.xlsx'
    return response



# @login_required
# def export_earning_per_adjusted_month_to_excel(request):
#     context = get_earning_adjustment_context(request)
#     data = context.get('earning_per_adjusted_month', [])
#
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Earnings Adjustment By adjusted month"
#
#     # Title row (merged)
#     ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=18)
#     title_cell = ws.cell(row=1, column=1, value="Earnings Adjustment By Adjusted Month")
#     title_cell.font = Font(size=14, bold=True)
#     title_cell.alignment = Alignment(horizontal='center', vertical='center')
#
#     # Headers row
#     headers = [
#         "Record Month", "Adjusted Month", "Personnel ID",
#         "First Name", "Father Name", "Last Name", "Original Gross/Original Tax",
#         "Gross Taxable", "Gross Non-Taxable",
#         "Gross Pay", "Total Taxable Pay", "Employment Tax Total",
#         "Employment Tax", "Employee Pension", "Employer Pension",
#         "Total Pension Contribution", "Net Earning", "Total Expense"
#     ]
#
#     export_util = ExportUtilityService()
#     ws.append([export_util.split_header_to_lines(h) for h in headers])
#
#     # Header styling
#     header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
#     header_font = Font(bold=True)
#     header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#
#     thin_border = Border(right=Side(style='thin', color='FF000000'))
#
#     for cell in ws[2]:
#         cell.fill = header_fill
#         cell.font = header_font
#         cell.alignment = header_alignment
#         cell.border = thin_border
#
#     # Safe dict get helper
#     def safe_get(d, key, default=0):
#         return d.get(key, default) if d else default
#
#     # Data rows
#     for adj in data:
#         ws.append([
#             safe_get(adj, "payroll_to_record__payroll_month__payroll_month__payroll_month", ""),
#             safe_get(adj, "payroll_needing_adjustment__payroll_month__payroll_month__payroll_month", ""),
#             safe_get(adj, "payroll_to_record__personnel_full_name__personnel_id", ""),
#             safe_get(adj, "payroll_to_record__personnel_full_name__first_name", ""),
#             safe_get(adj, "payroll_to_record__personnel_full_name__father_name", ""),
#             safe_get(adj, "payroll_to_record__personnel_full_name__last_name", ""),
#             f"Gross Taxable: {safe_get(adj, 'payroll_needing_adjustment__gross_taxable_pay', 0)} / "
#             f"Tax: {safe_get(adj, 'payroll_needing_adjustment__employment_income_tax', 0)}",
#             safe_get(adj, "adjusted_month_gross_taxable_pay", 0),
#             safe_get(adj, "adjusted_month_gross_non_taxable_pay", 0),
#             safe_get(adj, "adjusted_month_gross_pay", 0),
#             safe_get(adj, "adjusted_month_total_taxable_pay", 0),
#             safe_get(adj, "adjusted_month_employment_income_tax_total", 0),
#             safe_get(adj, "adjusted_month_employment_income_tax", 0),
#             safe_get(adj, "adjusted_month_employee_pension_contribution", 0),
#             safe_get(adj, "adjusted_month_employer_pension_contribution", 0),
#             safe_get(adj, "adjusted_month_total_pension", 0),
#             safe_get(adj, "adjusted_month_total_earning_deduction", 0),
#             safe_get(adj, "adjusted_month_expense", 0),
#         ])
#
#     # Adjust column widths safely (works even with merged cells)
#     for col_idx in range(1, ws.max_column + 1):
#         col_letter = get_column_letter(col_idx)
#         max_length = 0
#         for cell in ws[col_letter]:
#             if cell.value:
#                 max_length = max(max_length, len(str(cell.value)))
#         ws.column_dimensions[col_letter].width = min(max_length + 2, 35)
#
#     # Output to HTTP response
#     output = BytesIO()
#     wb.save(output)
#     output.seek(0)
#
#     response = HttpResponse(
#         output.read(),
#         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#     )
#     response["Content-Disposition"] = "attachment; filename=earning_per_adjusted_month.xlsx"
#     return response





#monthly export
@login_required
def export_monthly_earning_adjustment_to_excel(request):
    context = get_earning_adjustment_context(request)
    data = context.get('monthly_earning_adjustment', [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Earning Adjustment"

    # Title row (1st)
    total_columns = 16  # Number of columns in headers
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Monthly Earning Adjustment Report"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    headers = [
        "Record Month", "Personnel ID", "First Name", "Father Name", "Last Name",
        "Taxable Gross", "Non-Taxable Gross",
        "Gross Pay", "Total Taxable Pay",
        "Employment Income Tax Total", "Employment Income Tax",
        "Employee Pension", "Employer Pension", "Total Pension Contribution",
        "Net Adjustment", "Expense"
    ]

    # Use ExportUtilityService to split header lines
    export_util = ExportUtilityService()
    ws.append([export_util.split_header_to_lines(h) for h in headers])

    # Header styles: amber fill, bold, center align, wrap text
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(right=Side(style='thin', color='FF000000'))

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border


    # Append data rows (no index variable since unused)
    for row in data:
        ws.append([
            row.get("payroll_to_record__payroll_month__payroll_month__payroll_month", ""),
            row.get("payroll_to_record__personnel_full_name__personnel_id", ""),
            row.get("payroll_to_record__personnel_full_name__first_name", ""),
            row.get("payroll_to_record__personnel_full_name__father_name", ""),
            row.get("payroll_to_record__personnel_full_name__last_name", ""),
            row.get("recorded_month_taxable_gross_pay", 0),
            row.get("recorded_month_non_taxable_gross_pay", 0),
            row.get("recorded_month_gross_pay", 0),
            row.get("recorded_month_total_taxable_pay", 0),
            row.get("recorded_month_employment_income_tax_total", 0),
            row.get("recorded_month_employment_income_tax", 0),
            row.get("recorded_month_employee_pension_contribution", 0),
            row.get("recorded_month_employer_pension_contribution", 0),
            row.get("recorded_month_total_pension_contribution", 0),
            row.get("recorded_month_total_earning_deduction", 0),
            row.get("recorded_month_expense", 0),
        ])

    # Adjust column widths with limits
    MIN_WIDTH = 10
    MAX_WIDTH = 15
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except Exception:
                pass
        adjusted_width = max(MIN_WIDTH, min(MAX_WIDTH, max_length + 2))
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=monthly_earning_adjustment.xlsx'
    return response

#

@login_required
def export_monthly_earning_adjustment_aggregate(request):
    # Your aggregated data
    context = get_earning_adjustment_context(request)
    monthly_earning_adjustment_aggregate = context.get('monthly_earning_adjustment_aggregate', [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Earning Adjustment Aggregate"

    # Title row
    total_columns = 12
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Monthly Earning Adjustment Aggregate Report"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Headers
    headers = [
        "Payroll Month",
        "Adjusted Taxable Gross",
        "Adjusted Non-Taxable Gross",
        "Adjusted Gross Pay",
        "Adjusted Total Taxable Pay",
        "Adjusted Income Tax Total",
        "Adjusted Income Tax",
        "Employee Pension",
        "Employer Pension",
        "Total Pension",
        "Earning Deduction",
        "Expense",

    ]

    ws.append(headers)

    # Header styles
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type='solid')
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(right=Side(style='thin'))

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row in monthly_earning_adjustment_aggregate:
        ws.append([
            row.get('payroll_to_record__payroll_month__payroll_month__payroll_month', ''),
            row.get('adjusted_gross_taxable_pay', 0),
            row.get('adjusted_gross_non_taxable_pay', 0),
            row.get('adjusted_gross_pay', 0),
            row.get('adjusted_taxable_pay_total', 0),
            row.get('adjusted_income_tax_total', 0),
            row.get('adjusted_income_tax', 0),
            row.get('employee_pension', 0),
            row.get('employer_pension', 0),
            row.get('total_pension', 0),
            row.get('earning_deduction', 0),
            row.get('expense', 0),

        ])

    # Adjust column widths
    MIN_WIDTH = 12
    MAX_WIDTH = 15
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except Exception:
                pass
        adjusted_width = max(MIN_WIDTH, min(MAX_WIDTH, max_length + 2))
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Return response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=monthly_earning_adjustment_aggregate.xlsx'
    return response