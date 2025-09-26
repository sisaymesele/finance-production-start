from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from management_project.models import SwotAnalysis
from management_project.forms import SwotAnalysisForm
from django.db.models import Q

from django.core.paginator import Paginator

from django.http import HttpResponse
from django.core.paginator import Paginator
import openpyxl
from openpyxl.utils import get_column_letter


# -------------------- SWOT LIST --------------------

@login_required
def swot_analysis_list(request):
    query = request.GET.get('search', '').strip()
    selected_type = request.GET.get('swot_type', '').strip()
    page_number = request.GET.get('page', 1)

    # Base queryset: all SWOT analyses for this organization
    swots = SwotAnalysis.objects.filter(
        organization_name=request.user.organization_name
    )

    # Apply exact filter by SWOT type first
    if selected_type:
        swots = swots.filter(swot_type=selected_type)

    # Apply search filter across multiple fields
    if query:
        search_filter = (
            Q(swot_type__icontains=query) |
            Q(swot_pillar__icontains=query) |
            Q(swot_factor__icontains=query) |
            Q(description__icontains=query)
        )
        swots = swots.filter(search_filter)

    # Ordering
    swots = swots.order_by('swot_type', 'priority', '-created_at')

    # Provide SWOT types for dropdown filter
    swot_types = [choice[0] for choice in SwotAnalysis.SWOT_TYPES]

    # Pagination
    paginator = Paginator(swots, 10)
    page_obj = paginator.get_page(page_number)

    return render(request, 'swot_report/list.html', {
        'swots': page_obj,
        'page_obj': page_obj,
        'search_query': query,
        'swot_types': swot_types,
        'selected_type': selected_type,
    })

# -------------------- CREATE SWOT --------------------


@login_required
def create_swot_analysis(request):
    if request.method == 'POST':
        form = SwotAnalysisForm(request.POST)
        # Only save if the Save button is clicked
        if 'save' in request.POST and form.is_valid():
            swot_entry = form.save(commit=False)
            swot_entry.organization_name = request.user.organization_name
            swot_entry.save()
            messages.success(request, "SWOT entry created successfully!")
            return redirect('swot_analysis_list')
    else:
        form = SwotAnalysisForm()

    return render(request, 'swot_analysis/form.html', {'form': form})


# -------------------- UPDATE SWOT --------------------
@login_required
def update_swot_analysis(request, pk):
    # Fetch the entry only if it belongs to the user's organization
    entry = get_object_or_404(SwotAnalysis, pk=pk, organization_name=request.user.organization_name)

    if request.method == 'POST':
        form = SwotAnalysisForm(request.POST, instance=entry)
        if 'save' in request.POST and form.is_valid():
            form.save()
            messages.success(request, "SWOT entry updated successfully!")
            return redirect('swot_analysis_list')
        else:
            messages.error(request, "Error updating SWOT entry. Please check the form.")
    else:
        form = SwotAnalysisForm(instance=entry)

    return render(request, 'swot_analysis/form.html', {'form': form})


# -------------------- DELETE SWOT --------------------

@login_required
def delete_swot_analysis(request, pk):
    entry = get_object_or_404(
        SwotAnalysis,
        pk=pk,
        organization_name=request.user.organization_name
    )

    if request.method == 'POST':
        entry.delete()
        messages.success(request, "SWOT entry deleted successfully!")
        return redirect('swot_analysis_list')

    return render(request, 'swot_analysis/delete_confirm.html', {'entry': entry})



@login_required
def swot_analysis_list(request):
    query = request.GET.get("search", "").strip()
    page_number = request.GET.get("page", 1)
    filter_type = request.GET.get("swot_type", "").strip()
    export = request.GET.get("export", "").lower()

    # Base queryset for the organization
    swots = SwotAnalysis.objects.filter(
        organization_name=request.user.organization_name
    )

    # Filter by SWOT type if provided
    if filter_type:
        swots = swots.filter(swot_type=filter_type)

    # Apply search filter
    if query:
        swots = swots.filter(
            Q(swot_pillar__icontains=query) |
            Q(swot_factor__icontains=query) |
            Q(description__icontains=query)
        )

    # Ordering
    swots = swots.order_by("swot_type", "priority", "-created_at")

    # Excel export
    if export == "excel":
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=swot_analysis.xlsx'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SWOT Analysis"

        # Headers
        headers = [
            "SWOT Type", "Pillar", "Factor", "Description",
            "Priority", "Impact", "Likelihood", "Created At", "Updated At"
        ]
        ws.append(headers)

        # Data rows
        for s in swots:
            ws.append([
                s.swot_type,
                s.swot_pillar,
                s.swot_factor,
                s.description or "",
                s.priority,
                s.impact,
                s.likelihood or "",
                s.created_at.strftime("%Y-%m-%d"),
                s.updated_at.strftime("%Y-%m-%d"),
            ])

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(response)
        return response

    # Pagination
    paginator = Paginator(swots, 10)
    page_obj = paginator.get_page(page_number)

    # SWOT types for filter dropdown
    swot_types = [choice[0] for choice in SwotAnalysis.SWOT_TYPES]

    return render(request, "swot_analysis/list.html", {
        "swots": page_obj,
        "page_obj": page_obj,
        "search_query": query,
        "swot_types": swot_types,
        "selected_type": filter_type,
    })
