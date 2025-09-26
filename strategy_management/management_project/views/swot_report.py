from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Q
import calendar
from django.core.paginator import Paginator
from django.urls import reverse

from management_project.models import SwotReport, StrategicCycle, StrategicReport
from management_project.forms import SwotReportForm

from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter



@login_required
def swot_report_by_cycle_list(request):
    """List distinct strategic cycles for the current organization with all info."""
    cycles_qs = StrategicCycle.objects.filter(
        organization_name=request.user.organization_name
    ).order_by('-start_date')

    # Build a list of dicts including calculated properties
    cycles = []
    for cycle in cycles_qs:
        cycles.append({
            'name': cycle.name,
            'time_horizon': cycle.time_horizon,
            'time_horizon_type': cycle.time_horizon_type,
            'start_date': cycle.start_date,
            'end_date': cycle.end_date,
            'slug': cycle.slug,
            'duration_days': (cycle.end_date - cycle.start_date).days if cycle.start_date and cycle.end_date else None,
            'start_month_name': calendar.month_name[cycle.start_date.month] if cycle.start_date else None,
            'start_quarter': (cycle.start_date.month - 1) // 3 + 1 if cycle.start_date else None,
            'start_year': cycle.start_date.year if cycle.start_date else None,
        })

    return render(request, 'swot_report/cycle_list.html', {
        'strategic_cycles': cycles
    })


# -------------------- SWOT LIST --
@login_required
def swot_report_list(request):
    query = request.GET.get("search", "").strip()
    cycle_slug = request.GET.get("cycle", "").strip()
    page_number = request.GET.get("page", 1)

    # Get all SWOT reports for the organization
    swots = SwotReport.objects.filter(
        organization_name=request.user.organization_name
    )

    # Filter by strategic cycle if provided
    if cycle_slug:
        try:
            strategic_cycle = StrategicCycle.objects.get(
                slug=cycle_slug,
                organization_name=request.user.organization_name
            )
            strategic_reports = StrategicReport.objects.filter(
                action_plan__strategic_cycle=strategic_cycle
            )
            swots = swots.filter(strategic_report_period__in=strategic_reports)
        except StrategicCycle.DoesNotExist:
            pass

    # Apply search filter
    if query:
        swots = swots.filter(
            Q(swot_type__icontains=query) |
            Q(swot_pillar__icontains=query) |
            Q(swot_factor__icontains=query) |
            Q(description__icontains=query) |
            Q(strategic_report_period__action_plan__strategic_cycle__name__icontains=query) |
            Q(strategic_report_period__action_plan__strategic_cycle__time_horizon__icontains=query) |
            Q(strategic_report_period__action_plan__strategic_cycle__time_horizon_type__icontains=query)
        )

    # Available strategic cycles for filter dropdown
    strategic_cycles = StrategicCycle.objects.filter(
        organization_name=request.user.organization_name
    ).order_by('-start_date')

    # Ordering
    swots = swots.order_by("swot_type", "priority", "-created_at")

    # Pagination
    paginator = Paginator(swots, 10)  # 10 items per page
    page_obj = paginator.get_page(page_number)

    return render(request, "swot_report/list.html", {
        "swots": page_obj,
        "page_obj": page_obj,
        "search_query": query,
        "strategic_cycles": strategic_cycles,
        "selected_cycle": cycle_slug,
    })
#
#
# -------------------- CREATE SWOT --------------------


@login_required
def create_swot_report(request):
    org = request.user.organization_name
    # Get the latest StrategicReport for this organization
    latest_report = StrategicReport.objects.filter(organization_name=org).order_by('-id').first()

    if request.method == 'POST':
        form = SwotReportForm(request.POST, request=request)
        if 'save' in request.POST and form.is_valid():
            swot_entry = form.save(commit=False)
            swot_entry.organization_name = org
            # Assign the selected strategic report or fallback to latest
            if not swot_entry.strategic_report_period:
                swot_entry.strategic_report_period = latest_report
            swot_entry.save()
            messages.success(request, "SWOT entry created successfully!")
            return redirect('swot_report_list')
        else:
            messages.error(request, "Error creating SWOT entry. Please check the form.")
    else:
        form = SwotReportForm(request=request)
        if latest_report:
            form.fields['strategic_report_period'].initial = latest_report

    return render(request, 'swot_report/form.html', {'form': form})

# -------------------- UPDATE SWOT --------------------
@login_required
def update_swot_report(request, pk):
    entry = get_object_or_404(
        SwotReport,
        pk=pk,
        organization_name=request.user.organization_name
    )

    if request.method == 'POST':
        form = SwotReportForm(request.POST, instance=entry, request=request)
        if 'save' in request.POST and form.is_valid():
            form.save()
            messages.success(request, "SWOT entry updated successfully!")
            return redirect('swot_report_list')
        else:
            messages.error(request, "Error updating SWOT entry. Please check the form.")
    else:
        form = SwotReportForm(instance=entry, request=request)

    return render(request, 'swot_report/form.html', {'form': form})


# -------------------- DELETE SWOT --------------------
@login_required
def delete_swot_report(request, pk):
    entry = get_object_or_404(
        SwotReport,
        pk=pk,
        organization_name=request.user.organization_name
    )

    if request.method == 'POST':
        entry.delete()
        messages.success(request, "SWOT entry deleted successfully!")
        return redirect('swot_report_list')

    return render(request, 'swot_report/delete_confirm.html', {'entry': entry})



@login_required
def swot_report_list(request):
    query = request.GET.get("search", "").strip()
    cycle_slug = request.GET.get("cycle", "").strip()
    page_number = request.GET.get("page", 1)
    export = request.GET.get("export", "").lower()

    # Get all SWOT reports for the organization
    swots = SwotReport.objects.filter(
        organization_name=request.user.organization_name
    )

    # Filter by strategic cycle if provided
    if cycle_slug:
        try:
            strategic_cycle = StrategicCycle.objects.get(
                slug=cycle_slug,
                organization_name=request.user.organization_name
            )
            strategic_reports = StrategicReport.objects.filter(
                action_plan__strategic_cycle=strategic_cycle
            )
            swots = swots.filter(strategic_report_period__in=strategic_reports)
        except StrategicCycle.DoesNotExist:
            pass

    # Apply search filter
    if query:
        swots = swots.filter(
            Q(swot_type__icontains=query) |
            Q(swot_pillar__icontains=query) |
            Q(swot_factor__icontains=query) |
            Q(description__icontains=query) |
            Q(strategic_report_period__action_plan__strategic_cycle__name__icontains=query) |
            Q(strategic_report_period__action_plan__strategic_cycle__time_horizon__icontains=query) |
            Q(strategic_report_period__action_plan__strategic_cycle__time_horizon_type__icontains=query)
        )

    # Ordering
    swots = swots.order_by("swot_type", "priority", "-created_at")

    # Excel Export
    if export == "excel":
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=swot_reports.xlsx'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SWOT Reports"

        # Headers
        headers = [
            "Strategic Cycle", "SWOT Type", "Pillar", "Factor",
            "Description", "Priority", "Impact", "Likelihood", "Created At"
        ]
        ws.append(headers)

        # Data rows
        for s in swots:
            ws.append([
                s.strategic_report_period.action_plan.strategic_cycle.name,
                s.swot_type,
                s.swot_pillar,
                s.swot_factor,
                s.description,
                s.priority,
                s.impact,
                s.likelihood or "",
                s.created_at.strftime("%Y-%m-%d"),
            ])

        # Optional: adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(response)
        return response

    # Pagination
    paginator = Paginator(swots, 10)  # 10 items per page
    page_obj = paginator.get_page(page_number)

    # Available strategic cycles for filter dropdown
    strategic_cycles = StrategicCycle.objects.filter(
        organization_name=request.user.organization_name
    ).order_by('-start_date')

    return render(request, "swot_report/list.html", {
        "swots": page_obj,
        "page_obj": page_obj,
        "search_query": query,
        "strategic_cycles": strategic_cycles,
        "selected_cycle": cycle_slug,
    })

