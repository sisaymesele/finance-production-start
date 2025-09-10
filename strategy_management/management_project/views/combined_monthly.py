# Django core
from django.shortcuts import render
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
# Built-in and standard libraries
from decimal import Decimal
from datetime import datetime
# Third-party
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.cell.cell import MergedCell
import openpyxl
from openpyxl.utils import get_column_letter
import plotly.graph_objs as go
import plotly.offline as opy
# Project-specific
from management_project.services.combined.monthly_context import get_combined_monthly_detail

#
@login_required
def monthly_combined_detail(request):
    context = get_combined_monthly_detail(request)
    return render(request, 'combined_payroll/monthly_detail.html', context)


@login_required
def monthly_combined_summary(request):
    context = get_combined_monthly_detail(request)
    return render(request, 'combined_payroll/monthly_summary.html', context)



@login_required
def export_combined_monthly_detail_to_excel(request):
    context = get_combined_monthly_detail(request)
    monthly_list = context['page_obj'].object_list

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Combined Monthly Payroll"

    # ======================
    # STYLE DEFINITIONS
    # ======================
    # Fonts
    title_font = Font(bold=True, size=14, color="FFFFFF")
    section_font = Font(bold=True, size=12, color="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF")
    normal_font = Font(size=11)
    bold_font = Font(bold=True, size=11)
    total_font = Font(bold=True, size=11, color="FFFFFF")

    # Colors (matching HTML template)
    color_palette = {
        'primary': '4472C4',  # Blue (card header)
        'success': '70AD47',  # Green (regular payroll)
        'warning': 'FFC000',  # Yellow (earning adjustments)
        'danger': 'ED7D31',  # Orange (deduction adjustments)
        'info': '5B9BD5',  # Light blue (adjustment summary)
        'purple': '7030A0',  # Purple (severance)
        'totals': '4472C4',  # Blue (totals)
    }

    # Fills
    title_fill = PatternFill(start_color=color_palette['primary'], end_color=color_palette['primary'],
                             fill_type="solid")
    regular_fill = PatternFill(start_color=color_palette['success'], end_color=color_palette['success'],
                               fill_type="solid")
    earning_fill = PatternFill(start_color=color_palette['warning'], end_color=color_palette['warning'],
                               fill_type="solid")
    deduction_fill = PatternFill(start_color=color_palette['danger'], end_color=color_palette['danger'],
                                 fill_type="solid")
    summary_fill = PatternFill(start_color=color_palette['info'], end_color=color_palette['info'], fill_type="solid")
    severance_fill = PatternFill(start_color=color_palette['purple'], end_color=color_palette['purple'],
                                 fill_type="solid")
    totals_fill = PatternFill(start_color=color_palette['totals'], end_color=color_palette['totals'], fill_type="solid")

    # Borders
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    # Number formats
    money_format = '#,##0.00'

    # Alignment
    center_align = Alignment(horizontal="center")
    left_align = Alignment(horizontal="left")
    right_align = Alignment(horizontal="right")

    row_num = 1

    for item in monthly_list:
        # ======================
        # MONTH HEADER
        # ======================
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
        header_cell = ws.cell(row=row_num, column=1, value=f"Combined Monthly Payroll Detail for {item['month']}")
        header_cell.font = title_font
        header_cell.fill = title_fill
        header_cell.alignment = center_align
        row_num += 2

        # ======================
        # REGULAR PAYROLL SECTION
        # ======================
        if 'regular' in item:
            # Section Header
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
            header_cell = ws.cell(row=row_num, column=1, value="REGULAR PAYROLL")
            header_cell.font = Font(bold=True, size=12, color=color_palette['success'])
            row_num += 1

            # Regular Components
            if 'regular_item_by_component' in item:
                headers = ["Component", "Amount"]
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=header)
                    cell.font = header_font
                    cell.fill = regular_fill
                    cell.alignment = center_align
                    cell.border = border
                row_num += 1

                for comp, amount in item['regular_item_by_component'].items():
                    if amount:
                        ws.cell(row=row_num, column=1, value=comp).font = normal_font
                        ws.cell(row=row_num, column=1).alignment = left_align
                        ws.cell(row=row_num, column=1).border = border

                        amt_cell = ws.cell(row=row_num, column=2, value=float(amount))
                        amt_cell.number_format = money_format
                        amt_cell.alignment = right_align
                        amt_cell.border = border
                        row_num += 1
                row_num += 1

            # Regular Summary
            headers = ["Regular Payroll Summary Item", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = regular_fill
                cell.alignment = center_align
                cell.border = border
            row_num += 1

            regular_summary = [
                ("Taxable Gross", item['regular'].get('taxable_gross')),
                ("Non-Taxable Gross", item['regular'].get('non_taxable_gross')),
                ("Gross Pay", item['regular'].get('gross'), 'primary'),
                ("Pensionable Amount", item['regular'].get('pensionable')),
                ("Employee Pension", item['regular'].get('employee_pension')),
                ("Employer Pension", item['regular'].get('employer_pension')),
                ("Total Pension", item['regular'].get('total_pension')),
                ("Employment Income Tax", item['regular'].get('employment_income_tax')),
                ("Total Regular Deduction", item['regular'].get('total_regular_deduction')),
                ("Net Pay", item['regular'].get('net_pay'), 'success'),
                ("Expense", item['regular'].get('expense')),
            ]

            for label, value, *style in regular_summary:
                if value:
                    cell = ws.cell(row=row_num, column=1, value=label)
                    cell.font = bold_font
                    cell.alignment = left_align
                    cell.border = border

                    amt_cell = ws.cell(row=row_num, column=2, value=float(value))
                    amt_cell.number_format = money_format
                    amt_cell.alignment = right_align
                    amt_cell.border = border

                    if style:
                        amt_cell.font = total_font
                        amt_cell.fill = PatternFill(start_color=color_palette[style[0]],
                                                    end_color=color_palette[style[0]], fill_type="solid")

                    row_num += 1
            row_num += 2

        # ======================
        # EARNING ADJUSTMENTS SECTION
        # ======================
        if item.get('show_earning', False):
            # Section Header
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
            header_cell = ws.cell(row=row_num, column=1, value="EARNING ADJUSTMENTS")
            header_cell.font = Font(bold=True, size=12, color=color_palette['warning'])
            row_num += 1

            # Earning Components
            if 'earning_adj_by_component' in item:
                headers = ["Component", "Earning Amount", "Taxable", "Non-Taxable",
                           "Employee Pension", "Employer Pension", "Total Pension"]
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=header)
                    cell.font = header_font
                    cell.fill = earning_fill
                    cell.alignment = center_align
                    cell.border = border
                row_num += 1

                for comp, vals in item['earning_adj_by_component'].items():
                    if vals.get('earning_amount'):
                        ws.cell(row=row_num, column=1, value=comp).font = normal_font
                        ws.cell(row=row_num, column=1).alignment = left_align
                        ws.cell(row=row_num, column=1).border = border

                        columns = [
                            ('earning_amount', 2),
                            ('taxable', 3),
                            ('non_taxable', 4),
                            ('employee_pension_contribution', 5),
                            ('employer_pension_contribution', 6),
                            ('total_pension', 7)
                        ]

                        for key, col in columns:
                            val = float(vals.get(key, Decimal('0.00')))
                            ws.cell(row=row_num, column=col, value=val).number_format = money_format
                            ws.cell(row=row_num, column=col).alignment = right_align
                            ws.cell(row=row_num, column=col).border = border

                        row_num += 1
                row_num += 1

            # Earning Summary
            headers = ["Earning Adjustment Summary Item", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = earning_fill
                cell.alignment = center_align
                cell.border = border
            row_num += 1

            earning_summary = [
                ("Taxable Gross", item['adjustment'].get('taxable_gross')),
                ("Non-Taxable Gross", item['adjustment'].get('non_taxable_gross')),
                ("Gross Adjustment", item['adjustment'].get('gross'), 'primary'),
                ("Adjusted Pensionable", item['adjustment'].get('adjusted_pensionable')),
                ("Employee Pension", item['adjustment'].get('employee_pension')),
                ("Employer Pension", item['adjustment'].get('employer_pension')),
                ("Total Pension", item['adjustment'].get('total_pension')),
                ("Employment Income Tax", item['adjustment'].get('employment_income_tax')),
                ("Earning Adjustment Deduction", item['adjustment'].get('earning_adjustment_deduction')),
                ("Expense", item['adjustment'].get('expense')),
            ]

            for label, value, *style in earning_summary:
                if value:
                    cell = ws.cell(row=row_num, column=1, value=label)
                    cell.font = bold_font
                    cell.alignment = left_align
                    cell.border = border

                    amt_cell = ws.cell(row=row_num, column=2, value=float(value))
                    amt_cell.number_format = money_format
                    amt_cell.alignment = right_align
                    amt_cell.border = border

                    if style:
                        amt_cell.font = total_font
                        amt_cell.fill = PatternFill(start_color=color_palette[style[0]],
                                                    end_color=color_palette[style[0]], fill_type="solid")

                    row_num += 1
            row_num += 2

        # ======================
        # DEDUCTION ADJUSTMENTS SECTION
        # ======================
        if item.get('show_deduction', False):
            # Section Header
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
            header_cell = ws.cell(row=row_num, column=1, value="DEDUCTION ADJUSTMENTS")
            header_cell.font = Font(bold=True, size=12, color=color_palette['danger'])
            row_num += 1

            # Deduction Components
            if 'deduction_adj_by_component' in item:
                headers = ["Component", "Amount"]
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=header)
                    cell.font = header_font
                    cell.fill = deduction_fill
                    cell.alignment = center_align
                    cell.border = border
                row_num += 1

                for comp, amount in item['deduction_adj_by_component'].items():
                    if amount:
                        ws.cell(row=row_num, column=1, value=comp).font = normal_font
                        ws.cell(row=row_num, column=1).alignment = left_align
                        ws.cell(row=row_num, column=1).border = border

                        amt_cell = ws.cell(row=row_num, column=2, value=float(amount))
                        amt_cell.number_format = money_format
                        amt_cell.alignment = right_align
                        amt_cell.border = border
                        row_num += 1
                row_num += 1

            # Deduction Summary
            headers = ["Deduction Adjustment Summary Item", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = deduction_fill
                cell.alignment = center_align
                cell.border = border
            row_num += 1

            deduction_summary = [
                ("Individual Adjustment Deduction", item['adjustment'].get('individual_adjustment_deduction')),
            ]

            for label, value in deduction_summary:
                if value:
                    cell = ws.cell(row=row_num, column=1, value=label)
                    cell.font = bold_font
                    cell.alignment = left_align
                    cell.border = border

                    amt_cell = ws.cell(row=row_num, column=2, value=float(value))
                    amt_cell.number_format = money_format
                    amt_cell.alignment = right_align
                    amt_cell.border = border
                    row_num += 1
            row_num += 1

        # ======================
        # ======================
        # ADJUSTMENT SUMMARY SECTION
        # ======================
        if (
            item.get('show_earning', False) or
            item.get('show_deduction', False)
        ) and 'adjustment' in item:
            # Section Header
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
            header_cell = ws.cell(row=row_num, column=1, value="ADJUSTMENT SUMMARY")
            header_cell.font = Font(bold=True, size=12, color=color_palette['info'])
            row_num += 1

            # Adjustment Summary
            headers = ["Total Adjustment Summary Item", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = summary_fill
                cell.alignment = center_align
                cell.border = border
            row_num += 1

            adjustment_summary = [
                ("Taxable Gross", item['adjustment'].get('taxable_gross')),
                ("Non-Taxable Gross", item['adjustment'].get('non_taxable_gross')),
                ("Gross Adjustment", item['adjustment'].get('gross'), 'primary'),
                ("Adjusted Pensionable", item['adjustment'].get('adjusted_pensionable')),
                ("Employee Pension", item['adjustment'].get('employee_pension')),
                ("Employer Pension", item['adjustment'].get('employer_pension')),
                ("Total Pension", item['adjustment'].get('total_pension')),
                ("Employment Income Tax", item['adjustment'].get('employment_income_tax')),
                ("Earning Adjustment Deduction", item['adjustment'].get('earning_adjustment_deduction')),
                ("Individual Adjustment Deduction", item['adjustment'].get('individual_adjustment_deduction')),
                ("Total Monthly Adjustment Deduction", item['adjustment'].get('total_monthly_adjustment_deduction'),
                 'primary'),
                ("Net Monthly Adjustment", item['adjustment'].get('net_monthly_adjustment'), 'success'),
                ("Expense", item['adjustment'].get('expense')),
            ]

            for label, value, *style in adjustment_summary:
                if value:
                    cell = ws.cell(row=row_num, column=1, value=label)
                    cell.font = bold_font
                    cell.alignment = left_align
                    cell.border = border

                    amt_cell = ws.cell(row=row_num, column=2, value=float(value))
                    amt_cell.number_format = money_format
                    amt_cell.alignment = right_align
                    amt_cell.border = border

                    if style:
                        amt_cell.font = total_font
                        amt_cell.fill = PatternFill(
                            start_color=color_palette[style[0]],
                            end_color=color_palette[style[0]],
                            fill_type="solid"
                        )
                    row_num += 1
            row_num += 2


        # ======================
        # SEVERANCE PAY SECTION
        # ======================
        if 'severance' in item and any(item['severance'].values()):
            # Section Header
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
            header_cell = ws.cell(row=row_num, column=1, value="SEVERANCE PAY")
            header_cell.font = Font(bold=True, size=12, color=color_palette['purple'])
            row_num += 1

            # Severance Details
            headers = ["Component", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = severance_fill
                cell.alignment = center_align
                cell.border = border
            row_num += 1

            severance_items = [
                ("Taxable Gross", item['severance'].get('taxable_gross')),
                ("Gross Pay", item['severance'].get('gross'), 'primary'),
                ("Employment Income Tax", item['severance'].get('employment_income_tax')),
                ("Total Severance Deduction", item['severance'].get('total_severance_deduction')),
                ("Net Pay", item['severance'].get('net'), 'success'),
                ("Expense", item['severance'].get('expense')),
            ]

            for label, value, *style in severance_items:
                if value:
                    cell = ws.cell(row=row_num, column=1, value=label)
                    cell.font = bold_font
                    cell.alignment = left_align
                    cell.border = border

                    amt_cell = ws.cell(row=row_num, column=2, value=float(value))
                    amt_cell.number_format = money_format
                    amt_cell.alignment = right_align
                    amt_cell.border = border

                    if style:
                        amt_cell.font = total_font
                        amt_cell.fill = PatternFill(start_color=color_palette[style[0]],
                                                    end_color=color_palette[style[0]], fill_type="solid")

                    row_num += 1
            row_num += 2

        # ======================
        # TOTAL SUMMARY SECTION
        # ======================
        if 'totals' in item:
            # Section Header
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=8)
            header_cell = ws.cell(row=row_num, column=1, value="TOTAL SUMMARY")
            header_cell.font = Font(bold=True, size=12, color=color_palette['totals'])
            row_num += 1

            # Total Details
            headers = ["Total Summary Items", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = totals_fill
                cell.alignment = center_align
                cell.border = border
            row_num += 1

            summary_items = [
                ("Taxable Gross", item['totals'].get('taxable_gross'), 'info'),
                ("Non-Taxable Gross", item['totals'].get('non_taxable_gross'), 'info'),
                ("Total Gross", item['totals'].get('gross'), 'info'),
                ("Pensionable", item['totals'].get('pensionable'), 'warning'),
                ("Employee Pension", item['totals'].get('employee_pension'), 'warning'),
                ("Employer Pension", item['totals'].get('employer_pension'), 'warning'),
                ("Total Pension", item['totals'].get('total_pension'), 'warning'),
                ("Employment Income Tax", item['totals'].get('employment_income_tax')),
                ("Total Deduction", item['totals'].get('total_deduction')),
                ("Final Net Pay", item['totals'].get('final_net_pay'), 'success'),
                ("Expense", item['totals'].get('expense')),
            ]

            for label, value, *style in summary_items:
                if value:
                    cell = ws.cell(row=row_num, column=1, value=label)
                    cell.font = bold_font
                    cell.alignment = left_align
                    cell.border = border

                    amt_cell = ws.cell(row=row_num, column=2, value=float(value))
                    amt_cell.number_format = money_format
                    amt_cell.alignment = right_align
                    amt_cell.border = border

                    if style:
                        amt_cell.font = total_font
                        amt_cell.fill = PatternFill(start_color=color_palette[style[0]],
                                                    end_color=color_palette[style[0]], fill_type="solid")

                    row_num += 1
            row_num += 3  # Extra space between months

    # ======================
    # FINAL FORMATTING
    # ======================
    # Adjust column widths - handle merged cells properly
    column_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue  # Skip merged cells

            if cell.value:
                col_letter = openpyxl.utils.get_column_letter(cell.column)
                content_length = len(str(cell.value))
                if col_letter in column_widths:
                    if content_length > column_widths[col_letter]:
                        column_widths[col_letter] = content_length
                else:
                    column_widths[col_letter] = content_length

    for col_letter, max_len in column_widths.items():
        adjusted_width = (max_len + 2) * 1.2
        ws.column_dimensions[col_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"combined_monthly_payroll_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

#

#


#
@login_required
def export_combined_monthly_summary_to_excel(request):
    context = get_combined_monthly_detail(request)
    monthly_list = context['page_obj'].object_list

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Payroll Summary"

    # ======================
    # STYLE DEFINITIONS MATCHING HTML TEMPLATE
    # ======================
    # Fonts
    title_font = Font(bold=True, size=14, color="FFFFFF")
    section_font = Font(bold=True, size=12)
    header_font = Font(bold=True, color="FFFFFF")
    normal_font = Font(size=11)
    bold_font = Font(bold=True, size=11)
    positive_font = Font(size=11, color="000000")  # Black for amounts
    negative_font = Font(size=11, color="000000")  # Black for amounts
    neutral_font = Font(size=11, color="000000")  # Black for neutral amounts

    # Colors matching HTML template classes
    color_palette = {
        'title': '3498DB',  # Blue (card header)
        'regular': '2E86C1',  # Primary blue (section header)
        'adjustment': 'F39C12',  # Orange (section header)
        'severance': 'E74C3C',  # Red (section header)
        'totals': '2ECC71',  # Green (section header)
        'primary': 'D6EAF8',  # Light blue (table-primary)
        'success': 'D5F5E3',  # Light green (table-success)
        'info': 'D6EAF8',  # Light blue (table-info)
        'warning': 'FDEBD0',  # Light orange (table-warning)
        'danger': 'FADBD8',  # Light red (table-danger)
        'neutral': 'FFFFFF',  # White
    }

    # Fills
    title_fill = PatternFill(start_color=color_palette['title'], end_color=color_palette['title'], fill_type="solid")
    regular_fill = PatternFill(start_color=color_palette['regular'], end_color=color_palette['regular'],
                               fill_type="solid")
    adjustment_fill = PatternFill(start_color=color_palette['adjustment'], end_color=color_palette['adjustment'],
                                  fill_type="solid")
    severance_fill = PatternFill(start_color=color_palette['severance'], end_color=color_palette['severance'],
                                 fill_type="solid")
    totals_fill = PatternFill(start_color=color_palette['totals'], end_color=color_palette['totals'], fill_type="solid")
    primary_fill = PatternFill(start_color=color_palette['primary'], end_color=color_palette['primary'],
                               fill_type="solid")
    success_fill = PatternFill(start_color=color_palette['success'], end_color=color_palette['success'],
                               fill_type="solid")
    info_fill = PatternFill(start_color=color_palette['info'], end_color=color_palette['info'], fill_type="solid")
    warning_fill = PatternFill(start_color=color_palette['warning'], end_color=color_palette['warning'],
                               fill_type="solid")
    danger_fill = PatternFill(start_color=color_palette['danger'], end_color=color_palette['danger'], fill_type="solid")
    neutral_fill = PatternFill(start_color=color_palette['neutral'], end_color=color_palette['neutral'],
                               fill_type="solid")

    # Borders
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    # Number formats
    money_format = '#,##0.00'

    # Alignment
    center_align = Alignment(horizontal="center")
    left_align = Alignment(horizontal="left")
    right_align = Alignment(horizontal="right")

    row_num = 1

    for item in monthly_list:
        # ======================
        # MONTH HEADER - matches card header
        # ======================
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
        header_cell = ws.cell(row=row_num, column=1, value=f"Monthly Payroll Summary For {item['month']}")
        header_cell.font = title_font
        header_cell.fill = title_fill
        header_cell.alignment = center_align
        row_num += 2

        # ======================
        # REGULAR PAYROLL SUMMARY - matches table-primary/success
        # ======================
        if 'regular' in item and any(val != Decimal('0.00') for val in item['regular'].values()):
            # Section Header
            ws.cell(row=row_num, column=1, value="REGULAR PAYROLL SUMMARY").font = Font(bold=True, size=12,
                                                                                        color=color_palette['regular'])
            row_num += 1

            headers = ["Item", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = regular_fill
                cell.alignment = center_align
                cell.border = border
            row_num += 1

            regular_summary = [
                ("taxable_gross", item['regular'].get('taxable_gross'), 'neutral'),
                ("non_taxable_gross", item['regular'].get('non_taxable_gross'), 'neutral'),
                ("gross", item['regular'].get('gross'), 'primary'),
                ("pensionable", item['regular'].get('pensionable'), 'neutral'),
                ("employee_pension", item['regular'].get('employee_pension'), 'neutral'),
                ("employer_pension", item['regular'].get('employer_pension'), 'neutral'),
                ("total_pension", item['regular'].get('total_pension'), 'neutral'),
                ("employment_income_tax", item['regular'].get('employment_income_tax'), 'neutral'),
                ("total_regular_deduction", item['regular'].get('total_regular_deduction'), 'neutral'),
                ("net_pay", item['regular'].get('net_pay'), 'success'),
                ("expense", item['regular'].get('expense'), 'neutral'),
            ]

            for field_name, value, color_type in regular_summary:
                if value != Decimal('0.00'):
                    label = field_name.replace('_', ' ').title()
                    ws.cell(row=row_num, column=1, value=label).font = bold_font
                    ws.cell(row=row_num, column=1).alignment = left_align
                    ws.cell(row=row_num, column=1).border = border

                    amt_cell = ws.cell(row=row_num, column=2, value=float(value))
                    amt_cell.number_format = money_format
                    amt_cell.alignment = right_align
                    amt_cell.border = border

                    # Apply color coding
                    if color_type == 'primary':
                        amt_cell.fill = primary_fill
                        amt_cell.font = bold_font
                    elif color_type == 'success':
                        amt_cell.fill = success_fill
                        amt_cell.font = bold_font
                    else:
                        amt_cell.fill = neutral_fill
                        amt_cell.font = normal_font

                    row_num += 1
            row_num += 2

        # ======================
        # ADJUSTMENT SUMMARY - matches table-primary/success
        # ======================
        if 'adjustment' in item:
            adj_summary = [
                ("taxable_gross", item['adjustment'].get('taxable_gross'), 'neutral'),
                ("non_taxable_gross", item['adjustment'].get('non_taxable_gross'), 'neutral'),
                ("gross", item['adjustment'].get('gross'), 'primary'),
                ("adjusted_pensionable", item['adjustment'].get('adjusted_pensionable'), 'neutral'),
                ("employee_pension", item['adjustment'].get('employee_pension'), 'neutral'),
                ("employer_pension", item['adjustment'].get('employer_pension'), 'neutral'),
                ("total_pension", item['adjustment'].get('total_pension'), 'neutral'),
                ("employment_income_tax", item['adjustment'].get('employment_income_tax'), 'neutral'),
                ("earning_adjustment_deduction", item['adjustment'].get('earning_adjustment_deduction'), 'neutral'),
                ("individual_adjustment_deduction", item['adjustment'].get('individual_adjustment_deduction'),
                 'neutral'),
                ("total_monthly_adjustment_deduction", item['adjustment'].get('total_monthly_adjustment_deduction'),
                 'primary'),
                ("net_monthly_adjustment", item['adjustment'].get('net_monthly_adjustment'), 'success'),
                ("expense", item['adjustment'].get('expense'), 'neutral'),
            ]

            # Check if there are any non-zero values in the adjustment items
            has_non_zero_adjustments = any(value != Decimal('0.00') for _, value, _ in adj_summary)

            if has_non_zero_adjustments:
                # Section Header
                ws.cell(row=row_num, column=1, value="ADJUSTMENT SUMMARY").font = Font(bold=True, size=12,
                                                                                       color=color_palette[
                                                                                           'adjustment'])
                row_num += 1

                headers = ["Item", "Amount"]
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=header)
                    cell.font = header_font
                    cell.fill = adjustment_fill
                    cell.alignment = center_align
                    cell.border = border
                row_num += 1

                for field_name, value, color_type in adj_summary:
                    if value != Decimal('0.00'):
                        label = field_name.replace('_', ' ').title()
                        ws.cell(row=row_num, column=1, value=label).font = bold_font
                        ws.cell(row=row_num, column=1).alignment = left_align
                        ws.cell(row=row_num, column=1).border = border

                        amt_cell = ws.cell(row=row_num, column=2, value=float(value))
                        amt_cell.number_format = money_format
                        amt_cell.alignment = right_align
                        amt_cell.border = border

                        # Apply color coding
                        if color_type == 'primary':
                            amt_cell.fill = primary_fill
                            amt_cell.font = bold_font
                        elif color_type == 'success':
                            amt_cell.fill = success_fill
                            amt_cell.font = bold_font
                        else:
                            amt_cell.fill = neutral_fill
                            amt_cell.font = normal_font

                        row_num += 1
                row_num += 2

        # ======================
        # SEVERANCE SUMMARY - matches table-primary/success/danger
        # ======================
        if 'severance' in item and any(val != Decimal('0.00') for val in item['severance'].values()):
            # Section Header
            ws.cell(row=row_num, column=1, value="SEVERANCE SUMMARY").font = Font(bold=True, size=12,
                                                                                  color=color_palette['severance'])
            row_num += 1

            headers = ["Item", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = severance_fill
                cell.alignment = center_align
                cell.border = border
            row_num += 1

            severance_items = [
                ("taxable_gross", item['severance'].get('taxable_gross'), 'neutral'),
                ("gross", item['severance'].get('gross'), 'primary'),
                ("employment_income_tax", item['severance'].get('employment_income_tax'), 'neutral'),
                ("total_severance_deduction", item['severance'].get('total_severance_deduction'), 'neutral'),
                ("net", item['severance'].get('net'), 'success'),
                ("expense", item['severance'].get('expense'), 'neutral'),
            ]

            for field_name, value, color_type in severance_items:
                if value != Decimal('0.00'):
                    label = field_name.replace('_', ' ').title()
                    ws.cell(row=row_num, column=1, value=label).font = bold_font
                    ws.cell(row=row_num, column=1).alignment = left_align
                    ws.cell(row=row_num, column=1).border = border

                    amt_cell = ws.cell(row=row_num, column=2, value=float(value))
                    amt_cell.number_format = money_format
                    amt_cell.alignment = right_align
                    amt_cell.border = border

                    # Apply color coding
                    if color_type == 'primary':
                        amt_cell.fill = primary_fill
                        amt_cell.font = bold_font
                    elif color_type == 'success':
                        amt_cell.fill = success_fill
                        amt_cell.font = bold_font
                    else:
                        amt_cell.fill = neutral_fill
                        amt_cell.font = normal_font

                    row_num += 1
            row_num += 2

        # ======================
        # TOTAL SUMMARY - matches table-info/warning/success
        # ======================
        if 'totals' in item and any(val != Decimal('0.00') for val in item['totals'].values()):
            # Section Header
            ws.cell(row=row_num, column=1, value="TOTAL SUMMARY").font = Font(bold=True, size=12,
                                                                              color=color_palette['totals'])
            row_num += 1

            headers = ["Item", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = totals_fill
                cell.alignment = center_align
                cell.border = border
            row_num += 1

            summary_items = [
                ("taxable_gross", item['totals'].get('taxable_gross'), 'info'),
                ("non_taxable_gross", item['totals'].get('non_taxable_gross'), 'info'),
                ("gross", item['totals'].get('gross'), 'info'),
                ("pensionable", item['totals'].get('pensionable'), 'warning'),
                ("employee_pension", item['totals'].get('employee_pension'), 'warning'),
                ("employer_pension", item['totals'].get('employer_pension'), 'warning'),
                ("total_pension", item['totals'].get('total_pension'), 'warning'),
                ("employment_income_tax", item['totals'].get('employment_income_tax'), 'neutral'),
                ("total_deduction", item['totals'].get('total_deduction'), 'neutral'),
                ("final_net_pay", item['totals'].get('final_net_pay'), 'success'),
                ("expense", item['totals'].get('expense'), 'neutral'),
            ]

            for field_name, value, color_type in summary_items:
                if value != Decimal('0.00'):
                    label = field_name.replace('_', ' ').title()
                    ws.cell(row=row_num, column=1, value=label).font = bold_font
                    ws.cell(row=row_num, column=1).alignment = left_align
                    ws.cell(row=row_num, column=1).border = border

                    amt_cell = ws.cell(row=row_num, column=2, value=float(value))
                    amt_cell.number_format = money_format
                    amt_cell.alignment = right_align
                    amt_cell.border = border

                    # Apply color coding
                    if color_type == 'info':
                        amt_cell.fill = info_fill
                        amt_cell.font = bold_font
                    elif color_type == 'warning':
                        amt_cell.fill = warning_fill
                        amt_cell.font = bold_font
                    elif color_type == 'success':
                        amt_cell.fill = success_fill
                        amt_cell.font = bold_font
                    else:
                        amt_cell.fill = neutral_fill
                        amt_cell.font = normal_font

                    row_num += 1
            row_num += 3  # Extra space between months

    # ======================
    # FINAL FORMATTING
    # ======================
    # Adjust column widths
    column_max_widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                col_index = cell.column
                content_length = len(str(cell.value))
                column_max_widths[col_index] = max(column_max_widths.get(col_index, 0), content_length)

    for col_index, max_len in column_max_widths.items():
        col_letter = get_column_letter(col_index)
        if max_len <= 10:
            adjusted_width = max_len + 8
        elif max_len <= 20:
            adjusted_width = int(max_len * 1.2)
        else:
            adjusted_width = int(max_len * 1.4)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"monthly_payroll_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response




#graph combined monthly
@login_required
def combined_monthly_graph_view(request):
    data_dict = get_combined_monthly_detail(request)
    monthly_summary = data_dict.get("monthly_summary", {})

    # --- Sort months properly ---
    def parse_key(key):
        try:
            month_str, month_str = key.split('-')
            month = int(month_str)
            month_order = {
                "January": 1, "February": 2, "March": 3, "April": 4,
                "May": 5, "June": 6, "July": 7, "August": 8,
                "September": 9, "October": 10, "November": 11, "December": 12
            }
            month_num = month_order.get(month_str.capitalize(), 0)
            return (month, month_num)
        except:
            return (0, 0)

    sorted_summary = sorted(monthly_summary.items(), key=lambda x: parse_key(x[0]))

    # --- Prepare data for chart ---
    months = []
    gross_list = []
    net_list = []
    deduction_list = []
    expense_list = []

    for month_key, data in sorted_summary:
        months.append(month_key)
        totals = data['totals']
        gross_list.append(float(totals.get('gross', 0)))
        net_list.append(float(totals.get('final_net_pay', 0)))
        deduction_list.append(float(totals.get('total_deduction', 0)))
        expense_list.append(float(totals.get('expense', 0)))

    # --- Create traces ---
    traces = [
        go.Bar(name='Gross Pay', x=months, y=gross_list, marker_color='royalblue'),
        go.Bar(name='Net Pay', x=months, y=net_list, marker_color='green'),
        go.Bar(name='Total Deduction', x=months, y=deduction_list, marker_color='crimson'),
        go.Bar(name='Expense', x=months, y=expense_list, marker_color='orange'),

    ]

    layout = go.Layout(
        title='Combined Monthly Payroll Overview',
        barmode='group',
        xaxis=dict(title='Month'),
        yaxis=dict(title='Amount (ETB)'),
        template='plotly_white',
        margin=dict(l=40, r=40, t=60, b=60),
    )

    fig = go.Figure(data=traces, layout=layout)
    div = opy.plot(fig, auto_open=False, output_type='div')

    return render(request, 'combined_payroll/monthly_graph.html', {
        'graph_div': div
    })


    # monthly_summary = get_combined_monthly_detail(request).get("monthly_summary", {})
    #
    # # Order by "Month-YYYY"
    # def parse_key(key):
    #     try:
    #         month_str, month_str = key.split('-')
    #         month = int(month_str)
    #         month_order_full = {
    #             "January": 1, "February": 2, "March": 3, "April": 4,
    #             "May": 5, "June": 6, "July": 7, "August": 8,
    #             "September": 9, "October": 10, "November": 11, "December": 12
    #         }
    #         month_num = month_order_full.get(month_str.capitalize(), 0)
    #         return (month, month_num)
    #     except Exception:
    #         return (0, 0)
    #
    # sorted_summary = sorted(monthly_summary.items(), key=lambda x: parse_key(x[0]))
    #
    # # Components to include in the bar chart
    # components = [
    #     'taxable_gross',
    #     'non_taxable_gross',
    #     'gross',
    #     'pensionable',
    #     'employee_pension',
    #     'employer_pension',
    #     'total_pension',
    #     'employment_income_tax',
    #     'expense',
    #     'final_net_pay',
    # ]
    #
    # months = []
    # data_by_component = {comp: [] for comp in components}
    #
    # for month_key, data in sorted_summary:
    #     months.append(month_key)  # Use raw key like "July-2024"
    #     totals = data['totals']
    #     for comp in components:
    #         data_by_component[comp].append(float(totals.get(comp, 0)))
    #
    # # Create Bar chart traces
    # bar_traces = [
    #     go.Bar(x=months, y=data_by_component[comp], name=comp.replace('_', ' ').title())
    #     for comp in components
    # ]
    #
    # bar_layout = go.Layout(
    #     title='Monthly Payroll Summary By Components',
    #     barmode='group',
    #     xaxis=dict(title='Month'),
    #     yaxis=dict(title='Amount (ETB)'),
    #     template='plotly_white',
    #     margin=dict(l=40, r=40, t=60, b=60),
    # )
    #
    # bar_fig = go.Figure(data=bar_traces, layout=bar_layout)
    # bar_div = opy.plot(bar_fig, auto_open=False, output_type='div')
    #
    # # --- Pie chart for selected month ---
    # selected_month = request.GET.get('month')
    # if not selected_month and sorted_summary:
    #     selected_month = sorted_summary[-1][0]
    #
    # month_options = [key for key, _ in sorted_summary]
    # selected_data = dict(sorted_summary).get(selected_month)
    #
    # if selected_data:
    #     totals = selected_data['totals']
    #     gross = float(totals.get('gross', 0))
    #     net = float(totals.get('final_net_pay', 0))
    #     emp_pension = float(totals.get('employer_pension', 0))
    #     emp_deduction = gross - net
    #
    #     labels = [
    #         'Net Pay (Employee Take-home)',
    #         'Employee Deductions (Tax, Pension, etc.)',
    #         'Employer Pension Contribution',
    #     ]
    #     values = [net, emp_deduction, emp_pension]
    #
    #     if gross <= 0 or sum(values) <= 0:
    #         pie_div = "<p>Insufficient data for pie chart.</p>"
    #     else:
    #         pie_trace = go.Pie(
    #             labels=labels,
    #             values=values,
    #             hole=0.3,
    #             hoverinfo='label+percent+value',
    #             textinfo='label+percent'
    #         )
    #         pie_layout = go.Layout(
    #             title=f'Payroll Cost Breakdown for {selected_month}',
    #             template='plotly_white',
    #             margin=dict(l=40, r=40, t=60, b=60),
    #         )
    #         pie_fig = go.Figure(data=[pie_trace], layout=pie_layout)
    #         pie_div = opy.plot(pie_fig, auto_open=False, output_type='div')
    # else:
    #     pie_div = "<p>No data available for selected month.</p>"
    #
    # return render(request, 'combined_payroll/monthly_graph.html', {
    #     'month_options': month_options,
    #     'selected_month': selected_month,
    #     'pie_div': pie_div,
    #     'bar_div': bar_div,
    # })