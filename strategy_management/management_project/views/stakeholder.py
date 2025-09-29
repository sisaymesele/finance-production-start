# Standard library imports
from datetime import datetime

# Django imports
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Avg, Case, Count, IntegerField, Q, Value, When
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.core.paginator import Paginator

# Third-party imports
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Local project imports
from management_project.forms import StakeholderForm
from management_project.models import Department, Stakeholder


#
@login_required
def stakeholder_list(request):
    query = request.GET.get('search', '').strip()
    selected_type = request.GET.get('stakeholder_type', '').strip()
    page_number = request.GET.get('page', 1)

    # Base queryset: all stakeholders for this organization
    stakeholders = Stakeholder.objects.filter(
        organization_name=request.user.organization_name
    )

    # Apply exact filter by stakeholder type (internal/external)
    if selected_type:
        stakeholders = stakeholders.filter(stakeholder_type=selected_type)

    # Apply search filter across multiple fields
    if query:
        search_filter = (
            Q(stakeholder_name__icontains=query) |
            Q(role__icontains=query) |
            Q(department__name__icontains=query) |  # FK field now properly searched
            Q(description__icontains=query) |
            Q(contact_info__icontains=query)
        )
        stakeholders = stakeholders.filter(search_filter)

    # Ordering
    stakeholders = stakeholders.order_by('stakeholder_type', 'priority', 'stakeholder_name')

    # Provide stakeholder types for dropdown filter
    stakeholder_types = [choice[0] for choice in Stakeholder.STAKEHOLDER_TYPE_CHOICES]

    # Pagination
    paginator = Paginator(stakeholders, 10)
    page_obj = paginator.get_page(page_number)

    return render(request, 'stakeholder_list/list.html', {
        'stakeholders': page_obj,
        'page_obj': page_obj,
        'search_query': query,
        'stakeholder_types': stakeholder_types,
        'selected_type': selected_type,
    })


@login_required
def create_stakeholder(request):
    next_url = request.GET.get('next')
    if request.method == 'POST':
        form = StakeholderForm(request.POST, request=request)  # Pass request
        if form.is_valid():
            stakeholder = form.save(commit=False)
            stakeholder.organization_name = request.user.organization_name
            stakeholder.save()
            messages.success(request, "Stakeholder created successfully!")
            return redirect(next_url or 'stakeholder_list')
    else:
        form = StakeholderForm(request=request)  # Pass request

    context = {
        'form': form,
        'form_title': 'Create Stakeholder',
        'submit_button_text': 'Create Stakeholder',
        'back_url': next_url or reverse('stakeholder_list'),
    }
    return render(request, 'stakeholder_list/form.html', context)


@login_required
def update_stakeholder(request, pk):
    stakeholder = get_object_or_404(Stakeholder, pk=pk, organization_name=request.user.organization_name)

    if request.method == 'POST':
        form = StakeholderForm(request.POST, instance=stakeholder, request=request)  # Pass request
        if form.is_valid():
            form.save()
            messages.success(request, "Stakeholder updated successfully!")
            return redirect('stakeholder_list')
    else:
        form = StakeholderForm(instance=stakeholder, request=request)  # Pass request

    context = {
        'form': form,
        'form_title': 'Update Stakeholder',
        'submit_button_text': 'Update Stakeholder',
        'back_url': request.GET.get('next', reverse('stakeholder_list')),
        'edit_mode': True,
        'editing_stakeholder': stakeholder,
    }
    return render(request, 'stakeholder_list/form.html', context)


@login_required
def delete_stakeholder(request, pk):
    stakeholder = get_object_or_404(Stakeholder, pk=pk, organization_name=request.user.organization_name)

    if request.method == 'POST':
        stakeholder.delete()
        messages.success(request, "Stakeholder deleted successfully!")
        return redirect('stakeholder_list')

    return render(request, 'stakeholder_list/delete_confirm.html', {'stakeholder': stakeholder})

@login_required
def export_stakeholders_to_excel(request):
    query = request.GET.get('search', '').strip()
    selected_type = request.GET.get('stakeholder_type', '').strip()

    stakeholders = Stakeholder.objects.filter(
        organization_name=request.user.organization_name
    )

    if selected_type:
        stakeholders = stakeholders.filter(stakeholder_type=selected_type)

    if query:
        search_filter = (
            Q(stakeholder_name__icontains=query) |
            Q(role__icontains=query) |
            Q(department__name__icontains=query) |
            Q(description__icontains=query) |
            Q(notes__icontains=query) |
            Q(contact_info__icontains=query)
        )
        stakeholders = stakeholders.filter(search_filter)

    # Fields to include
    excluded_fields = {'slug', 'depends_on'}
    field_names = [
        field.name for field in Stakeholder._meta.get_fields()
        if not field.is_relation or isinstance(field, models.ForeignKey)
        if field.name not in excluded_fields
    ]

    headers = [field.replace('_', ' ').title() for field in field_names]

    # Workbook setup
    wb = Workbook()
    ws = wb.active
    ws.title = "Stakeholders"

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value="Stakeholder List")
    title_cell.font = Font(size=14, bold=True, color="FFFFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(start_color="FF305496", end_color="FF305496", fill_type="solid")

    # Headers
    ws.append(headers)
    for cell in ws[2]:
        cell.fill = PatternFill(start_color="FF0070C0", end_color="FF0070C0", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    for s in stakeholders:
        row = []
        for field in field_names:
            value = getattr(s, field, "")
            # Handle FK field (department) and None
            if isinstance(getattr(Stakeholder, field, None), models.ForeignKey):
                value = getattr(value, 'name', "") if value else ""
            elif isinstance(value, datetime):
                value = value.date()
            elif value is None:
                value = ""
            row.append(value)
        ws.append(row)

    # Column widths
    for i, column_cells in enumerate(ws.columns, 1):
        max_length = max((len(str(cell.value)) for cell in column_cells if cell.value), default=0) + 2
        ws.column_dimensions[get_column_letter(i)].width = max(max_length, 12)

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename=stakeholders.xlsx'
    wb.save(response)
    return response


# Model imports

@login_required
def stakeholder_graph_view(request):
    qs = Stakeholder.objects.filter(organization_name=request.user.organization_name)

    # Color palette for consistent styling
    color_palette = ["#1f77b4", "#ff7f0e", "#2ca02c", "#d62728", "#9467bd",
                     "#8c564b", "#e377c2", "#7f7f7f", "#bcbd22", "#17becf"]

    # ------------------ ROLE DISTRIBUTION ------------------
    role_data = {}
    for stakeholder in qs:
        for role in stakeholder.role or ['None']:
            label = role.replace('_', ' ').title() if role != 'other' else 'Other'
            role_data[label] = role_data.get(label, 0) + 1

    role_labels = list(role_data.keys())
    role_values = list(role_data.values())

    fig_role = go.Figure(data=[go.Bar(
        x=role_labels,
        y=role_values,
        marker_color=color_palette[:len(role_labels)],
        text=role_values,
        textposition='auto'
    )])
    fig_role.update_layout(
        title='Stakeholder Distribution by Role',
        xaxis_title='Role',
        yaxis_title='Count',
        template='plotly_white',
        height=500
    )

    # ------------------ DEPARTMENT DISTRIBUTION ------------------
    dept_counts = qs.values('department').annotate(
        count=Count('department'),
        avg_impact=Avg(Case(
            When(impact_level='very_low', then=1),
            When(impact_level='low', then=2),
            When(impact_level='medium', then=3),
            When(impact_level='high', then=4),
            When(impact_level='critical', then=5),
            default=3,
            output_field=IntegerField()
        )),
        avg_satisfaction=Avg(Case(
            When(satisfaction_level='very_low', then=1),
            When(satisfaction_level='low', then=2),
            When(satisfaction_level='medium', then=3),
            When(satisfaction_level='high', then=4),
            When(satisfaction_level='very_high', then=5),
            default=3,
            output_field=IntegerField()
        ))
    )

    dept_labels = [
        Department.objects.filter(pk=d['department']).first().department_name
        if d['department'] else 'Unknown'
        for d in dept_counts
    ]
    dept_values = [d['count'] for d in dept_counts]

    fig_dept = go.Figure(data=[go.Bar(
        x=dept_labels,
        y=dept_values,
        marker_color=color_palette[:len(dept_labels)],
        text=dept_values,
        textposition='auto',
        hovertemplate='<b>%{x}</b><br>Count: %{y}<br>Avg Impact: %{customdata[0]:.1f}<br>Avg Satisfaction: %{customdata[1]:.1f}<extra></extra>',
        customdata=[[d['avg_impact'] or 0, d['avg_satisfaction'] or 0] for d in dept_counts]
    )])
    fig_dept.update_layout(
        title='Stakeholders by Department',
        xaxis_title='Department',
        yaxis_title='Count',
        template='plotly_white',
        height=500
    )

    # ------------------ STAKEHOLDER TYPE DISTRIBUTION ------------------
    type_metrics = qs.values('stakeholder_type').annotate(
        count=Count('id'),
        avg_risk=Avg(Case(
            When(risk_level='very_low', then=1),
            When(risk_level='low', then=2),
            When(risk_level='medium', then=3),
            When(risk_level='high', then=4),
            When(risk_level='critical', then=5),
            default=3,
            output_field=IntegerField()
        ))
    )

    type_labels = [t['stakeholder_type'].title() for t in type_metrics]
    type_values = [t['count'] for t in type_metrics]
    type_risk = [t['avg_risk'] or 0 for t in type_metrics]

    fig_type = go.Figure(data=[go.Pie(
        labels=type_labels,
        values=type_values,
        hole=0.4,
        marker_colors=color_palette[:len(type_labels)],
        hovertemplate='<b>%{label}</b><br>Count: %{value}<br>Avg Risk: %{customdata:.1f}<extra></extra>',
        customdata=type_risk
    )])
    fig_type.update_layout(
        title='Stakeholder Distribution by Type<br><sub>Hover for average risk levels</sub>',
        height=500
    )

    # ------------------ IMPACT LEVEL ------------------
    impact_counts = qs.values('impact_level').annotate(count=Count('id')).order_by('impact_level')
    impact_labels = [item['impact_level'].replace('_', ' ').title() for item in impact_counts]
    impact_values = [item['count'] for item in impact_counts]
    impact_colors = ['#ff9999', '#ffcc99', '#ffff99', '#99ff99', '#66b366']

    fig_impact = go.Figure(data=[go.Bar(
        x=impact_labels,
        y=impact_values,
        marker_color=impact_colors[:len(impact_labels)],
        text=impact_values,
        textposition='auto',
        hovertemplate='<b>%{x}</b><br>Count: %{y}<extra></extra>'
    )])
    fig_impact.update_layout(
        title='Stakeholder Impact Level Distribution',
        xaxis_title='Impact Level',
        yaxis_title='Count',
        template='plotly_white',
        height=500
    )

    # ------------------ SATISFACTION LEVEL ------------------
    satisfaction_counts = qs.values('satisfaction_level').annotate(count=Count('id')).order_by('satisfaction_level')
    satisfaction_labels = [item['satisfaction_level'].replace('_', ' ').title() for item in satisfaction_counts]
    satisfaction_values = [item['count'] for item in satisfaction_counts]
    satisfaction_colors = ['#ff6666', '#ff9999', '#ffff99', '#99ff99', '#33cc33']

    fig_satisfaction = go.Figure(data=[go.Bar(
        x=satisfaction_labels,
        y=satisfaction_values,
        marker_color=satisfaction_colors[:len(satisfaction_labels)],
        text=satisfaction_values,
        textposition='auto',
        hovertemplate='<b>%{x}</b><br>Count: %{y}<extra></extra>'
    )])
    fig_satisfaction.update_layout(
        title='Stakeholder Satisfaction Level Distribution',
        xaxis_title='Satisfaction Level',
        yaxis_title='Count',
        template='plotly_white',
        height=500
    )

    # ------------------ ENGAGEMENT STRATEGY ------------------
    satisfaction_mapping = {'very_low': 1, 'low': 2, 'medium': 3, 'high': 4, 'very_high': 5}
    engagement_data = {}

    for stakeholder in qs:
        satisfaction_value = satisfaction_mapping.get(stakeholder.satisfaction_level, 3)
        for strategy in stakeholder.engagement_strategy or ['None']:
            label = strategy.replace('_', ' ').title() if strategy != 'other' else 'Other'
            if label not in engagement_data:
                engagement_data[label] = {'count': 0, 'total_satisfaction': 0}
            engagement_data[label]['count'] += 1
            engagement_data[label]['total_satisfaction'] += satisfaction_value

    engagement_labels = list(engagement_data.keys())
    engagement_counts = [data['count'] for data in engagement_data.values()]
    engagement_avg_satisfaction = [data['total_satisfaction'] / data['count'] for data in engagement_data.values()]

    fig_engagement = go.Figure()
    fig_engagement.add_trace(go.Bar(
        name='Count',
        x=engagement_labels,
        y=engagement_counts,
        marker_color=color_palette[0],
        yaxis='y',
        offsetgroup=1
    ))
    fig_engagement.add_trace(go.Scatter(
        name='Avg Satisfaction',
        x=engagement_labels,
        y=engagement_avg_satisfaction,
        marker_color=color_palette[1],
        yaxis='y2',
        mode='lines+markers',
        line=dict(width=3)
    ))
    fig_engagement.update_layout(
        title='Engagement Strategy Analysis',
        xaxis_title='Engagement Strategy',
        yaxis=dict(title='Count', side='left'),
        yaxis2=dict(title='Average Satisfaction', side='right', overlaying='y'),
        template='plotly_white',
        height=500
    )

    # ------------------ SUMMARY DATA ------------------
    total_stakeholders = qs.count()
    internal_count = qs.filter(stakeholder_type='internal').count()
    external_count = qs.filter(stakeholder_type='external').count()

    avg_impact = qs.aggregate(avg=Avg(Case(
        When(impact_level='very_low', then=1),
        When(impact_level='low', then=2),
        When(impact_level='medium', then=3),
        When(impact_level='high', then=4),
        When(impact_level='critical', then=5),
        default=3,
        output_field=IntegerField()
    )))['avg'] or 0

    avg_satisfaction = qs.aggregate(avg=Avg(Case(
        When(satisfaction_level='very_low', then=1),
        When(satisfaction_level='low', then=2),
        When(satisfaction_level='medium', then=3),
        When(satisfaction_level='high', then=4),
        When(satisfaction_level='very_high', then=5),
        default=3,
        output_field=IntegerField()
    )))['avg'] or 0

    high_risk_count = qs.filter(risk_level__in=['high', 'critical']).count()
    high_priority_count = qs.filter(priority__in=['high', 'critical']).count()

    summary_data = {
        'total_stakeholders': total_stakeholders,
        'internal_count': internal_count,
        'external_count': external_count,
        'avg_impact': round(avg_impact, 1),
        'avg_satisfaction': round(avg_satisfaction, 1),
        'high_risk_count': high_risk_count,
        'high_priority_count': high_priority_count,
        'high_risk_percentage': round((high_risk_count / total_stakeholders * 100) if total_stakeholders > 0 else 0, 1),
    }

    return render(request, 'stakeholder_list/graph.html', {
        'plot_html_role': fig_role.to_html(full_html=False),
        'plot_html_department': fig_dept.to_html(full_html=False),
        'plot_html_type': fig_type.to_html(full_html=False),
        'plot_html_impact': fig_impact.to_html(full_html=False),
        'plot_html_satisfaction': fig_satisfaction.to_html(full_html=False),
        'plot_html_engagement': fig_engagement.to_html(full_html=False),
        'summary_data': summary_data,
    })
