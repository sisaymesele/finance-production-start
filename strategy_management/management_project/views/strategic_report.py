from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.core.paginator import Paginator
from django.urls import reverse
from django.db.models import Q
import calendar

from django.contrib.auth.decorators import login_required
from management_project.models import StrategicReport, StrategicActionPlan
from management_project.forms import StrategicReportForm

from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from management_project.models import StrategicReport
#chart

from django.shortcuts import render
from django.db.models import Avg, Count
from management_project.models import StrategicReport, StrategicCycle
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from plotly.offline import plot
from collections import defaultdict
from datetime import datetime
from plotly.colors import qualitative


@login_required
def strategy_report_by_cycle_list(request):
    """List distinct strategic cycles for the current organization with all info."""
    cycles_qs = StrategicCycle.objects.filter(
        organization_name=request.user.organization_name
    ).order_by('-start_date')

    # Build a list of dicts including calculated properties
    cycles = []
    for cycle in cycles_qs:
        cycles.append({
            'name': cycle.name,
            'time_horizon': cycle.time_horizon,
            'time_horizon_type': cycle.time_horizon_type,
            'start_date': cycle.start_date,
            'end_date': cycle.end_date,
            'slug': cycle.slug,
            'duration_days': (cycle.end_date - cycle.start_date).days if cycle.start_date and cycle.end_date else None,
            'start_month_name': calendar.month_name[cycle.start_date.month] if cycle.start_date else None,
            'start_quarter': (cycle.start_date.month - 1) // 3 + 1 if cycle.start_date else None,
            'start_year': cycle.start_date.year if cycle.start_date else None,
        })

    return render(request, 'strategic_report/cycle_list.html', {
        'strategic_cycles': cycles
    })


@login_required
def strategic_report_list(request, cycle_slug):
    """List all strategic reports by cycle, grouped by action plan."""
    strategy_by_cycle = get_object_or_404(
        StrategicCycle,
        slug=cycle_slug,
        organization_name=request.user.organization_name
    )

    # Base queryset
    reports = StrategicReport.objects.filter(
        action_plan__strategic_cycle=strategy_by_cycle,
        organization_name=request.user.organization_name
    ).select_related("action_plan").order_by("-id")

    # Search query
    search_query = request.GET.get("search", "").strip()
    if search_query:
        reports = reports.filter(
            Q(action_plan__strategy_hierarchy__strategic_perspective__icontains=search_query) |
            Q(action_plan__strategy_hierarchy__focus_area__icontains=search_query) |
            Q(action_plan__strategy_hierarchy__objective__icontains=search_query) |
            Q(action_plan__strategy_hierarchy__kpi__icontains=search_query) |
            Q(achievement__icontains=search_query) |
            Q(status__icontains=search_query)
        ).distinct()

    # Pagination
    paginator = Paginator(reports, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Form for creating new report
    form = StrategicReportForm(request=request, cycle=strategy_by_cycle)

    context = {
        "strategy_by_cycle": strategy_by_cycle,
        "page_obj": page_obj,
        "form": form,
        "search_query": search_query,
    }
    return render(request, "strategic_report/list_by_cycle.html", context)

@login_required
def strategic_report_detail(request, cycle_slug, pk):
    # Get the strategic cycle by slug
    strategy_by_cycle = get_object_or_404(StrategicCycle, slug=cycle_slug)

    # Get the specific strategic report for this cycle and organization
    strategic_report = get_object_or_404(
        StrategicReport,
        pk=pk,
        action_plan__strategic_cycle=strategy_by_cycle,
        organization_name=request.user.organization_name
    )

    return render(request, 'strategic_report/detail.html', {
        'strategy_by_cycle': strategy_by_cycle,
        'strategic_report': strategic_report
    })


@login_required
def create_strategic_report(request, cycle_slug):
    strategy_by_cycle = get_object_or_404(StrategicCycle, slug=cycle_slug)

    if request.method == "POST":
        form = StrategicReportForm(request.POST, request=request, cycle=strategy_by_cycle)
        if form.is_valid():
            report = form.save(commit=False)
            report.organization_name = request.user.organization_name
            report.save()
            form.save_m2m()
            messages.success(request, "Strategic report created successfully!")
            return redirect("strategic_report_list", cycle_slug=strategy_by_cycle.slug)
    else:
        form = StrategicReportForm(request=request, cycle=strategy_by_cycle)

    return render(request, "strategic_report/form.html", {
        "form": form,
        "form_title": f"Create Strategic Report for {strategy_by_cycle.name}",
        "submit_button_text": "Create Strategic Report",
        "back_url": reverse("strategic_report_list", kwargs={"cycle_slug": strategy_by_cycle.slug}),
        "strategy_by_cycle": strategy_by_cycle,
    })


@login_required
def update_strategic_report(request, cycle_slug, pk):
    strategy_by_cycle = get_object_or_404(StrategicCycle, slug=cycle_slug)
    report = get_object_or_404(
        StrategicReport,
        pk=pk,
        action_plan__strategic_cycle=strategy_by_cycle,
        organization_name=request.user.organization_name,
    )

    if request.method == "POST":
        form = StrategicReportForm(request.POST, instance=report, request=request, cycle=strategy_by_cycle)
        if form.is_valid():
            form.save()
            messages.success(request, "Strategic report updated successfully!")
            return redirect("strategic_report_list", cycle_slug=strategy_by_cycle.slug)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = StrategicReportForm(instance=report, request=request, cycle=strategy_by_cycle)

    return render(request, "strategic_report/form.html", {
        "form": form,
        "form_title": f"Update Strategic Report for {strategy_by_cycle.name}",
        "submit_button_text": "Update Strategic Report",
        "back_url": reverse("strategic_report_list", kwargs={"cycle_slug": strategy_by_cycle.slug}),
        "strategy_by_cycle": strategy_by_cycle,
        "edit_strategic_report": report,
    })


@login_required
def delete_strategic_report(request, cycle_slug, pk):
    strategy_by_cycle = get_object_or_404(StrategicCycle, slug=cycle_slug)
    report = get_object_or_404(
        StrategicReport,
        pk=pk,
        action_plan__strategic_cycle=strategy_by_cycle,
        organization_name=request.user.organization_name,
    )

    if request.method == "POST":
        report.delete()
        messages.success(request, "Strategic report deleted successfully!")
        return redirect("strategic_report_list", cycle_slug=strategy_by_cycle.slug)

    return render(request, "strategic_report/delete_confirm.html", {
        "strategic_report": report,
        "strategy_by_cycle": strategy_by_cycle,
    })

#

#export
def split_two_lines(text):
    """Split a long text roughly into two lines."""
    if not text:
        return ""
    words = text.split()
    mid = len(words) // 2
    return " ".join(words[:mid]) + "\n" + " ".join(words[mid:])


@login_required
def export_strategic_report_to_excel(request, cycle_slug):
    """Export strategic reports to Excel for a specific cycle."""
    cycle = get_object_or_404(StrategicCycle, slug=cycle_slug, organization_name=request.user.organization_name)
    reports = StrategicReport.objects.filter(
        action_plan__strategic_cycle=cycle,
        organization_name=request.user.organization_name
    ).select_related("action_plan", "action_plan__strategy_hierarchy", "action_plan__strategic_cycle")

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Strategic Reports {cycle.name}"

    # Title row
    title_text = f"Strategic Reports For: {cycle.name}"
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=22)
    title_cell = sheet.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(bold=True, size=14, color="FFFFFF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    title_cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    title_cell.border = Border(
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin")
    )

    # Column headers
    headers = [
        "#", "Perspective", "Pillar", "Objective", "KPI",
        "Baseline", "Target", "Achievement", "Percent Achieved",
        "Variance", "Weighted Score", "Responsible Bodies",
        "Data Source", "Data Collector", "Progress Summary",
        "Performance Summary", "Status", "Created At", "Update At"
    ]

    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=2, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill(start_color="4BACC6", end_color="4BACC6", fill_type="solid")
        cell.border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

    # Data rows
    for row_idx, report in enumerate(reports, start=3):
        action_plan = report.action_plan
        strategy_hierarchy = action_plan.strategy_hierarchy if action_plan else None
        strategic_cycle = action_plan.strategic_cycle if action_plan else None

        data = [
            row_idx - 2,
            strategy_hierarchy.strategic_perspective if strategy_hierarchy else "",
            split_two_lines(strategy_hierarchy.focus_area if strategy_hierarchy else ""),
            split_two_lines(strategy_hierarchy.objective if strategy_hierarchy else ""),
            split_two_lines(strategy_hierarchy.kpi if strategy_hierarchy else ""),
            action_plan.baseline if action_plan else 0,
            action_plan.target if action_plan else 0,
            report.achievement,
            report.percent_achieved,
            report.variance,
            report.weighted_score,
            ", ".join([body.stakeholder_name for body in action_plan.responsible_bodies.all()]) if action_plan else "",
            report.data_source or "",
            report.data_collector or "",
            split_two_lines(report.progress_summary or ""),
            split_two_lines(report.performance_summary or ""),
            report.get_status_display(),
            report.created_at.strftime('%B %d, %Y %H:%M') if report.created_at else "",
            report.updated_at.strftime('%B %d, %Y %H:%M') if report.updated_at else ""

        ]

        for col_num, value in enumerate(data, start=1):
            cell = sheet.cell(row=row_idx, column=col_num, value=value)
            # Wrap text for certain fields
            if col_num in [3, 4, 5, 12, 15, 16]:  # Text-heavy columns
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            # Borders
            cell.border = Border(
                left=Side(border_style="thin"),
                right=Side(border_style="thin"),
                top=Side(border_style="thin"),
                bottom=Side(border_style="thin")
            )

    # Column widths
    min_width = 10
    max_width = 25
    for col_idx in range(1, sheet.max_column + 1):
        width = min_width
        for row_idx in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if cell.value:
                width = max(width, min(len(str(cell.value)) + 2, max_width))
        sheet.column_dimensions[get_column_letter(col_idx)].width = width

    # Save to buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Strategic_Reports_{cycle.name}.xlsx"'
    return response


# views.py

def strategic_report_chart(request):
    user = request.user  # logged-in user

    # ---------------- Filters ----------------
    cycle_filter = request.GET.get("strategic_cycle", "all")
    body_filter = request.GET.get("responsible_body", "all")
    horizon_filter = request.GET.get("time_horizon", "all")

    # Filter reports for the logged-in user
    reports = StrategicReport.objects.select_related(
        "action_plan",
        "action_plan__strategic_cycle",
        "organization_name"
    ).prefetch_related("action_plan__responsible_bodies").filter(organization_name=request.user.organization_name)

    if cycle_filter != "all":
        reports = reports.filter(action_plan__strategic_cycle_id=cycle_filter)
    if body_filter != "all":
        reports = reports.filter(action_plan__responsible_bodies__stakeholder_name=body_filter)
    if horizon_filter != "all":
        reports = reports.filter(action_plan__strategic_cycle__time_horizon_type=horizon_filter)

    # ---------------- 1. Summary Stats ----------------
    total_reports = reports.count()
    avg_percent_achieved = reports.aggregate(avg=Avg("percent_achieved"))["avg"] or 0
    avg_variance = reports.aggregate(avg=Avg("variance"))["avg"] or 0
    avg_weighted_score = reports.aggregate(avg=Avg("weighted_score"))["avg"] or 0

    # ---------------- 2. Status Pie Chart ----------------
    status_counts = reports.values("status").annotate(count=Count("id"))
    status_map = {
        "pending": {"name": "Pending", "color": "#FFCE56"},
        "in_progress": {"name": "In Progress", "color": "#36A2EB"},
        "completed": {"name": "Completed", "color": "#4BC0C0"},
        "on_hold": {"name": "On Hold", "color": "#FF6384"},
        "cancelled": {"name": "Cancelled", "color": "#9966FF"},
    }
    if status_counts:
        status_fig = go.Figure(
            data=[
                go.Pie(
                    labels=[status_map[i["status"]]["name"] for i in status_counts],
                    values=[i["count"] for i in status_counts],
                    hole=0.3,
                    marker=dict(colors=[status_map[i["status"]]["color"] for i in status_counts]),
                    textinfo="percent+label",
                )
            ]
        )
        status_fig.update_layout(title_text="Report Status Distribution")
        status_plot = plot(status_fig, output_type="div", config={"displayModeBar": False})
    else:
        status_plot = "<div class='no-data'>No status data available</div>"

    # ---------------- 3. Cycle-Level Metrics ----------------
    cycle_data = (
        reports.values("action_plan__strategic_cycle__end_date")
        .annotate(
            avg_percent_achieved=Avg("percent_achieved"),
            avg_variance=Avg("variance"),
            avg_weighted_score=Avg("weighted_score"),
        )
        .order_by("action_plan__strategic_cycle__end_date")
    )
    cycle_dates = [
        c["action_plan__strategic_cycle__end_date"].strftime("%B %Y") if c["action_plan__strategic_cycle__end_date"] else "N/A"
        for c in cycle_data
    ]
    cycle_fig = go.Figure()
    cycle_fig.add_trace(go.Bar(
        name="Percent Achieved",
        x=cycle_dates,
        y=[c["avg_percent_achieved"] or 0 for c in cycle_data],
        marker_color="#4BC0C0"
    ))
    cycle_fig.add_trace(go.Bar(
        name="Variance",
        x=cycle_dates,
        y=[c["avg_variance"] or 0 for c in cycle_data],
        marker_color="#FF6384"
    ))
    cycle_fig.add_trace(go.Bar(
        name="Weighted Score",
        x=cycle_dates,
        y=[c["avg_weighted_score"] or 0 for c in cycle_data],
        marker_color="#36A2EB"
    ))
    cycle_fig.update_layout(
        title="Cycle-Level Averages",
        barmode="group",
        xaxis_title="Cycle Months",
        yaxis_title="Value",
        xaxis_tickangle=-45,
        template="plotly_white"
    )
    cycle_plot = plot(cycle_fig, output_type="div", config={"displayModeBar": False})

    # ---------------- 4. Body + Cycle Metrics ----------------

    kpi_data = reports.values(
        "action_plan__strategy_hierarchy__kpi",
        "action_plan__strategic_cycle__end_date",
        "achievement"
    ).order_by("action_plan__strategic_cycle__end_date")

    kpi_names = sorted({r['action_plan__strategy_hierarchy__kpi'] for r in kpi_data if r['action_plan__strategy_hierarchy__kpi']})
    cycle_dates_set = sorted({r['action_plan__strategic_cycle__end_date'] for r in kpi_data if r['action_plan__strategic_cycle__end_date']})
    cycle_labels = [d.strftime("%b %Y") for d in cycle_dates_set]

    kpi_matrix = defaultdict(lambda: {d: 0 for d in cycle_dates_set})
    for r in kpi_data:
        kpi = r['action_plan__strategy_hierarchy__kpi']
        date = r['action_plan__strategic_cycle__end_date']
        if kpi and date:
            kpi_matrix[kpi][date] = r['achievement'] or 0

    kpi_fig = go.Figure()
    palette = qualitative.Set2
    for i, kpi in enumerate(kpi_names):
        y_vals = [kpi_matrix[kpi][d] for d in cycle_dates_set]
        kpi_fig.add_trace(go.Bar(
            x=cycle_labels,
            y=y_vals,
            name=kpi,
            marker_color=palette[i % len(palette)],
            text=y_vals,
            textposition="auto",
            hovertemplate=f"KPI: {kpi}<br>Date: %{{x}}<br>Achievement: %{{y}}<extra></extra>"
        ))

    kpi_fig.update_layout(
        title="KPI Achievement by Cycle",
        xaxis_title="Cycle Months",
        yaxis_title="Achievement",
        yaxis=dict(type="category", categoryorder="array", categoryarray=kpi_names[::-1]),
        barmode="group",
        height=600,
        legend_title_text="KPIs",
        xaxis_tickangle=-45,
        template="plotly_white",
        margin=dict(l=150, r=50, t=80, b=100)
    )
    kpi_plot = plot(kpi_fig, output_type="div", config={"displayModeBar": False})

    # ---------------- Context ----------------
    context = {
        "total_reports": total_reports,
        "avg_percent_achieved": avg_percent_achieved,
        "avg_variance": avg_variance,
        "avg_weighted_score": avg_weighted_score,
        "status_plot": status_plot,
        "cycle_plot": cycle_plot,
        "kpi_plot": kpi_plot,
        "strategic_cycles": StrategicCycle.objects.all(),
        "responsible_bodies": sorted({
            b.stakeholder_name
            for r in StrategicReport.objects.filter(organization_name=request.user.organization_name).prefetch_related("action_plan__responsible_bodies")
            for b in r.action_plan.responsible_bodies.all()
        }),
        "time_horizons": sorted({
            r.action_plan.strategic_cycle.time_horizon_type
            for r in StrategicReport.objects.filter(organization_name=request.user.organization_name).select_related("action_plan__strategic_cycle")
            if r.action_plan.strategic_cycle and r.action_plan.strategic_cycle.time_horizon_type
        }),
        "selected_cycle": cycle_filter,
        "selected_body": body_filter,
        "selected_horizon": horizon_filter,
        "current_date": datetime.now(),
    }

    return render(request, "strategic_report/chart.html", context)








