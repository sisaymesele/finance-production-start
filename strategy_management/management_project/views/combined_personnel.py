from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from openpyxl import Workbook
from decimal import Decimal
from management_project.services.combined.personnel_context import get_combined_personnel_payroll_context
from openpyxl.utils import get_column_letter
import openpyxl
from io import BytesIO
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from management_project.services.excel_export import ExportUtilityService


@login_required
def combined_personnel_detail(request):
    context = get_combined_personnel_payroll_context(request)
    return render(request, 'combined_payroll/personnel/detail.html', context)

@login_required
def combined_personnel_adjustment_list(request):
    context = get_combined_personnel_payroll_context(request)
    return render(request, 'combined_payroll/personnel/adjustment_list.html', context)

@login_required
def combined_personnel_payroll_list(request):
    context = get_combined_personnel_payroll_context(request)
    return render(request, 'combined_payroll/personnel/payroll_list.html', context)

@login_required
def combined_personnel_total(request):
    context = get_combined_personnel_payroll_context(request)
    return render(request, 'combined_payroll/personnel/total_list.html', context)

@login_required
def combined_personnel_expense(request):
    context = get_combined_personnel_payroll_context(request)
    return render(request, 'combined_payroll/personnel/expense_list.html', context)

@login_required
def combined_personnel_net_income(request):
    context = get_combined_personnel_payroll_context(request)
    return render(request, 'combined_payroll/personnel/net_income_list.html', context)


@login_required
def combined_personnel_employment_income_tax(request):
    context = get_combined_personnel_payroll_context(request)
    return render(request, 'combined_payroll/personnel/employment_income_tax_list.html', context)


@login_required
def combined_employee_pension(request):
    context = get_combined_personnel_payroll_context(request)
    return render(request, 'combined_payroll/personnel/pension_list.html', context)

#detail
def export_combined_personnel_detail(request):

    # Get the same data as the combined view
    context = get_combined_personnel_payroll_context(request)
    payroll_data = context['payroll_data']

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Combined Personnel Payroll"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'))

    # Use built-in number format instead of numbers.FORMAT_NUMBER_COMMA_SEPARATED2
    money_format = '#,##0.00'

    row_num = 1

    for item in payroll_data:
        # Header with merged cells
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=7)
        header_cell = ws.cell(row=row_num, column=1,
                              value=f"Combined Personnel Payslip For {item['payroll'].personnel_full_name} | {item['payroll'].payroll_month}")

        header_cell.font = Font(bold=True, size=14)
        header_cell.alignment = Alignment(horizontal="center")
        row_num += 2

        # Regular Payroll Section
        ws.cell(row=row_num, column=1, value="Regular Payroll").font = Font(bold=True, color="0070C0")
        row_num += 1

        # Regular Payroll Headers
        headers = ["Component", "Amount"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        row_num += 1

        # Regular Payroll Data
        for component, amount in item['regular_item_by_component'].items():
            if amount:
                ws.cell(row=row_num, column=1, value=component)
                amount_cell = ws.cell(row=row_num, column=2, value=float(amount))
                amount_cell.number_format = money_format
                row_num += 1

        row_num += 1

        # Earning Adjustment Section (if exists)
        if item['show_earning']:
            ws.cell(row=row_num, column=1, value="Earning Adjustment").font = Font(bold=True, color="00B050")
            row_num += 1

            # Earning Adjustment Headers
            headers = ["Component", "Total", "Taxable", "Non-Taxable",
                       "Employee Pension", "Employer Pension", "Total Pension Contribution"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            row_num += 1

            # Earning Adjustment Data
            for component, amounts in item['earning_adj_by_component'].items():
                if amounts['earning_amount']:
                    ws.cell(row=row_num, column=1, value=component)
                    ws.cell(row=row_num, column=2, value=float(amounts['earning_amount'])).number_format = money_format
                    ws.cell(row=row_num, column=3, value=float(amounts['taxable'])).number_format = money_format
                    ws.cell(row=row_num, column=4, value=float(amounts['non_taxable'])).number_format = money_format
                    ws.cell(row=row_num, column=5,
                            value=float(amounts['employee_pension_contribution'])).number_format = money_format
                    ws.cell(row=row_num, column=6,
                            value=float(amounts['employer_pension_contribution'])).number_format = money_format
                    ws.cell(row=row_num, column=7, value=float(amounts['total_pension'])).number_format = money_format
                    row_num += 1

            row_num += 1

            # Individual Adjustment Summary (like in HTML)
            ws.cell(row=row_num, column=1, value="Earning Adjustment Income tax").font = Font(bold=True, color="7030A0")
            row_num += 1

            adjustment_summary = [
                ("Employment Income Tax", item['earning_adjustment_item']['employment_income_tax']),
            ]

            for label, value in adjustment_summary:
                ws.cell(row=row_num, column=1, value=label)
                cell = ws.cell(row=row_num, column=2, value=float(value))
                cell.number_format = money_format
                row_num += 1

            row_num += 1


        # Deduction Adjustment Section (if exists)
        if item['show_deduction']:
            ws.cell(row=row_num, column=1, value="Deduction Adjustment").font = Font(bold=True, color="FF0000")
            row_num += 1

            # Deduction Adjustment Headers
            headers = ["Component", "Amount"]
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            row_num += 1

            # Deduction Adjustment Data
            for component, amount in item['deduction_adj_by_component'].items():
                if amount:
                    ws.cell(row=row_num, column=1, value=component)
                    amount_cell = ws.cell(row=row_num, column=2, value=float(amount))
                    amount_cell.number_format = money_format
                    row_num += 1

            row_num += 1

        # Summary Section
        ws.cell(row=row_num, column=1, value="Total Summary").font = Font(bold=True)
        row_num += 1

        # Summary Headers
        headers = ["Component", "Amount"]


        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        row_num += 1

        # Summary Data
        summary_items = [
            ("Taxable Gross Pay", item['totals']['taxable_gross']),
            ("Non-Taxable Gross Pay", item['totals']['non_taxable_gross']),
            ("Total Gross Pay", item['totals']['gross_pay']),
            ("Total Pensionable", item['totals']['pensionable']),
            ("Employee Pension", item['totals']['employee_pension']),
            ("Employer Pension", item['totals']['employer_pension']),
            ("Total Pension Contribution", item['totals']['total_pension']),
            ("Income Tax", item['totals']['employment_income_tax']),
            ("Total Deduction", item['totals']['deduction']),
            ("Total Expense", item['totals']['expense']),
            ("Final Net Pay", item['totals']['final_net_pay']),
        ]

        for component, amount in summary_items:
            ws.cell(row=row_num, column=1, value=component)
            amount_cell = ws.cell(row=row_num, column=2, value=float(amount))
            amount_cell.number_format = money_format
            row_num += 1

        # Add space between records
        row_num += 3

    # Auto-size columns
    for col in ws.columns:
        # Skip merged cells in column width calculation
        valid_cells = [cell for cell in col if not isinstance(cell, openpyxl.cell.cell.MergedCell)]
        if not valid_cells:
            continue

        max_length = 0
        column = valid_cells[0].column_letter  # Get letter from first valid cell

        for cell in valid_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=combined_personnel_payroll.xlsx'
    wb.save(response)

    return response
#


@login_required
def export_combined_personnel_list(request):
    payroll_data = get_combined_personnel_payroll_context(request).get('payroll_data', [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Combined Personnel Payroll"

    # Styles
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    total_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold_font = Font(bold=True)
    currency_format = '#,##0.00'
    thin_border = Border(right=Side(style='thin', color='FF000000'))

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=17)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Combined Personnel Payroll Summary"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = center_align

    # Headers
    headers = [
        "Month", "Personnel ID", "First Name", "Father Name", "Last Name", "Type",
        "Taxable Gross", "Non-Taxable Gross", "Total Gross", "Pensionable",
        "Employee Pension", "Employer Pension", "Total Pension Contribution",
        "Employment Income Tax", "Total Deduction", "Net Pay", "Total Expense"
    ]

    export_util = ExportUtilityService()
    decorated_headers = [export_util.split_header_to_lines(h) for h in headers]

    for col_num, header in enumerate(decorated_headers, 1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    def fmt_dec(val):
        return Decimal(val) if val is not None else Decimal('0.00')

    def write_row(row_idx, values, is_total=False):
        for col_offset, val in enumerate(values, 6):
            cell = ws.cell(row=row_idx, column=col_offset, value=val)
            cell.number_format = currency_format
            if is_total:
                cell.fill = total_fill
                cell.font = bold_font

    current_row = 3

    for row in payroll_data:
        payroll = row.get('payroll')
        p = payroll.personnel_full_name if payroll else None
        r = row.get('regular_totals', {})
        a = row.get('earning_adjustment_item', {})
        c = row.get('combined_adjustment', {})
        t = row.get('totals', {})

        payroll_month = str(getattr(getattr(payroll, 'payroll_month', None), 'payroll_month', '')) if payroll else ''

        # Merge Month and Personnel columns
        for col in range(1, 6):
            ws.merge_cells(start_row=current_row, start_column=col, end_row=current_row + 2, end_column=col)

        # Personnel Info
        ws.cell(row=current_row, column=1, value=payroll_month).alignment = center_align
        ws.cell(row=current_row, column=2, value=getattr(p, 'personnel_id', '')).alignment = center_align
        ws.cell(row=current_row, column=3, value=getattr(p, 'first_name', '')).alignment = center_align
        ws.cell(row=current_row, column=4, value=getattr(p, 'father_name', '')).alignment = center_align
        ws.cell(row=current_row, column=5, value=getattr(p, 'last_name', '')).alignment = center_align

        # Regular
        write_row(current_row, [
            "Regular",
            fmt_dec(r.get('taxable_gross')),
            fmt_dec(r.get('non_taxable_gross')),
            fmt_dec(r.get('gross_pay')),
            fmt_dec(r.get('pensionable')),
            fmt_dec(r.get('employee_pension')),
            fmt_dec(r.get('employer_pension')),
            fmt_dec(r.get('total_pension')),
            fmt_dec(r.get('employment_income_tax')),
            fmt_dec(r.get('deduction')),
            fmt_dec(r.get('net_pay')),
            fmt_dec(r.get('expense')),
        ])

        # Adjustment
        write_row(current_row + 1, [
            "Adjustment",
            fmt_dec(a.get('taxable_gross')),
            fmt_dec(a.get('non_taxable_gross')),
            fmt_dec(a.get('gross_pay')),
            fmt_dec(a.get('adjusted_pensionable')),
            fmt_dec(a.get('employee_pension')),
            fmt_dec(a.get('employer_pension')),
            fmt_dec(a.get('total_pension')),
            fmt_dec(a.get('employment_income_tax')),
            fmt_dec(c.get('total_adjustment_deduction')),
            fmt_dec(c.get('net_monthly_adjustment')),
            fmt_dec(a.get('expense')),
        ])

        # Total
        write_row(current_row + 2, [
            "Total",
            fmt_dec(t.get('taxable_gross')),
            fmt_dec(t.get('non_taxable_gross')),
            fmt_dec(t.get('gross_pay')),
            fmt_dec(t.get('pensionable')),
            fmt_dec(t.get('employee_pension')),
            fmt_dec(t.get('employer_pension')),
            fmt_dec(t.get('total_pension')),
            fmt_dec(t.get('employment_income_tax')),
            fmt_dec(t.get('deduction')),
            fmt_dec(t.get('final_net_pay')),
            fmt_dec(t.get('expense')),
        ], is_total=True)

        # Type column (6) in Total row should also be styled
        type_cell = ws.cell(row=current_row + 2, column=6)
        type_cell.fill = total_fill
        type_cell.font = bold_font

        current_row += 3

    # Set column widths
    for col, width in {1: 15, 2: 15, 3: 15, 4: 15, 5: 15}.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    for col in range(6, 18):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Return response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=combined_personnel_payroll.xlsx'
    wb.save(response)
    return response

    #

@login_required
def export_personnel_total_adjustment(request):
    context = get_combined_personnel_payroll_context(request)
    data = context.get('payroll_data', [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Personnel Total Adjustment"

    # Title and subtitle rows
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=20)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Combined Personnel Total Adjustment"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=20)
    subtitle_cell = ws.cell(row=2, column=1)
    subtitle_cell.value = "Details of personnel combined total earning and deduction adjustments per payroll month"
    subtitle_cell.font = Font(size=10, italic=True)
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')

    headers = [
        "Payroll Month", "Component", "Record Type",
        "Personnel ID", "First Name", "Father Name", "Last Name",
        "Taxable Gross", "Non-Taxable Gross", "Gross Pay",
        "Adjusted Pensionable", "Employee Pension", "Employer Pension",
        "Total Pension Contribution", "Employment Income Tax",
        "Earning Adjustment Deduction", "Other Adjustment Deduction",
        "Total Adjustment Deduction", "Net Adjustment Pay", "Adjustment Expense"
    ]

    export_util = ExportUtilityService()
    ws.append([export_util.split_header_to_lines(h) for h in headers])

    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(right=Side(style='thin', color='FF000000'))

    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    def safe_getattr(obj, attr, default=None):
        try:
            return getattr(obj, attr, default)
        except Exception:
            return default

    for item in data:
        payroll = item.get('payroll')
        earning = item.get('earning_adjustment_item', {})
        combined_adj = item.get('combined_adjustment', {})
        deduction_adj = item.get('deduction_adjustment')

        personnel = safe_getattr(payroll, 'personnel_full_name', None)

        # Payroll Month as string
        payroll_month = str(getattr(getattr(payroll, 'payroll_month', None), 'payroll_month', '')) if payroll else ''

        # Total deduction from deduction adjustment
        deduction_total = safe_getattr(deduction_adj, 'recorded_month_total_deduction', Decimal('0.00'))

        # Write Regular component row
        ws.append([
            payroll_month,
            "Regular Payroll",
            "Regular",
            safe_getattr(personnel, 'personnel_id', ''),
            safe_getattr(personnel, 'first_name', ''),
            safe_getattr(personnel, 'father_name', ''),
            safe_getattr(personnel, 'last_name', ''),
            earning.get('taxable_gross', Decimal('0.00')),
            earning.get('non_taxable_gross', Decimal('0.00')),
            earning.get('gross_pay', Decimal('0.00')),
            earning.get('adjusted_pensionable', Decimal('0.00')),
            earning.get('employee_pension', Decimal('0.00')),
            earning.get('employer_pension', Decimal('0.00')),
            earning.get('total_pension', Decimal('0.00')),
            earning.get('employment_income_tax', Decimal('0.00')),
            earning.get('earning_adjustment_deduction', Decimal('0.00')),
            deduction_total,
            combined_adj.get('total_adjustment_deduction', Decimal('0.00')),
            combined_adj.get('net_monthly_adjustment', Decimal('0.00')),
            earning.get('expense', Decimal('0.00')),
        ])

        # Write Adjustment component row
        ws.append([
            payroll_month,
            "Adjustment Payroll",
            "Adjustment",
            safe_getattr(personnel, 'personnel_id', ''),
            safe_getattr(personnel, 'first_name', ''),
            safe_getattr(personnel, 'father_name', ''),
            safe_getattr(personnel, 'last_name', ''),
            earning.get('taxable_gross', Decimal('0.00')),
            earning.get('non_taxable_gross', Decimal('0.00')),
            earning.get('gross_pay', Decimal('0.00')),
            earning.get('adjusted_pensionable', Decimal('0.00')),
            earning.get('employee_pension', Decimal('0.00')),
            earning.get('employer_pension', Decimal('0.00')),
            earning.get('total_pension', Decimal('0.00')),
            earning.get('employment_income_tax', Decimal('0.00')),
            earning.get('earning_adjustment_deduction', Decimal('0.00')),
            deduction_total,
            combined_adj.get('total_adjustment_deduction', Decimal('0.00')),
            combined_adj.get('net_monthly_adjustment', Decimal('0.00')),
            earning.get('expense', Decimal('0.00')),
        ])

        # If you have severance data similarly, add like this:
        severance = item.get('severance_adjustment', {})
        if severance:
            ws.append([
                payroll_month,
                "Severance Payroll",
                "Severance",
                safe_getattr(personnel, 'personnel_id', ''),
                safe_getattr(personnel, 'first_name', ''),
                safe_getattr(personnel, 'father_name', ''),
                safe_getattr(personnel, 'last_name', ''),
                severance.get('taxable_gross', Decimal('0.00')),
                severance.get('non_taxable_gross', Decimal('0.00')),
                severance.get('gross_pay', Decimal('0.00')),
                severance.get('adjusted_pensionable', Decimal('0.00')),
                severance.get('employee_pension', Decimal('0.00')),
                severance.get('employer_pension', Decimal('0.00')),
                severance.get('total_pension', Decimal('0.00')),
                severance.get('employment_income_tax', Decimal('0.00')),
                severance.get('earning_adjustment_deduction', Decimal('0.00')),
                severance.get('other_adjustment_deduction', Decimal('0.00')),
                severance.get('total_adjustment_deduction', Decimal('0.00')),
                severance.get('net_monthly_adjustment', Decimal('0.00')),
                severance.get('expense', Decimal('0.00')),
            ])

    # Adjust column widths
    MIN_WIDTH = 10
    MAX_WIDTH = 18
    for idx, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max(MIN_WIDTH, min(MAX_WIDTH, max_length + 2))
        ws.column_dimensions[get_column_letter(idx)].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=personnel_total_adjustment.xlsx"
    return response


@login_required
def export_combined_personnel_total(request):
    context = get_combined_personnel_payroll_context(request)
    payroll_data = context.get('payroll_data', [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Total Payroll Summary"

    # Title row (merged across all columns)
    total_cols = 16
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Combined Personnel Total Payroll Summary"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Header row
    headers = [
        "Month", "Personnel ID", "First Name", "Father Name", "Last Name",
        "Taxable Gross", "Non-Taxable Gross", "Total Gross Pay", "Total Pensionable",
        "Employee Pension", "Employer Pension", "Total Pension Contribution",
        "Employment Income Tax", "Total Deduction", "Final Net Pay", "Total Expense"
    ]

    # Use your export utility for multiline headers if needed
    export_util = ExportUtilityService()
    ws.append([export_util.split_header_to_lines(h) for h in headers])

    # Style header row (row 2)
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(right=Side(style='thin', color='FF000000'))

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Helper function for safe attribute access
    def safe_getattr(obj, attr, default=None):
        try:
            return getattr(obj, attr, default)
        except Exception:
            return default

    # Write data rows
    for item in payroll_data:
        payroll = item.get('payroll')
        personnel = safe_getattr(payroll, 'personnel_full_name', None)
        payroll_month_obj = safe_getattr(payroll, 'payroll_month', None)
        totals = item.get('totals', {})

        # Convert payroll_month to string (avoid Excel errors)
        payroll_month = str(getattr(getattr(payroll, 'payroll_month', None), 'payroll_month', '')) if payroll else ''

        row = [
            payroll_month,
            safe_getattr(personnel, 'personnel_id', ''),
            safe_getattr(personnel, 'first_name', ''),
            safe_getattr(personnel, 'father_name', ''),
            safe_getattr(personnel, 'last_name', ''),
            float(totals.get('taxable_gross', 0)),
            float(totals.get('non_taxable_gross', 0)),
            float(totals.get('gross_pay', 0)),
            float(totals.get('pensionable', 0)),
            float(totals.get('employee_pension', 0)),
            float(totals.get('employer_pension', 0)),
            float(totals.get('total_pension', 0)),
            float(totals.get('employment_income_tax', 0)),
            float(totals.get('deduction', 0)),
            float(totals.get('final_net_pay', 0)),
            float(totals.get('expense', 0)),
        ]
        ws.append(row)

    # Adjust column widths (start from col 1)
    MIN_WIDTH = 10
    MAX_WIDTH = 15
    for idx, col_cells in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
        adjusted_width = max(MIN_WIDTH, min(MAX_WIDTH, max_length + 2))
        ws.column_dimensions[get_column_letter(idx)].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=total_combined_payroll_summary.xlsx"
    return response


@login_required
def export_combined_personnel_expense(request):
    context = get_combined_personnel_payroll_context(request)
    payroll_data = context.get('payroll_data', [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Total Payroll Expense Summary"

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Combined Personnel Expense Summary"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    headers = [
        "Month", "Personnel ID", "First Name", "Father Name", "Last Name",
        "Total Gross Pay", "Employer Pension", "Total Expense"
    ]

    export_util = ExportUtilityService()
    ws.append([export_util.split_header_to_lines(h) for h in headers])


    # Header styling
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(right=Side(style='thin', color='FF000000'))

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    for item in payroll_data:

        payroll = item.get('payroll')
        p = getattr(item.get('payroll'), 'personnel_full_name', None)
        t = item.get('totals', {})

        # Convert payroll_month to string if it's an object
        payroll_month = str(getattr(getattr(payroll, 'payroll_month', None), 'payroll_month', '')) if payroll else ''

        ws.append([
            payroll_month,
            getattr(p, 'personnel_id', '') if p else '',
            getattr(p, 'first_name', '') if p else '',
            getattr(p, 'father_name', '') if p else '',
            getattr(p, 'last_name', '') if p else '',
            float(t.get('gross_pay', 0)),
            float(t.get('employer_pension', 0)),
            float(t.get('expense', 0)),
        ])

    # Set simple fixed column widths (optional)
    widths = [15, 15, 15, 15, 15, 15, 15, 15]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=combined_payroll_expense_summary.xlsx"
    return response


@login_required
def export_combined_personnel_net_income(request):
    context = get_combined_personnel_payroll_context(request)
    payroll_data = context.get('payroll_data', [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Total Net Income Summary"

    # Title row (merged)
    total_cols = 8
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Combined Personnel Net Income Summary"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    headers = [
        "Month", "Personnel ID", "First Name", "Father Name", "Last Name",
        "Total Gross Pay", "Total Deduction", "Final Net Pay"
    ]

    export_util = ExportUtilityService()
    ws.append([export_util.split_header_to_lines(h) for h in headers])

    # Header styling
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(right=Side(style='thin', color='FF000000'))

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    def safe_getattr(obj, attr, default=None):
        try:
            return getattr(obj, attr, default)
        except Exception:
            return default

    for item in payroll_data:
        payroll = item.get('payroll')
        personnel = safe_getattr(payroll, 'personnel_full_name', None)

        payroll_month = str(getattr(getattr(payroll, 'payroll_month', None), 'payroll_month', '')) if payroll else ''

        totals = item.get('totals', {})

        row = [
            payroll_month,
            safe_getattr(personnel, 'personnel_id', ''),
            safe_getattr(personnel, 'first_name', ''),
            safe_getattr(personnel, 'father_name', ''),
            safe_getattr(personnel, 'last_name', ''),
            float(totals.get('gross_pay', 0)),
            float(totals.get('deduction', 0)),
            float(totals.get('final_net_pay', 0)),
        ]
        ws.append(row)

    # Adjust column widths with reasonable limits
    MIN_WIDTH = 10
    MAX_WIDTH = 15
    for idx, col_cells in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
        adjusted_width = max(MIN_WIDTH, min(MAX_WIDTH, max_length + 2))
        ws.column_dimensions[get_column_letter(idx)].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response['Content-Disposition'] = 'attachment; filename=combined_net_income_summary.xlsx'
    return response

@login_required
def export_combined_personnel_employment_tax(request):
    context = get_combined_personnel_payroll_context(request)
    payroll_data = context.get('payroll_data', [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Employment Tax Summary"

    # Title row (merged)
    total_cols = 7
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Combined Personnel Total Employment Income Tax Summary"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    headers = [
        "Month", "Personnel ID", "First Name", "Father Name", "Last Name",
        "Total Taxable Gross", "Total Employment Income Tax"
    ]

    export_util = ExportUtilityService()
    ws.append([export_util.split_header_to_lines(h) for h in headers])

    # Header styling
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(right=Side(style='thin', color='FF000000'))

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    def safe_getattr(obj, attr, default=None):
        try:
            return getattr(obj, attr, default)
        except Exception:
            return default

    for item in payroll_data:
        payroll = item.get('payroll')
        personnel = safe_getattr(payroll, 'personnel_full_name', None)

        payroll_month = str(getattr(getattr(payroll, 'payroll_month', None), 'payroll_month', '')) if payroll else ''

        totals = item.get('totals', {})

        row = [
            payroll_month,
            safe_getattr(personnel, 'personnel_id', ''),
            safe_getattr(personnel, 'first_name', ''),
            safe_getattr(personnel, 'father_name', ''),
            safe_getattr(personnel, 'last_name', ''),
            float(totals.get('taxable_gross', 0)),
            float(totals.get('employment_income_tax', 0)),
        ]
        ws.append(row)

    # Adjust column widths within limits
    MIN_WIDTH = 10
    MAX_WIDTH = 15
    for idx, col_cells in enumerate(ws.columns, 1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
        adjusted_width = max(MIN_WIDTH, min(MAX_WIDTH, max_length + 2))
        ws.column_dimensions[get_column_letter(idx)].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response['Content-Disposition'] = 'attachment; filename=employment_tax_summary.xlsx'
    return response


@login_required
def export_combined_personnel_pension(request):
    """
    Export Combined Total Pension Contribution Summary (Per Employee Per Month) as Excel.
    """
    context = get_combined_personnel_payroll_context(request)
    payroll_data = context.get('payroll_data', [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Pension Summary"

    # Title row (merged)
    total_cols = 9
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Combined Personnel Pension Summary"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    headers = [
        'Payroll Month', 'Personnel ID', 'First Name', 'Father Name', 'Last Name',
        'Total Pensionable', 'Total Employee Pension',
        'Total Employer Pension', 'Total Pension Contribution'
    ]

    export_util = ExportUtilityService()
    ws.append([export_util.split_header_to_lines(h) for h in headers])

    # Header style
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Orange
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(right=Side(style='thin', color='FF000000'))

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    def safe_float(value):
        if value is None:
            return 0.0
        if isinstance(value, Decimal):
            return float(value)
        try:
            return float(value)
        except Exception:
            return 0.0

    for item in payroll_data:
        payroll = item.get('payroll')
        totals = item.get('totals', {})
        personnel = getattr(payroll, 'personnel_full_name', None)

        payroll_month = str(getattr(getattr(payroll, 'payroll_month', None), 'payroll_month', '')) if payroll else ''

        if not personnel:
            continue

        row = [
            payroll_month,
            getattr(personnel, 'personnel_id', ''),
            getattr(personnel, 'first_name', ''),
            getattr(personnel, 'father_name', ''),
            getattr(personnel, 'last_name', ''),
            safe_float(totals.get('pensionable')),
            safe_float(totals.get('employee_pension')),
            safe_float(totals.get('employer_pension')),
            safe_float(totals.get('total_pension')),
        ]
        ws.append(row)

    # Adjust column widths
    MIN_WIDTH = 12
    MAX_WIDTH = 15
    for i, col_cells in enumerate(ws.columns, 1):
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        adjusted_width = max(MIN_WIDTH, min(MAX_WIDTH, max_len + 2))
        ws.column_dimensions[get_column_letter(i)].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=pension_summary.xlsx'
    return response