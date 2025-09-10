from django.shortcuts import render, get_object_or_404, redirect
from django.core.paginator import Paginator
from django.urls import reverse
from django.contrib import messages
from django.db.models import Sum, F, Q
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from io import BytesIO
import plotly.graph_objs as go
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from management_project.models import PayrollMonthComponent, RegularPayroll
from management_project.forms import RegularPayrollForm
from management_project.services.excel_export import ExportUtilityService


# payroll month list
@login_required
def payroll_month_list(request):
    payroll_months = PayrollMonthComponent.objects.filter(
        organization_name=request.user.organization_name
    ).values(
        'payroll_month__year',
        'payroll_month__payroll_month',
        'payroll_month__slug',
    ).order_by(
        '-payroll_month__year',
        '-payroll_month__month'
    ).distinct()

    return render(request, 'payroll_month/list.html', {
        'payroll_months': payroll_months
    })

#
@login_required
def regular_payroll_list(request, payroll_month_slug):
    payroll_month = get_object_or_404(PayrollMonthComponent, slug=payroll_month_slug)

    # Filter the regular payrolls for the user and payroll month
    regular_payrolls = RegularPayroll.objects.filter(
        payroll_month=payroll_month,
        organization_name=request.user.organization_name
    ).order_by('-payroll_month__payroll_month__year', '-payroll_month__payroll_month__month')

    # Apply search functionality if search query exists
    search_query = request.GET.get('search', '')
    if search_query:
        regular_payrolls = regular_payrolls.filter(
            Q(personnel_full_name__personnel_id__icontains=search_query) |
            Q(personnel_full_name__first_name__icontains=search_query) |
            Q(personnel_full_name__father_name__icontains=search_query) |
            Q(personnel_full_name__last_name__icontains=search_query)
        )
    regular_payrolls = regular_payrolls.order_by('-id')

    # Pagination
    paginator = Paginator(regular_payrolls, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Form with user-specific personnel filtering
    form = RegularPayrollForm(
        initial={'payroll_month': payroll_month},
        request=request  # Pass request to filter personnels
    )

    context = {
        'payroll_month': payroll_month,
        'page_obj': page_obj,
        'form': form,
        'search_query': search_query,
    }
    return render(request, 'regular_payroll/list_by_month.html', context)


@login_required
def regular_payroll_detail(request, payroll_month_slug, regular_payroll_pk):
    payroll_month = get_object_or_404(PayrollMonthComponent, slug=payroll_month_slug)
    regular_payroll = get_object_or_404(
        RegularPayroll,
        pk=regular_payroll_pk,
        payroll_month=payroll_month,
        organization_name=request.user.organization_name  # Add user filter for security
    )

    context = {
        'payroll_month': payroll_month,
        'regular_payroll': regular_payroll,
    }
    return render(request, 'regular_payroll/detail.html', context)


# #
@login_required
def create_regular_payroll(request, payroll_month_slug):
    payroll_month = get_object_or_404(PayrollMonthComponent, slug=payroll_month_slug)
    # Get payroll_month_slug however you determine it

    if request.method == 'POST':
        form = RegularPayrollForm(
            request.POST,
            initial={'payroll_month': payroll_month},
            request=request
        )
        if form.is_valid():
            regular_payroll = form.save(commit=False)
            regular_payroll.payroll_month = payroll_month
            regular_payroll.organization_name = request.user.organization_name
            regular_payroll.save()
            messages.success(request, "Payroll created successfully!")

            return redirect('regular_payroll_list', payroll_month_slug=payroll_month.slug)
    else:
        form = RegularPayrollForm(
            initial={'payroll_month': payroll_month},
            request=request
        )

    context = {
        'form': form,
        # Updated to use month and year fields instead of name
        'form_title': f'Create Regular Payroll for {payroll_month.payroll_month.payroll_month}',
        'submit_button_text': 'Create Regular Payroll',
        'back_url': reverse('regular_payroll_list', kwargs={'payroll_month_slug': payroll_month.slug}),
        'show_add_personnel': True,
        'current_url': request.path,  # Pass current URL for redirect back
        'payroll_month_slug': payroll_month.slug  # Pass the slug to the template

    }
    return render(request, 'regular_payroll/form.html', context)


#
#
@login_required
def update_regular_payroll(request, payroll_month_slug, pk):
    payroll_month = get_object_or_404(PayrollMonthComponent, slug=payroll_month_slug)
    regular_payroll = get_object_or_404(
        RegularPayroll,
        id=pk,
        payroll_month=payroll_month,
        organization_name=request.user.organization_name
    )

    if request.method == 'POST':
        form = RegularPayrollForm(request.POST, instance=regular_payroll, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, "Payroll updated successfully!")
            return redirect('regular_payroll_list', payroll_month_slug=payroll_month.slug)
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = RegularPayrollForm(instance=regular_payroll, request=request)

    context = {
        'form': form,
        'form_title': f'Update Regular Payroll for {payroll_month.payroll_month.month} {payroll_month.payroll_month.year}',
        'submit_button_text': 'Update Regular Payroll',
        'back_url': reverse('regular_payroll_list', kwargs={'payroll_month_slug': payroll_month.slug}),
        'show_add_personnel': True,
        'current_url': request.path,
        'payroll_month_slug': payroll_month.slug,
        'payroll_month': payroll_month,
        'edit_regular_payroll': regular_payroll
    }

    return render(request, 'regular_payroll/form.html', context)


@login_required
def delete_regular_payroll(request, payroll_month_slug, pk):
    payroll_month = get_object_or_404(PayrollMonthComponent, slug=payroll_month_slug)
    regular_payroll = get_object_or_404(
        RegularPayroll, id=pk, payroll_month=payroll_month, organization_name=request.user.organization_name
    )

    if request.method == 'POST':
        regular_payroll.delete()
        messages.success(request, "Payroll deleted successfully!")
        return redirect('regular_payroll_list', payroll_month_slug=payroll_month.slug)

    return render(
        request, 'regular_payroll/delete_confirm.html',
        {'regular_payroll': regular_payroll, 'payroll_month': payroll_month}
    )


@login_required
def payroll_month_for_journal_entry_report(request):
    # Get distinct payroll months for the current user
    payroll_month_journal_entry = RegularPayroll.objects.filter(
        organization_name=request.user.organization_name
    ).select_related('payroll_month').values(
        'payroll_month__slug',
        'payroll_month__payroll_month__payroll_month'
    ).distinct().order_by('-payroll_month__payroll_month__year', '-payroll_month__payroll_month__month')

    context = {
        'payroll_month_journal_entry': payroll_month_journal_entry,
    }
    return render(request, 'regular_payroll/journal_entry/month_list.html', context)


@login_required
def payroll_month_for_summary_report(request):
    # Get payroll months only for the current user's organization
    payroll_month_summary = RegularPayroll.objects.filter(
        organization_name=request.user.organization_name
    ).select_related('payroll_month').values(
        'payroll_month__slug',
        'payroll_month__payroll_month__payroll_month'
    ).distinct().order_by('-payroll_month__payroll_month__year', '-payroll_month__payroll_month__month')

    context = {
        'payroll_month_summary': payroll_month_summary,
    }
    return render(request, 'regular_payroll/summary/month_list.html', context)


#
@login_required
def get_regular_payroll_by_month_report(request, payroll_month_slug):
    payroll_month = get_object_or_404(PayrollMonthComponent, slug=payroll_month_slug)

    aggregate_by_month = RegularPayroll.objects.filter(
        payroll_month=payroll_month,
        organization_name=request.user.organization_name
    ).values(
        'payroll_month__payroll_month__year', 'payroll_month__payroll_month__payroll_month'
    ).annotate(
        aggregate_basic_salary=Sum('basic_salary'),
        aggregate_overtime=Sum('overtime'),
        aggregate_housing_allowance=Sum('housing_allowance'),
        aggregate_position_allowance=Sum('position_allowance'),
        aggregate_commission=Sum('commission'),
        aggregate_telephone_allowance=Sum('telephone_allowance'),
        aggregate_one_time_bonus=Sum('one_time_bonus'),
        aggregate_causal_labor_wage=Sum('causal_labor_wage'),
        aggregate_transport_home_to_office_taxable=Sum('transport_home_to_office_taxable'),
        aggregate_transport_home_to_office_non_taxable=Sum('transport_home_to_office_non_taxable'),
        aggregate_fuel_home_to_office_taxable=Sum('fuel_home_to_office_taxable'),
        aggregate_fuel_home_to_office_non_taxable=Sum('fuel_home_to_office_non_taxable'),
        aggregate_transport_for_work_taxable=Sum('transport_for_work_taxable'),
        aggregate_transport_for_work_non_taxable=Sum('transport_for_work_non_taxable'),
        aggregate_fuel_for_work_taxable=Sum('fuel_for_work_taxable'),
        aggregate_fuel_for_work_non_taxable=Sum('fuel_for_work_non_taxable'),
        aggregate_per_diem_taxable=Sum('per_diem_taxable'),
        aggregate_per_diem_non_taxable=Sum('per_diem_non_taxable'),
        aggregate_hardship_allowance_taxable=Sum('hardship_allowance_taxable'),
        aggregate_hardship_allowance_non_taxable=Sum('hardship_allowance_non_taxable'),
        aggregate_public_cash_award=Sum('public_cash_award'),
        aggregate_incidental_operation_allowance=Sum('incidental_operation_allowance'),
        aggregate_medical_allowance=Sum('medical_allowance'),
        aggregate_cash_gift=Sum('cash_gift'),
        aggregate_tuition_fees=Sum('tuition_fees'),
        aggregate_personal_injury=Sum('personal_injury'),
        aggregate_child_support_payment=Sum('child_support_payment'),
        aggregate_charitable_donation=Sum('charitable_donation'),
        aggregate_saving_plan=Sum('saving_plan'),
        aggregate_loan_payment=Sum('loan_payment'),
        aggregate_court_order=Sum('court_order'),
        aggregate_workers_association=Sum('workers_association'),
        aggregate_personnel_insurance_saving=Sum('personnel_insurance_saving'),
        aggregate_university_cost_share_pay=Sum('university_cost_share_pay'),
        aggregate_red_cross=Sum('red_cross'),
        aggregate_party_contribution=Sum('party_contribution'),
        aggregate_other_deduction=Sum('other_deduction'),
        aggregate_employment_income_tax=Sum('employment_income_tax'),
        aggregate_employee_pension_contribution=Sum('employee_pension_contribution'),
        aggregate_employer_pension_contribution=Sum('employer_pension_contribution'),
        aggregate_total_pension_contribution=Sum('total_pension_contribution'),
        aggregate_gross_pay=Sum('gross_pay'),
        aggregate_gross_non_taxable_pay=Sum('gross_non_taxable_pay'),
        aggregate_gross_taxable_pay=Sum('gross_taxable_pay'),
        aggregate_total_payroll_deduction=Sum('total_payroll_deduction'),
        aggregate_net_pay=Sum('net_pay'),
        aggregate_expense=Sum('expense')
    ).order_by('-payroll_month__payroll_month__year', '-payroll_month__payroll_month__month')

    return {
        'payroll_month': payroll_month,
        'aggregate_by_month': aggregate_by_month
    }



@login_required
def regular_payroll_summary(request, payroll_month_slug):
    context = get_regular_payroll_by_month_report(request, payroll_month_slug)
    return render(request,'regular_payroll/summary/month.html', context)


@login_required
def regular_payroll_journal_entry(request, payroll_month_slug):
    context = get_regular_payroll_by_month_report(request, payroll_month_slug)
    return render(request, 'regular_payroll/journal_entry/month.html', context)

#

#
@login_required
def payroll_by_year_summary_report(request):
    aggregate_by_year = RegularPayroll.objects.filter(
        organization_name=request.user.organization_name
    ).values(
        'payroll_month__payroll_month__year'
    ).annotate(
        aggregate_basic_salary=Sum('basic_salary'),
        aggregate_overtime=Sum('overtime'),
        aggregate_housing_allowance=Sum('housing_allowance'),
        aggregate_position_allowance=Sum('position_allowance'),
        aggregate_commission=Sum('commission'),
        aggregate_telephone_allowance=Sum('telephone_allowance'),
        aggregate_one_time_bonus=Sum('one_time_bonus'),
        aggregate_causal_labor_wage=Sum('causal_labor_wage'),
        aggregate_transport_home_to_office_taxable=Sum('transport_home_to_office_taxable'),
        aggregate_transport_home_to_office_non_taxable=Sum('transport_home_to_office_non_taxable'),
        aggregate_fuel_home_to_office_taxable=Sum('fuel_home_to_office_taxable'),
        aggregate_fuel_home_to_office_non_taxable=Sum('fuel_home_to_office_non_taxable'),
        aggregate_transport_for_work_taxable=Sum('transport_for_work_taxable'),
        aggregate_transport_for_work_non_taxable=Sum('transport_for_work_non_taxable'),
        aggregate_fuel_for_work_taxable=Sum('fuel_for_work_taxable'),
        aggregate_fuel_for_work_non_taxable=Sum('fuel_for_work_non_taxable'),
        aggregate_per_diem_taxable=Sum('per_diem_taxable'),
        aggregate_per_diem_non_taxable=Sum('per_diem_non_taxable'),
        aggregate_hardship_allowance_taxable=Sum('hardship_allowance_taxable'),
        aggregate_hardship_allowance_non_taxable=Sum('hardship_allowance_non_taxable'),
        aggregate_public_cash_award=Sum('public_cash_award'),
        aggregate_incidental_operation_allowance=Sum('incidental_operation_allowance'),
        aggregate_medical_allowance=Sum('medical_allowance'),
        aggregate_cash_gift=Sum('cash_gift'),
        aggregate_tuition_fees=Sum('tuition_fees'),
        aggregate_personal_injury=Sum('personal_injury'),
        aggregate_child_support_payment=Sum('child_support_payment'),
        aggregate_charitable_donation=Sum('charitable_donation'),
        aggregate_saving_plan=Sum('saving_plan'),
        aggregate_loan_payment=Sum('loan_payment'),
        aggregate_court_order=Sum('court_order'),
        aggregate_workers_association=Sum('workers_association'),
        aggregate_personnel_insurance_saving=Sum('personnel_insurance_saving'),
        aggregate_university_cost_share_pay=Sum('university_cost_share_pay'),
        aggregate_red_cross=Sum('red_cross'),
        aggregate_party_contribution=Sum('party_contribution'),
        aggregate_other_deduction=Sum('other_deduction'),
        aggregate_employment_income_tax=Sum('employment_income_tax'),
        aggregate_employee_pension_contribution=Sum('employee_pension_contribution'),
        aggregate_employer_pension_contribution=Sum('employer_pension_contribution'),
        aggregate_total_pension_contribution=Sum('total_pension_contribution'),
        aggregate_gross_pay=Sum('gross_pay'),
        aggregate_gross_non_taxable_pay=Sum('gross_non_taxable_pay'),
        aggregate_gross_taxable_pay=Sum('gross_taxable_pay'),
        aggregate_total_payroll_deduction=Sum('total_payroll_deduction'),
        aggregate_net_pay=Sum('net_pay'),
        aggregate_expense=Sum('expense')
    ).order_by('-payroll_month__payroll_month__year')

    return render(
        request,
        'regular_payroll/summary/year.html',
        {'aggregate_by_year': aggregate_by_year}
    )
#

@login_required
def payroll_processing_graphs(request):
    # Fetch all payroll months, ordered by the most recent first
    payroll_months = PayrollMonthComponent.objects.filter(organization_name=request.user.organization_name)

    # Labels for the x-axis (month names)
    month_labels = [payroll_month.payroll_month.payroll_month for payroll_month in payroll_months]

    # Lists to store values for each field (for all months)
    employment_income_tax_values = []
    employee_pension_contribution_values = []
    employer_pension_contribution_values = []
    total_pension_contribution_values = []
    gross_pay_values = []
    gross_non_taxable_pay_values = []
    gross_taxable_pay_values = []
    total_payroll_deduction_values = []
    net_pay_values = []
    expense_values = []

    # Aggregate data for each month
    for payroll_month in payroll_months:
        # Get payroll processing records for the selected payroll month
        payrolls = RegularPayroll.objects.filter(payroll_month=payroll_month,
                                                 organization_name=request.user.organization_name)

        # Aggregations for each payroll month
        aggregates = {
            'monthly_employment_income_tax': Sum('employment_income_tax'),
            'monthly_employee_pension_contribution': Sum('employee_pension_contribution'),
            'monthly_employer_pension_contribution': Sum('employer_pension_contribution'),
            'monthly_total_pension_contribution': Sum('total_pension_contribution'),
            'monthly_gross_pay': Sum('gross_pay'),
            'monthly_gross_non_taxable_pay': Sum('gross_non_taxable_pay'),
            'monthly_gross_taxable_pay': Sum('gross_taxable_pay'),
            'monthly_total_payroll_deduction': Sum('total_payroll_deduction'),
            'monthly_net_pay': Sum('net_pay'),
            'monthly_expense': Sum('expense'),
        }

        # Perform the aggregation and round the values
        aggregate_data = payrolls.aggregate(**aggregates)
        aggregate_data = {key: round(value or 0, 2) for key, value in aggregate_data.items()}

        # Append the aggregated values for each field
        employment_income_tax_values.append(aggregate_data['monthly_employment_income_tax'])
        employee_pension_contribution_values.append(
            aggregate_data['monthly_employee_pension_contribution'])
        employer_pension_contribution_values.append(
            aggregate_data['monthly_employer_pension_contribution'])
        total_pension_contribution_values.append(aggregate_data['monthly_total_pension_contribution'])
        gross_pay_values.append(aggregate_data['monthly_gross_pay'])
        gross_non_taxable_pay_values.append(aggregate_data['monthly_gross_non_taxable_pay'])
        gross_taxable_pay_values.append(aggregate_data['monthly_gross_taxable_pay'])
        total_payroll_deduction_values.append(aggregate_data['monthly_total_payroll_deduction'])
        net_pay_values.append(aggregate_data['monthly_net_pay'])
        expense_values.append(aggregate_data['monthly_expense'])

    # Create line graph traces for each component
    trace_employment_income_tax = go.Scatter(
        x=month_labels,
        y=employment_income_tax_values,
        mode='lines+markers',  # Line graph with markers
        name='Employment Income Tax'
    )

    trace_employee_pension_contribution = go.Scatter(
        x=month_labels,
        y=employee_pension_contribution_values,
        mode='lines+markers',
        name='Employee Pension Contribution'
    )

    trace_employer_pension_contribution = go.Scatter(
        x=month_labels,
        y=employer_pension_contribution_values,
        mode='lines+markers',
        name='Employer Pension Contribution'
    )

    trace_total_pension_contribution = go.Scatter(
        x=month_labels,
        y=total_pension_contribution_values,
        mode='lines+markers',
        name='Total Pension Contribution'
    )

    trace_gross_pay = go.Scatter(
        x=month_labels,
        y=gross_pay_values,
        mode='lines+markers',
        name='Gross Pay'
    )

    trace_gross_non_taxable_pay = go.Scatter(
        x=month_labels,
        y=gross_non_taxable_pay_values,
        mode='lines+markers',
        name='Gross Non-Taxable Pay'
    )

    trace_gross_taxable_pay = go.Scatter(
        x=month_labels,
        y=gross_taxable_pay_values,
        mode='lines+markers',
        name='Gross Taxable Pay'
    )

    trace_total_payroll_deduction = go.Scatter(
        x=month_labels,
        y=total_payroll_deduction_values,
        mode='lines+markers',
        name='Total Payroll Deduction'
    )

    trace_net_pay = go.Scatter(
        x=month_labels,
        y=net_pay_values,
        mode='lines+markers',
        name='Net Pay'
    )

    trace_expense = go.Scatter(
        x=month_labels,
        y=expense_values,
        mode='lines+markers',
        name='Regular Personnel Regular Payroll Cost'
    )

    # Layout for the chart
    layout = go.Layout(
        title='Regular Payroll Report By Months',
        xaxis=dict(title='Payroll Month'),
        yaxis=dict(title='Amount'),
        xaxis_tickangle=-45
    )

    # Create the figure with all traces
    figure = go.Figure(data=[
        trace_employment_income_tax, trace_employee_pension_contribution,
        trace_employer_pension_contribution, trace_total_pension_contribution,
        trace_gross_pay, trace_gross_non_taxable_pay,
        trace_gross_taxable_pay, trace_total_payroll_deduction,
        trace_net_pay, trace_expense
    ], layout=layout)

    # Convert the figure to HTML
    chart_json = figure.to_html(full_html=False)

    # Pass the chart to the template
    return render(request, 'regular_payroll/graph/regular_payroll.html', {
        'chart_json': chart_json,
    })




@login_required
def university_cost_sharing_view(request):
    # Querying the RegularPayroll model and aggregating the required fields
    university_cost_shares = (
        RegularPayroll.objects.filter(organization_name=request.user.organization_name)
        .values('personnel_full_name__first_name', 'personnel_full_name__father_name', 'personnel_full_name__last_name')
        .annotate(
            university_cost_sharing_debt=F('personnel_full_name__university_cost_sharing_debt'),
            sum_cost_sharing_payed=Sum('university_cost_share_pay')
        )
        .annotate(
            remaining_cost_sharing=F('university_cost_sharing_debt') - F('sum_cost_sharing_payed')
        )
        # Exclude records where the university_cost_sharing_debt is None or 0
        .filter(university_cost_sharing_debt__gt=0)  # Debt should be greater than 0
    )

    # If no data is found after filtering, you can choose to handle it by either setting an empty queryset or returning a default message.
    if not university_cost_shares:
        university_cost_shares = []

    context = {
        'university_cost_shares': university_cost_shares
    }
    return render(request, 'regular_payroll/university_cost_sharing.html', context)


#export regular
#

@login_required
def export_regular_payroll_to_excel(request, payroll_month_slug):
    payroll_month = get_object_or_404(
        PayrollMonthComponent,
        slug=payroll_month_slug,
        organization_name=request.user.organization_name
    )

    regular_payrolls = RegularPayroll.objects.filter(
        payroll_month=payroll_month,
        organization_name=request.user.organization_name
    )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"Payroll process for {payroll_month.payroll_month.payroll_month}"

    # Build header list dynamically based on payroll_month flags
    header = ["ID"]
    header.extend(['Personnel ID', 'First Name', 'Father Name', 'Last Name'])

    # Add fully taxable fields
    if payroll_month.use_basic_salary:
        header.append("Basic Salary")
    if payroll_month.use_overtime:
        header.append("Overtime")
    if payroll_month.use_housing_allowance:
        header.append("Housing Allowance")
    if payroll_month.use_position_allowance:
        header.append("Position Allowance")
    if payroll_month.use_commission:
        header.append("Commission")
    if payroll_month.use_telephone_allowance:
        header.append("Telephone Allowance")
    if payroll_month.use_one_time_bonus:
        header.append("One-Time Bonus")
    if payroll_month.use_causal_labor_wage:
        header.append("Causal Labor Wage")

    # Add partially taxable fields
    if payroll_month.use_transport_home_to_office:
        header.append("Transport Home to Office Taxable")
        header.append("Transport Home to Office Non-Taxable")
    if payroll_month.use_fuel_home_to_office:
        header.append("Fuel Home to Office Taxable")
        header.append("Fuel Home to Office Non-Taxable")
    if payroll_month.use_transport_for_work:
        header.append("Transport for Work Taxable")
        header.append("Transport for Work Non-Taxable")
    if payroll_month.use_fuel_for_work:
        header.append("Fuel for Work Taxable")
        header.append("Fuel for Work Non-Taxable")
    if payroll_month.use_per_diem:
        header.append("Per Diem Taxable")
        header.append("Per Diem Non-Taxable")
    if payroll_month.use_hardship_allowance:
        header.append("Hardship Allowance Taxable")
        header.append("Hardship Allowance Non-Taxable")

    # Fully non-taxable fields
    if payroll_month.use_public_cash_award:
        header.append("Public Cash Award")
    if payroll_month.use_incidental_operation_allowance:
        header.append("Incidental Operation Allowance")
    if payroll_month.use_medical_allowance:
        header.append("Medical Allowance")
    if payroll_month.use_cash_gift:
        header.append("Cash Gift")
    if payroll_month.use_tuition_fees:
        header.append("Tuition Fees")
    if payroll_month.use_personal_injury:
        header.append("Personal Injury")
    if payroll_month.use_child_support_payment:
        header.append("Child Support Payment")

    # Deduction fields
    if payroll_month.use_charitable_donation:
        header.append("Charitable Donation")
    if payroll_month.use_saving_plan:
        header.append("Saving Plan")
    if payroll_month.use_loan_payment:
        header.append("Loan Payment")
    if payroll_month.use_court_order:
        header.append("Court Order")
    if payroll_month.use_workers_association:
        header.append("Workers Association")
    if payroll_month.use_personnel_insurance_saving:
        header.append("Personnel Insurance Saving")
    if payroll_month.use_university_cost_share_pay:
        header.append("University Cost Share Pay")
    if payroll_month.use_red_cross:
        header.append("Red Cross")
    if payroll_month.use_party_contribution:
        header.append("Party Contribution")
    if payroll_month.use_other_deduction:
        header.append("Other Deduction")

    # Always include summary fields
    summary_fields = [
        'Employment Income Tax',
        'Employee Pension Contribution',
        'Employer Pension Contribution',
        'Total Pension Contribution',
        'Gross Pay',
        'Gross Taxable Pay',
        'Gross Non-Taxable Pay',
        'Total Payroll Deduction',
        'Net Pay',
        'Expense'
    ]
    header.extend(summary_fields)


    # Add Title (row 1)
    total_cols = len(header)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    title_cell = sheet.cell(row=1, column=1)
    title_cell.value = f"Regular Payroll For - {payroll_month.payroll_month.payroll_month}"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    sheet.row_dimensions[1].height = 30

    # Add Subtitle (row 2)
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    subtitle_cell = sheet.cell(row=2, column=1)
    subtitle_cell.value = "Detail of earnings and deductions"
    subtitle_cell.font = Font(size=10, italic=True)
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
    sheet.row_dimensions[2].height = 20

    # Add Headers (row 3)
    export_util = ExportUtilityService()
    sheet.append([export_util.split_header_to_lines(h) for h in header])

    # Header styles
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Amber
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(right=Side(style='thin', color='FF000000'))

    for cell in sheet[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Append payroll data rows starting from row 4
    for payroll in regular_payrolls:
        row = [payroll.id]
        row.extend([
            payroll.personnel_full_name.personnel_id,
            payroll.personnel_full_name.first_name,
            payroll.personnel_full_name.father_name,
            payroll.personnel_full_name.last_name,
        ])

        # Fully taxable fields
        if payroll_month.use_basic_salary:
            row.append(payroll.basic_salary)
        if payroll_month.use_overtime:
            row.append(payroll.overtime)
        if payroll_month.use_housing_allowance:
            row.append(payroll.housing_allowance)
        if payroll_month.use_position_allowance:
            row.append(payroll.position_allowance)
        if payroll_month.use_commission:
            row.append(payroll.commission)
        if payroll_month.use_telephone_allowance:
            row.append(payroll.telephone_allowance)
        if payroll_month.use_one_time_bonus:
            row.append(payroll.one_time_bonus)
        if payroll_month.use_causal_labor_wage:
            row.append(payroll.causal_labor_wage)

        # Partially taxable fields
        if payroll_month.use_transport_home_to_office:
            row.append(payroll.transport_home_to_office_taxable)
            row.append(payroll.transport_home_to_office_non_taxable)
        if payroll_month.use_fuel_home_to_office:
            row.append(payroll.fuel_home_to_office_taxable)
            row.append(payroll.fuel_home_to_office_non_taxable)
        if payroll_month.use_transport_for_work:
            row.append(payroll.transport_for_work_taxable)
            row.append(payroll.transport_for_work_non_taxable)
        if payroll_month.use_fuel_for_work:
            row.append(payroll.fuel_for_work_taxable)
            row.append(payroll.fuel_for_work_non_taxable)
        if payroll_month.use_per_diem:
            row.append(payroll.per_diem_taxable)
            row.append(payroll.per_diem_non_taxable)
        if payroll_month.use_hardship_allowance:
            row.append(payroll.hardship_allowance_taxable)
            row.append(payroll.hardship_allowance_non_taxable)

        # Fully non-taxable fields
        if payroll_month.use_public_cash_award:
            row.append(payroll.public_cash_award)
        if payroll_month.use_incidental_operation_allowance:
            row.append(payroll.incidental_operation_allowance)
        if payroll_month.use_medical_allowance:
            row.append(payroll.medical_allowance)
        if payroll_month.use_cash_gift:
            row.append(payroll.cash_gift)
        if payroll_month.use_tuition_fees:
            row.append(payroll.tuition_fees)
        if payroll_month.use_personal_injury:
            row.append(payroll.personal_injury)
        if payroll_month.use_child_support_payment:
            row.append(payroll.child_support_payment)

        # Deductions
        if payroll_month.use_charitable_donation:
            row.append(payroll.charitable_donation)
        if payroll_month.use_saving_plan:
            row.append(payroll.saving_plan)
        if payroll_month.use_loan_payment:
            row.append(payroll.loan_payment)
        if payroll_month.use_court_order:
            row.append(payroll.court_order)
        if payroll_month.use_workers_association:
            row.append(payroll.workers_association)
        if payroll_month.use_personnel_insurance_saving:
            row.append(payroll.personnel_insurance_saving)
        if payroll_month.use_university_cost_share_pay:
            row.append(payroll.university_cost_share_pay)
        if payroll_month.use_red_cross:
            row.append(payroll.red_cross)
        if payroll_month.use_party_contribution:
            row.append(payroll.party_contribution)
        if payroll_month.use_other_deduction:
            row.append(payroll.other_deduction)

        # Summary fields
        row.extend([
            payroll.employment_income_tax,
            payroll.employee_pension_contribution,
            payroll.employer_pension_contribution,
            payroll.total_pension_contribution,
            payroll.gross_pay,
            payroll.gross_taxable_pay,
            payroll.gross_non_taxable_pay,
            payroll.total_payroll_deduction,
            payroll.net_pay,
            payroll.expense
        ])

        sheet.append(row)

    # Adjust column widths dynamically with limits
    MIN_WIDTH = 10
    MAX_WIDTH = 20
    for i, col_cells in enumerate(sheet.columns, 1):
        max_length = 0
        for cell in col_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = max(MIN_WIDTH, min(MAX_WIDTH, max_length + 2))
        col_letter = get_column_letter(i)
        sheet.column_dimensions[col_letter].width = adjusted_width

    # Save workbook to buffer and return as response
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=payroll_for_{payroll_month.payroll_month.payroll_month}.xlsx'

    return response
