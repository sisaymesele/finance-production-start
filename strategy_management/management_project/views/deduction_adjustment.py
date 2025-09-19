from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
#exel
from django.http import HttpResponse
from openpyxl import Workbook
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from management_project.models import DeductionAdjustment
from management_project.forms import DeductionAdjustmentForm
from management_project.services.deduction_adjustment.context import get_deduction_adjustment_context
from management_project.services.excel_export import ExportUtilityService


@login_required
def deduction_object_list(request):
    context = get_deduction_adjustment_context(request)
    return render(request, 'deduction_adjustment/list.html', context)


@login_required
def deduction_object_detail(request):
    context = get_deduction_adjustment_context(request)
    return render(request, 'deduction_adjustment/detail.html', context)


@login_required
def deduction_per_adjusted_month(request):
    context = get_deduction_adjustment_context(request)
    return render(request, 'deduction_adjustment/per_adjusted_month.html', context)


@login_required
def monthly_deduction_adjustment(request):
    context = get_deduction_adjustment_context(request)
    return render(request, 'deduction_adjustment/monthly_deduction.html', context)

@login_required
def monthly_deduction_adjustment_total(request):
    context = get_deduction_adjustment_context(request)
    return render(request, 'deduction_adjustment/monthly_deduction_total.html', context)



# Create View
@login_required
def create_deduction_adjustment(request):
    if request.method == 'POST':
        form = DeductionAdjustmentForm(request.POST, request=request)
        if form.is_valid():
            deduction_adjustment = form.save(commit=False)
            deduction_adjustment.organization_name = request.user.organization_name
            deduction_adjustment.save()
            messages.success(request, "Deduction Adjustment created successfully!")
            return redirect('deduction_adjustment_list')
        else:
            messages.error(request, "Error creating the deduction adjustment. Check the form.")
    else:
        form = DeductionAdjustmentForm(request=request)

    context = {
        'form': form,
        'form_title': 'Create Deduction Adjustment',
        'submit_button_text': 'Create Deduction Adjustment',
    }
    return render(request, 'deduction_adjustment/form.html', context)


# Update View
@login_required
def update_deduction_adjustment(request, pk):
    deduction_adjustment = get_object_or_404(DeductionAdjustment, pk=pk,
                                             organization_name=request.user.organization_name)

    if request.method == 'POST':
        form = DeductionAdjustmentForm(request.POST, instance=deduction_adjustment, request=request)
        if form.is_valid():
            form.save()
            messages.success(request, "Deduction Adjustment updated successfully!")
            return redirect('deduction_adjustment_list')
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = DeductionAdjustmentForm(instance=deduction_adjustment, request=request)

    context = {
        'form': form,
        'form_title': 'Update Deduction Adjustment',
        'submit_button_text': 'Update Deduction Adjustment',
    }
    return render(request, 'deduction_adjustment/form.html', context)


# Delete View
@login_required
def delete_deduction_adjustment(request, pk):
    deduction_adjustment = get_object_or_404(DeductionAdjustment, pk=pk,
                                             organization_name=request.user.organization_name)

    if request.method == "POST":
        deduction_adjustment.delete()
        messages.success(request, "Deduction Adjustment deleted successfully!")
        return redirect('deduction_adjustment_list')

    context = {'deduction_adjustment': deduction_adjustment}

    return render(request, 'deduction_adjustment/delete_confirm.html', context)


# export deduction adjustment
@login_required
def export_deduction_adjustment_list_to_excel(request):
    context = get_deduction_adjustment_context(request)
    deductions = context.get('deduction_adjustments', [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Deduction Adjustments"

    headers = [
        "Record Month", "Adjusted Month",
        "First Name", "Father Name", "Last Name",
        "Case", "Component", "Deduction Amount",
        "Period Start", "Period End", "Months Covered",
        "Created At", "Updated At"
    ]

    total_columns = len(headers)

    # Title row (1st)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Detailed Deduction Adjustment"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Subtitle row (2nd)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_columns)
    subtitle_cell = ws.cell(row=2, column=1)
    subtitle_cell.value = "Detailed personnel deduction adjustments for each component"
    subtitle_cell.font = Font(size=10, italic=True)
    subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20

    # Headers row (3rd)
    export_util = ExportUtilityService()
    ws.append([export_util.split_header_to_lines(h) for h in headers])

    # Header styling
    header_fill = PatternFill(start_color="FF0070C0", end_color="FF0070C0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(right=Side(style='thin', color='FF000000'))

    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Helper function for safe nested getattr one level
    def safe_getattr(obj, attr, default=""):
        return getattr(obj, attr, default) if obj else default

    # Manually 3-level safe getattr for strategy_by_cycle fields
    for d in deductions:
        # Record Month nested 3 levels
        rec_month_obj = safe_getattr(d.payroll_to_record, 'strategy_by_cycle', None)
        rec_month_lvl2 = safe_getattr(rec_month_obj, 'strategy_by_cycle', None)
        rec_month_final = safe_getattr(rec_month_lvl2, 'strategy_by_cycle', "")

        # Adjusted Month nested 3 levels
        adj_month_obj = safe_getattr(d.payroll_needing_adjustment, 'strategy_by_cycle', None)
        adj_month_lvl2 = safe_getattr(adj_month_obj, 'strategy_by_cycle', None)
        adj_month_final = safe_getattr(adj_month_lvl2, 'strategy_by_cycle', "")

        ws.append([
            rec_month_final,
            adj_month_final,
            safe_getattr(safe_getattr(d.payroll_to_record, 'personnel_full_name', None), 'first_name', ""),
            safe_getattr(safe_getattr(d.payroll_to_record, 'personnel_full_name', None), 'father_name', ""),
            safe_getattr(safe_getattr(d.payroll_to_record, 'personnel_full_name', None), 'last_name', ""),
            d.get_case_display(),
            d.get_component_display(),
            float(d.deduction_amount or 0),
            d.period_start.strftime("%Y-%m-%d") if d.period_start else "",
            d.period_end.strftime("%Y-%m-%d") if d.period_end else "",
            d.months_covered or "",
            d.created_at.strftime("%Y-%m-%d %H:%M") if d.created_at else "",
            d.updated_at.strftime("%Y-%m-%d %H:%M") if d.updated_at else "",
        ])

    # Adjust column widths
    MIN_WIDTH = 10
    MAX_WIDTH = 25
    for idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                continue
        adjusted_width = min(max(max_length + 2, MIN_WIDTH), MAX_WIDTH)
        ws.column_dimensions[get_column_letter(idx)].width = adjusted_width

    # Prepare response
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response['Content-Disposition'] = 'attachment; filename=deduction_adjustments.xlsx'
    return response



@login_required
def export_deduction_per_adjusted_month_to_excel(request):
    context = get_deduction_adjustment_context(request)
    data = context.get('deduction_per_adjusted_month', [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Total Deduction By Adjusted Month"

    headers = [
        "Record Month",
        "Adjusted Month",
        "Personnel ID",
        "First Name",
        "Father Name",
        "Last Name",
        "Total Deduction"
    ]

    total_columns = len(headers)

    # Title row (1st)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Deduction Adjustment By Adjusted Month"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Use ExportUtilityService to split header lines
    export_util = ExportUtilityService()
    ws.append([export_util.split_header_to_lines(h) for h in headers])

    # Header style
    header_fill = PatternFill(start_color="FF0070C0", end_color="FF0070C0", fill_type="solid")  # Blue fill
    header_font = Font(bold=True, color="FFFFFFFF")  # White font
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(right=Side(style='thin', color='FF000000'))

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Append data rows starting from row 4
    for item in data:
        ws.append([
            item.get("payroll_to_record__payroll_month__payroll_month__payroll_month", ""),
            item.get("payroll_needing_adjustment__payroll_month__payroll_month__payroll_month", ""),
            item.get("payroll_to_record__personnel_full_name__personnel_id", ""),
            item.get("payroll_to_record__personnel_full_name__first_name", ""),
            item.get("payroll_to_record__personnel_full_name__father_name", ""),
            item.get("payroll_to_record__personnel_full_name__last_name", ""),
            float(item.get("adjusted_month_total_deduction", 0) or 0),
        ])

    # Auto-adjust column widths
    MIN_WIDTH = 12
    MAX_WIDTH = 15
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max(max_length + 2, MIN_WIDTH), MAX_WIDTH)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Save workbook to memory and return response
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=deduction_per_adjusted_month.xlsx"
    return response



@login_required
def export_monthly_deduction_adjustment_to_excel(request):
    context = get_deduction_adjustment_context(request)
    data = context.get('monthly_deduction_adjustment', [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Deduction Adjustment"

    headers = [
        "Payroll Month",
        "Personnel ID",
        "First Name",
        "Father Name",
        "Last Name",
        "Total Deduction",
    ]

    total_columns = len(headers)

    # Title row (1st)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Monthly Deduction Adjustment"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Use ExportUtilityService to split header lines
    export_util = ExportUtilityService()
    ws.append([export_util.split_header_to_lines(h) for h in headers])

    # Header style
    header_fill = PatternFill(start_color="FF0070C0", end_color="FF0070C0", fill_type="solid")  # Blue fill
    header_font = Font(bold=True, color="FFFFFFFF")  # White font
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin_border = Border(right=Side(style='thin', color='FF000000'))

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Append data rows
    for item in data:
        ws.append([
            item.get("payroll_to_record__payroll_month__payroll_month__payroll_month", ""),
            item.get("payroll_to_record__personnel_full_name__personnel_id", ""),
            item.get("payroll_to_record__personnel_full_name__first_name", ""),
            item.get("payroll_to_record__personnel_full_name__father_name", ""),
            item.get("payroll_to_record__personnel_full_name__last_name", ""),
            float(item.get("recorded_month_total_deduction", 0) or 0),
        ])

    # Adjust column widths
    MIN_WIDTH = 12
    MAX_WIDTH = 15
    for i, col_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in col_cells:
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except Exception:
                pass
        adjusted_width = max(MIN_WIDTH, min(MAX_WIDTH, max_length + 2))
        ws.column_dimensions[get_column_letter(i)].width = adjusted_width

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=monthly_deduction_adjustment.xlsx"
    return response


@login_required
def export_monthly_deduction_adjustment_aggregate(request):
    # Get aggregated data
    context = get_deduction_adjustment_context(request)
    monthly_deduction_adjustment_aggregate = context.get('monthly_deduction_adjustment_aggregate', [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Deduction Aggregate"

    # Title row
    total_columns = 3
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = ws.cell(row=1, column=1)
    title_cell.value = "Monthly Deduction Adjustment Aggregate Report"
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Headers
    headers = [
        "Payroll Month",
        "Total Deduction",
    ]
    ws.append(headers)

    # Header styles
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type='solid')
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(right=Side(style='thin'))

    for cell in ws[2]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row in monthly_deduction_adjustment_aggregate:
        payroll_month = row.get('payroll_to_record__payroll_month__payroll_month__payroll_month', '')
        total_deduction = row.get('total_deduction', 0)
        # Convert Decimal to float
        if total_deduction is None:
            total_deduction = 0
        else:
            total_deduction = float(total_deduction)

        ws.append([payroll_month, total_deduction])

    # Adjust column widths
    MIN_WIDTH = 15
    MAX_WIDTH = 25
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_length = max((len(str(cell.value)) for cell in col_cells if cell.value), default=0)
        adjusted_width = max(MIN_WIDTH, min(MAX_WIDTH, max_length + 2))
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # Return response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=monthly_deduction_aggregate.xlsx'
    return response